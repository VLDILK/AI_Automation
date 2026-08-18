"""Шар даних складу: SQLite-сховище (ExcelSqliteStore), синхронізація
SQLite -> Excel (create_excel_backup, sync_sheet(s)_to_excel), і бізнес-логіка
приходу/продажу (apply_sale_operation, apply_income_operation) разом з їхніми
допоміжними функціями (мапінг колонок складу, форматування розміру позиції,
формування рядка листа продажів тощо).

SQLite тут завжди головне джерело правди; Excel лише генерується з нього.
apply_sale_operation/apply_income_operation приймають sync_mode параметром
(а не читають налаштування самі) — рішення "коли синхронізувати Excel"
лишається за викликачем (TelegramBotWorker), цей модуль про Telegram не знає
нічого.
"""

import base64
import hashlib
import html
import json
import os
import re
import shutil
import sqlite3
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

import excel_source
import permissions
from paths import BACKUP_DIR, BACKUP_PASSWORD_PATH, DB_BACKUP_DIR, SETTINGS_PATH
from settings import SettingsStore
from utils import (
    _deserialize_row,
    _display_bot_number,
    _display_value,
    _normalize_phrase,
    _number_value,
    _parse_date_text,
    _priced_amount,
    _sanitize_excel_value,
    _serialize_row,
    piece_measure as _shared_piece_measure,
    row_measure_kind as _shared_row_measure_kind,
)

# Скільки останніх копій Excel-файлу тримати в backups/ — старіші видаляються
# автоматично після кожного нового бекапу.
EXCEL_BACKUP_LIMIT = 30

# Скільки останніх щоденних знімків app_data.sqlite3 тримати в
# db_backups/auto/ — старіші видаляються автоматично. 20 покриває і "10-20
# знімків", і "історію за тиждень" одним рішенням (~3 тижні, ~59 МБ на диску
# при сьогоднішньому розмірі бази).
DB_BACKUP_LIMIT = 20

# Команди бота "з коробки" — заповнюються в БД при першому запуску
# (ExcelSqliteStore._seed_builtin_bot_commands) і можуть далі редагуватись
# через UI (додавання власних синонімів-alias).
BUILTIN_BOT_COMMANDS = [
    {
        "code": "stock_balance",
        "title": "Остаток",
        "description": "Показать актуальный остаток по листу СКЛАД.",
        "aliases": [
            "остаток",
            "остатки",
            "покажи остаток",
            "показать остаток",
            "остаток склада",
            "склад остаток",
        ],
    },
    {
        "code": "add_income",
        "title": "Приход",
        "description": "Добавить приход материала в склад.",
        "aliases": [
            "приход",
            "прихід",
            "поступление",
            "добавить приход",
            "добавь приход",
        ],
    },
    {
        "code": "stock_sale",
        "title": "Продажа",
        "description": "Списать материал со склада как продажу.",
        "aliases": [
            "продажа",
            "продажи",
            "продать",
            "продай",
            "реализация",
        ],
    },
    {
        "code": "cancel_operation",
        "title": "Отмена",
        "description": "Отменить текущую операцию с ботом.",
        "aliases": [
            "отмена",
            "отменить",
            "стоп",
            "скасувати",
            "відміна",
            "відмінити",
        ],
    },
    {
        "code": "calculator",
        "title": "Калькулятор",
        "description": "Посчитать числа, кубатуру по размерам или количество штук по м3.",
        "aliases": [
            "калькулятор",
            "посчитай",
            "порахуй",
            "рахуй",
            "считай",
            "рассчитай",
            "сколько кубов",
            "сколько штук",
            "порахувати",
        ],
    },
    {
        "code": "help",
        "title": "Помощь",
        "description": "Показать список доступных команд.",
        "aliases": [
            "помощь",
            "help",
            "/help",
            "справка",
            "команды",
            "список команд",
        ],
    },
]

# Каталог дій, які можна прив'язати до кастомної кнопки (Кнопки бота, GUI).
# Кожна дія позначена "section" — кореневий кастомний вузол обирає ОДНУ
# секцію, і дії з інших секцій просто не пропонуються для жодної кнопки
# в цій гілці (щоб під "Реализация" не можна було прив'язати "Приход").
# Суто дані, без залежності від Telegram — фактичний виклик (яку функцію
# TelegramDialogMixin запустити для даного code) живе в telegram_dialog.py.
CUSTOM_BUTTON_ACTIONS = [
    {"code": "start_income", "section": "приход", "label": "Начать приход товара"},
    {"code": "start_income_form", "section": "приход", "label": "Начать приход одной формой"},
    {"code": "start_sale", "section": "реализация", "label": "Начать продажу"},
    {"code": "start_sale_form", "section": "реализация", "label": "Начать продажу одной формой"},
    {"code": "start_antiseptic_form", "section": "реализация", "label": "Начать антисептирование одной формой"},
    {"code": "start_stock_report", "section": "данные", "label": "Показать остаток склада"},
    {"code": "start_sales_report", "section": "данные", "label": "Показать отчёт по продажам"},
    {"code": "start_antiseptic_report", "section": "данные", "label": "Показать отчёт по антисептированию"},
    {"code": "start_sales_by_client_report", "section": "данные", "label": "Показать отчёт по клиентам"},
    {"code": "start_low_stock_report", "section": "данные", "label": "Показать отчёт по низкому остатку"},
    {"code": "start_writeoff", "section": "списание", "label": "Начать списание товара"},
    {"code": "start_writeoff_form", "section": "списание", "label": "Начать списание одной формой"},
    {"code": "start_data_browser_form", "section": "данные", "label": "Показать данные одной формой"},
    {"code": "start_calculator", "section": "прочее", "label": "Открыть калькулятор"},
    {"code": "show_help", "section": "прочее", "label": "Показать справку"},
]

CUSTOM_BUTTON_SECTIONS = ["приход", "реализация", "данные", "списание", "прочее"]

# Мітки, які кастомна кнопка НЕ може перевикористати (колізія з уже
# захардкодженими кнопками головного меню бота). Жодна з 5 колишніх
# кнопок головного меню (приход/реализация/данные/калькулятор/помощь)
# тут більше не потрібна — усі перенесені у custom_menu_buttons
# (BUILTIN_MIGRATED_CUSTOM_BUTTONS нижче) і більше не хардкоджені; дубль
# назви й так відсіється перевіркою на bot_commands-alias
# (find_command_code_by_phrase) у custom_button_label_collides для тих,
# що мають alias ("приход"/"реализация"/"калькулятор"/"помощь" — "данные"
# такого alias-у не має, це прийнятний залишковий пробіл).
# Аудит коду: цей набір був ПОРОЖНІМ, хоча custom_button_label_collides
# (нижче) уже звіряється саме з ним — кастомну кнопку можна було назвати
# "Данные"/"Склад"/"Помощь"/"Назад" тощо (набір проходив перевірку колізії
# з БУДЬ-якою командою), а сама кнопка потім НІКОЛИ не спрацьовувала, бо
# хардкоджені перевірки (_is_data_menu_request/_is_stock_data_menu_request/
# _is_main_menu_back_request/_is_help_request/_is_cancel_request у
# telegram_dialog_core.py/_menu.py) перехоплюють цей текст РАНІШЕ, ніж
# диспетчер доходить до кастомного дерева кнопок. Тут — рівно ті фрази, що
# перехоплюються ДО _custom_root_button_by_label.
_RESERVED_BUTTON_LABELS = {
    "данные", "data",
    "склад", "warehouse",
    "помощь", "help", "справка", "команды", "список команд",
    "назад", "back",
    "главное меню", "головне меню",
    "отмена", "отменить", "стоп", "скасувати", "відміна", "відмінити",
}

# Пункти головного меню, перенесені з хардкоду в custom_menu_buttons —
# Задача користувача: переносити по одному, кожен стає звичайною кнопкою,
# яку видно й можна редагувати в Редакторі кнопок. migration_key —
# ідентифікатор для ОДНОРАЗОВОГО сіяння (перевіряється перед вставкою в
# _seed_builtin_migrated_custom_buttons): після першого запуску це вже
# ЗВИЧАЙНА кнопка, подальші правки адміністратора (назва/текст/розмір/
# видалення) ніколи не перезаписуються повторним стартом застосунку.
# parent_migration_key (None = корінь) — ДАННЫЕ тепер БАТЬКІВСЬКИЙ вузол
# (без власної дії, показує дітей), а СКЛАД/ПРОДАЖИ — його мігровані ДІТИ
# (Задача користувача: "ті що є всередині — вимкни і додай у редактор").
# Записи-батьки мають йти в списку ПЕРЕД своїми дітьми — сіяння резолвить
# parent_id за вже вставленим migration_key батька.
BUILTIN_MIGRATED_CUSTOM_BUTTONS = [
    {"migration_key": "income", "label": "ПРИХОД", "action_code": "start_income", "layout": "full", "parent_migration_key": None},
    {"migration_key": "income_form", "label": "ПРИХОД (форма)", "action_code": "start_income_form", "layout": "full", "parent_migration_key": None},
    {"migration_key": "sale", "label": "РЕАЛИЗАЦИЯ", "action_code": "start_sale", "layout": "full", "parent_migration_key": None},
    {"migration_key": "sale_form", "label": "РЕАЛИЗАЦИЯ (форма)", "action_code": "start_sale_form", "layout": "full", "parent_migration_key": None},
    {"migration_key": "antiseptic_form", "label": "АНТИСЕПТИРОВАНИЕ (форма)", "action_code": "start_antiseptic_form", "layout": "full", "parent_migration_key": None},
    {"migration_key": "writeoff", "label": "СПИСАНИЕ", "action_code": "start_writeoff", "layout": "full", "parent_migration_key": None},
    {"migration_key": "writeoff_form", "label": "СПИСАНИЕ (форма)", "action_code": "start_writeoff_form", "layout": "full", "parent_migration_key": None},
    {"migration_key": "data_browser_form", "label": "ДАННЫЕ (форма)", "action_code": "start_data_browser_form", "layout": "full", "parent_migration_key": None},
    {"migration_key": "data_menu", "label": "ДАННЫЕ", "action_code": None, "layout": "full", "parent_migration_key": None},
    {"migration_key": "stock_report_section", "label": "СКЛАД", "action_code": "start_stock_report", "layout": "full", "parent_migration_key": "data_menu"},
    {"migration_key": "sales_report_section", "label": "ПРОДАЖИ", "action_code": "start_sales_report", "layout": "full", "parent_migration_key": "data_menu"},
    {"migration_key": "antiseptic_report_section", "label": "АНТИСЕПТИРОВАНИЕ", "action_code": "start_antiseptic_report", "layout": "full", "parent_migration_key": "data_menu"},
    {"migration_key": "sales_by_client_report_section", "label": "Клиенты", "action_code": "start_sales_by_client_report", "layout": "full", "parent_migration_key": "data_menu"},
    {"migration_key": "low_stock_report_section", "label": "Низкий остаток", "action_code": "start_low_stock_report", "layout": "full", "parent_migration_key": "data_menu"},
    {"migration_key": "calculator", "label": "Калькулятор", "action_code": "start_calculator", "layout": "half", "parent_migration_key": None},
    {"migration_key": "help", "label": "Помощь", "action_code": "show_help", "layout": "half", "parent_migration_key": None},
]


# Задача користувача: показати й дати редагувати текст, який бот пише в чат
# одразу після натискання кожної дії. Ключ — сам action_code (з
# CUSTOM_BUTTON_ACTIONS) — унікальний і так. Значення тут — саме ТИПОВИЙ
# (заводський) текст: якщо адміністратор ще нічого не міняв,
# store.get_message_template(action_code, BOT_MESSAGE_DEFAULTS[action_code])
# у telegram_dialog.py поверне саме цей рядок — тому він має бути
# СИНХРОНІЗОВАНИЙ з реальним хардкодженим текстом у відповідних функціях
# (_start_income_category_menu/_start_sale_payment_method_menu/
# _stock_data_menu_reply/_sales_period_prompt_reply/
# _start_calculator_operation/show_help-гілка в _custom_button_action_reply).
BOT_MESSAGE_DEFAULTS = {
    "start_income": "Приход. Выберите категорию товара:",
    # Задача користувача: "додай приход (форма)" - той самий мега-формат,
    # що вже мають РЕАЛИЗАЦИЯ (форма)/СПИСАНИЕ (форма).
    "start_income_form": (
        "Приход одной формой. Нажмите кнопку ниже и заполните всё сразу — "
        "категория и размеры."
    ),
    "start_sale": "Реализация. Выберите способ оплаты:",
    # Задача користувача (Крок 4, "Дії" не показує пропущений крок після
    # вибору способу оплати): "start_sale" вище — це текст ДО вибору
    # способу оплати, а цей — ОКРЕМИЙ текст, що бот показує ВЖЕ ПІСЛЯ
    # вибору способу оплати, разом з клавіатурою категорій товару. Раніше
    # був хардкоджений у ДВОХ місцях (_start_sale_category_menu і
    # "choose_sale_payment_method" в telegram_dialog.py) без жодного
    # редагованого представлення — тепер обидва читають цей ключ.
    "start_sale_category_prompt": "Реализация. Выберите категорию товара:",
    # Задача користувача: "розділити" РЕАЛИЗАЦИЯ на 2 кнопки — стара
    # (крок за кроком: спосіб оплати -> категорія -> чек-лист/форма) лишається
    # без змін, ця нова відкриває ОДНУ форму, де категорія (і спосіб оплати)
    # обираються прямо в самій формі, без жодних чатових кроків до неї.
    "start_sale_form": (
        "Продажа одной формой. Нажмите кнопку ниже и заполните всё сразу — "
        "категория, размеры, клиент и оплата."
    ),
    # Задача користувача: окрема форма для антисептирования (раніше було
    # лише доповнення всередині форми продажі) - товар/розмір/штук обираються
    # так само, як і в решті форм, об'єм рахується автоматично.
    "start_antiseptic_form": (
        "Антисептирование одной формой. Нажмите кнопку ниже и заполните всё сразу — "
        "товар, размеры, объём посчитается автоматически, цена, клиент и оплата."
    ),
    "start_stock_report": "Склад. Выберите категорию или Фильтры для точного поиска:",
    "start_sales_report": "Продажи. За какой период показать?",
    "start_antiseptic_report": "Антисептирование. За какой период показать?",
    "start_sales_by_client_report": "Продажи по клиентам. За какой период показать?",
    "start_low_stock_report": "Показываю позиции с низким остатком.",
    "start_writeoff": "Списание товара. Выберите категорию:",
    "start_writeoff_form": (
        "Списание одной формой. Нажмите кнопку ниже и заполните всё сразу — "
        "категория и размеры."
    ),
    "start_data_browser_form": "Данные склада одной формой.",
    "start_calculator": "Что посчитать?",
    "show_help": (
        "Доступные команды:\n"
        "Приход - принять товар на склад\n"
        "Реализация - оформить продажу\n"
        "Данные - склад по категориям/фильтру, продажи\n"
        "Калькулятор - посчитать числа, м3 или количество штук\n"
        "Отмена - отменить текущую операцию\n"
        "Помощь - показать все доступные команды"
    ),
}


# Крок 3+ "Дії": одноразовий сідінг bot_operations/bot_operation_fields/
# bot_operation_field_columns (той самий builtin_key-ідіом, що й
# BUILTIN_MIGRATED_CUSTOM_BUTTONS вище) — точно відтворює сьогоднішню
# поведінку apply_income_operation/apply_sale_operation ДО того, як
# адміністратор щось редагував. Ідентифікаційні поля складу (товар/порода/
# товщина/ширина/довжина) — ті самі для приходу й продажу; "condition_
# identity" відрізняє лише те, чи товщина типу (AD/KD) обов'язкова для
# пошуку рядка (продукти без типу — ОСБ/Вагонка — порівнюють тип, лише
# якщо він заданий, див. _warehouse_row_matches).
BUILTIN_OPERATIONS = [
    {"builtin_key": "income_doska_ad", "kind": "income", "parent_action_code": "start_income",
     "label": "ДОСКА AD", "prefill": {"product": "Доска", "condition": "AD"}, "condition_identity": True},
    {"builtin_key": "income_doska_kd", "kind": "income", "parent_action_code": "start_income",
     "label": "ДОСКА KD", "prefill": {"product": "Доска", "condition": "KD"}, "condition_identity": True},
    {"builtin_key": "income_osb", "kind": "income", "parent_action_code": "start_income",
     "label": "ОСБ", "prefill": {"product": "ОСБ"}, "condition_identity": False},
    {"builtin_key": "income_vagonka", "kind": "income", "parent_action_code": "start_income",
     "label": "ВАГОНКА", "prefill": {"product": "Вагонка"}, "condition_identity": False},
    {"builtin_key": "sale_doska_ad", "kind": "sale", "parent_action_code": "start_sale",
     "label": "ДОСКА AD", "prefill": {"product": "Доска", "condition": "AD"}, "condition_identity": True},
    {"builtin_key": "sale_doska_kd", "kind": "sale", "parent_action_code": "start_sale",
     "label": "ДОСКА KD", "prefill": {"product": "Доска", "condition": "KD"}, "condition_identity": True},
    {"builtin_key": "sale_osb", "kind": "sale", "parent_action_code": "start_sale",
     "label": "ОСБ", "prefill": {"product": "ОСБ"}, "condition_identity": False},
    {"builtin_key": "sale_vagonka", "kind": "sale", "parent_action_code": "start_sale",
     "label": "ВАГОНКА", "prefill": {"product": "Вагонка"}, "condition_identity": False},
    {"builtin_key": "writeoff_doska_ad", "kind": "writeoff", "parent_action_code": "start_writeoff",
     "label": "ДОСКА AD", "prefill": {"product": "Доска", "condition": "AD"}, "condition_identity": True},
    {"builtin_key": "writeoff_doska_kd", "kind": "writeoff", "parent_action_code": "start_writeoff",
     "label": "ДОСКА KD", "prefill": {"product": "Доска", "condition": "KD"}, "condition_identity": True},
    {"builtin_key": "writeoff_osb", "kind": "writeoff", "parent_action_code": "start_writeoff",
     "label": "ОСБ", "prefill": {"product": "ОСБ"}, "condition_identity": False},
    {"builtin_key": "writeoff_vagonka", "kind": "writeoff", "parent_action_code": "start_writeoff",
     "label": "ВАГОНКА", "prefill": {"product": "Вагонка"}, "condition_identity": False},
]

# АНТИСЕПТИРОВАНИЕ — окрема "service"-дія (requires_row_identity=0: не
# шукає й не чіпає жодного рядка складу, лист свій). Усі поля тут
# write_mode='ledger' (реальний запис і далі робить antiseptic_sheet_values,
# бо там бізнес-логіка — розподіл готівка/банк, нумерація документа тощо).
BUILTIN_SERVICE_OPERATIONS = [
    {"builtin_key": "sale_antiseptic", "kind": "service", "parent_action_code": "start_sale",
     "label": "АНТИСЕПТИРОВАНИЕ"},
]

# Мітки — РОСІЙСЬКОЮ (Задача користувача, знайдено при підключенні чек-
# листа до конфігурації: бот сам спілкується виключно російською — те, що
# зберігається тут, ЦЕ Й Є буквальний текст пункту в чат-повідомленні, той
# самий принцип, що й bot_message_templates.text для "Редагувати текст").
# Українською лишається тільки обгортка навколо цього поля в самому GUI
# (підписи кнопок/заголовків екрана), не сам текст, що йде в чат.
_WAREHOUSE_IDENTITY_FIELD_LABELS = [
    ("product", "Товар"),
    ("breed", "Порода"),
    ("condition", "Тип продукта"),
    ("thickness", "Толщина"),
    ("width", "Ширина"),
    ("length", "Длина"),
]

_ANTISEPTIC_FIELD_DEFS = [
    ("date", "Дата", "info"),
    ("service_number", "№ услуги", "info"),
    ("client", "Клиент", "info"),
    ("address", "Адрес выгрузки", "info"),
    ("service", "Услуга", "info"),
    ("unit", "Ед. изм.", "info"),
    # volume/payment_method — мітки саме такі, як у чек-листі бота
    # (_antiseptic_mandatory_fields_missing), НЕ як сирий заголовок колонки
    # аркуша (той показується окремо, у власному рядку прив'язки).
    ("volume", "Объем услуги, м3", "add"),
    ("price_per_unit", "Цена за м3", "info"),
    ("total_amount", "Стоимость", "add"),
    ("payment_method", "Способ оплаты", "info"),
    ("payment_status", "Статус оплаты", "info"),
    ("document", "№ документа", "info"),
    ("cash_amount", "Приход наличных", "add"),
    ("bank_amount", "Приход по банку", "add"),
    ("reflection", "Отражение в расчетах", "info"),
    ("manager", "Ответственный", "info"),
    ("comment", "Комментарий", "info"),
]

# Крок 3+ "Дії", Етап 3: ЗАКРИТИЙ (не довільний) перелік семантичних
# ключів, які адміністратор може обрати, додаючи НОВЕ поле-запит до вже
# існуючої дії. Це саме ті ключі, які бот УЖЕ вміє розпізнавати з тексту
# (product/breed/condition/thickness/width/length/quantity/measure з
# приходу-продажу, client/price_per_unit/total_amount/payment_method зі
# продажу, решта — з антисептирования) — розпізнавання тексту тут НЕ
# змінюється (Задача користувача, ще Крок 3+ старт), тож новий ключ
# вигадати не можна, лише додати вже наявний до дії, де його ще нема.
RECOGNIZED_OPERATION_FIELD_KEYS = [
    ("product", "Товар"),
    ("breed", "Порода"),
    ("condition", "Тип продукта"),
    ("thickness", "Толщина"),
    ("width", "Ширина"),
    ("length", "Длина"),
    ("quantity", "Количество, шт"),
    ("measure", "Количество, м3"),
    ("client", "Клиент"),
    ("address", "Адрес выгрузки"),
    ("price_per_unit", "Цена"),
    ("total_amount", "Сумма"),
    ("payment_method", "Способ оплаты"),
    ("date", "Дата"),
    ("service_number", "№ услуги"),
    ("service", "Услуга"),
    ("unit", "Ед. изм."),
    ("payment_status", "Статус оплаты"),
    ("document", "№ документа"),
    ("cash_amount", "Приход наличных"),
    ("bank_amount", "Приход по банку"),
    ("reflection", "Отражение в расчетах"),
    ("manager", "Ответственный"),
    ("comment", "Комментарий"),
    # Задача користувача: звіти (stock_report/sales_report) тепер теж
    # реально підключені — ці ключі відповідають РЕАЛЬНИМ колонкам звіту
    # (_STOCK_REPORT_COLUMN_META/_SALES_REPORT_COLUMN_META), потрібні тут,
    # щоб адмін міг ДОДАТИ назад видалену колонку звіту через "+ Додати
    # поле-запит".
    ("size", "Размер (товщина x ширина x длина)"),
    ("position", "Товар (позиция целиком)"),
    ("volume", "Объем, м3"),
    ("area", "Площадь, м2"),
    ("linear", "Длина, мп"),
    ("note", "Примечание"),
]

# Задача користувача: "Показать остаток склада"/"Показать отчёт по
# продажам" теж мають бути редаговані так само, як категорії приходу/
# продажу — додавати/прибирати внутрішні дії. Це ЗВІТИ (не дії, що
# пишуть у таблицю): "показує ВСІ рядки", а не "шукає ОДИН потрібний",
# тож requires_row_identity/is_identity тут не мають сенсу — жодне поле
# не захищене, все вільно додається/прибирається. Усі прив'язки
# write_mode='ledger' (лише показ у редакторі — сам текст звіту й далі
# формує telegram_dialog.py власним кодом, ця конфігурація поки НЕ
# підключена до генерації самого звіту, той самий "поки лише каркас",
# що й ledger-прив'язки в продажу/антисептируванні з Етапу 1).
# Задача користувача: перше редагування звітів (нижче) було ЧИСТО
# декоративним — _sales_report_spec/_stock_report_spec (telegram_dialog.py)
# лишались повністю захардкодженими списками колонок, ніколи не читаючи цю
# таблицю, тож перейменування/видалення поля-запиту тут не міняло НІЧОГО у
# справжньому звіті бота ("бутафорія"). Тепер поля-запити відповідають
# 1-в-1 реальним колонкам звіту (той самий field_key, що й "key" у
# _STOCK_REPORT_COLUMN_META/_SALES_REPORT_COLUMN_META), а мітки — це
# буквально заголовки, які бот друкує в реальному звіті. column_key (третій
# елемент кортежу) — семантичний ключ прив'язки-документації через
# warehouse_columns/sales_columns (як і в приході/продажу); None там, де
# значення КОМПОНУЄТЬСЯ з кількох колонок одразу (size/position/note) і
# немає ОДНІЄЇ реальної колонки, яку можна показати.
BUILTIN_REPORT_OPERATIONS = [
    {
        "builtin_key": "stock_report",
        "parent_action_code": "start_stock_report",
        "label": "Остаток склада",
        "sheet": "СКЛАД",
        "fields": [
            ("product", "Продукт", "product"),
            ("breed", "Порода", "breed"),
            ("condition", "Состояние", "condition"),
            ("size", "Размер", None),
            ("quantity", "Кол-во, шт", "balance_qty"),
            ("volume", "Кол-во, м3", "balance_volume"),
            ("area", "Площадь, м2", "balance_area"),
            ("linear", "Длина, мп", "balance_linear"),
            ("note", "Примечание", None),
        ],
    },
    {
        "builtin_key": "sales_report",
        "parent_action_code": "start_sales_report",
        "label": "Отчёт по продажам",
        "sheet": "ПРОДАЖА МАТЕРИАЛА",
        "fields": [
            ("date", "Дата", "date"),
            ("client", "Клиент", "client"),
            ("address", "Адрес выгрузки", "address"),
            ("position", "Товар", None),
            ("quantity", "Кол-во, шт", "quantity"),
            ("volume", "Объем, м3", "total_volume"),
            ("area", "Площадь, м2", "total_area"),
            ("linear", "Длина, мп", "total_linear"),
            ("total_amount", "Сумма, MDL", "total_amount"),
            ("payment_method", "Оплата", "payment_method"),
            ("manager", "Менеджер", "manager_final"),
        ],
    },
    {
        "builtin_key": "antiseptic_report",
        "parent_action_code": "start_antiseptic_report",
        "label": "Отчёт по антисептированию",
        "sheet": "АНТИСЕПТИРОВАНИЕ",
        "fields": [
            ("date", "Дата", "date"),
            ("client", "Клиент", "client"),
            ("address", "Адрес выгрузки", "address"),
            ("volume", "Объем, м3", "volume"),
            ("total_amount", "Сумма, MDL", "total_amount"),
            ("payment_method", "Оплата", "payment_method"),
            ("manager", "Менеджер", "manager"),
        ],
    },
    {
        "builtin_key": "sales_by_client_report",
        "parent_action_code": "start_sales_by_client_report",
        "label": "Отчёт по клиентам",
        "sheet": "ПРОДАЖА МАТЕРИАЛА",
        "fields": [
            ("client", "Клиент", "client"),
            ("count", "Кол-во продаж", None),
            ("quantity", "Кол-во, шт", "quantity"),
            ("volume", "Объем, м3", "total_volume"),
            ("area", "Площадь, м2", "total_area"),
            ("linear", "Длина, мп", "total_linear"),
            ("total_amount", "Сумма, MDL", "total_amount"),
        ],
    },
    {
        "builtin_key": "low_stock_report",
        "parent_action_code": "start_low_stock_report",
        "label": "Низкий остаток",
        "sheet": "СКЛАД",
        "fields": [
            ("product", "Продукт", "product"),
            ("breed", "Порода", "breed"),
            ("condition", "Состояние", "condition"),
            ("size", "Размер", None),
            ("quantity", "Остаток, шт", "balance_qty"),
            ("volume", "Остаток, м3", "balance_volume"),
            ("area", "Остаток, м2", "balance_area"),
        ],
    },
]


def create_excel_backup():
    BACKUP_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    stem, suffix = excel_source.backup_file_name_parts()
    backup_path = BACKUP_DIR / f"{stem}_{timestamp}{suffix}"
    backup_path.write_bytes(excel_source.backup_workbook_bytes())
    _rotate_excel_backups(stem, suffix)
    # Задача користувача (2026-08-18): "дзеркалити Excel-бекапи в OneDrive
    # так само, як БД" - client_app.py._schedule_db_backup_tick мірорить
    # цей шлях у AI_Automation_Backups/excel_backups щогодини, поруч з
    # db_backups/config_backups.
    return backup_path


def _rotate_excel_backups(stem, suffix, limit=EXCEL_BACKUP_LIMIT):
    backups = sorted(
        BACKUP_DIR.glob(f"{stem}_*{suffix}"),
        key=lambda path: path.stat().st_mtime,
    )
    excess = len(backups) - limit
    if excess <= 0:
        return
    for path in backups[:excess]:
        path.unlink(missing_ok=True)


def sync_sheets_to_excel(store, sheet_names):
    create_excel_backup()
    workbook = excel_source.open_workbook()
    try:
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row - 1)
            for values in store.fetch_all_rows(sheet_name):
                worksheet.append([_sanitize_excel_value(value) for value in values])
        excel_source.save_workbook(workbook)
    finally:
        workbook.close()


def sync_sheet_to_excel(store, sheet_name):
    sync_sheets_to_excel(store, [sheet_name])


# Задача користувача: "додай автоматичну перевірку при приєднанні нової
# таблиці чи є ця вкладка, якщо нема, то програма має створити сама і
# почати заповнювати" - викликається з gui.py на кожному старті (єдина
# точка, де взагалі перечитується джерело Excel - зміна джерела через
# "Таблиця Excel" й так вимагає перезапуску, щоб набути чинності). БЕЗ
# data_only=True (на відміну від читання-для-імпорту нижче) - інакше
# збереження назад тихо стерло б формули на ІНШИХ листах, замінивши їх
# кешованими значеннями (той самий принцип, що вже застосований у
# sync_sheets_to_excel/create_excel_backup - save_workbook лише на
# workbook, відкритому БЕЗ data_only).
# Той самий принцип, що вже має ExcelSqliteStore.import_sheet
# (_SHEETS_WITH_HEADER_BLOCK) - АНТИСЕПТИРОВАНИЕ має інформаційний блок
# (назва/зведення KPI/примітка) ПЕРЕД реальним рядком заголовків, тож
# перевіряти/дописувати колонку можна лише знайшовши РЕАЛЬНИЙ заголовок
# (перша клітинка = "Дата"), а не сліпо вважати рядок 1 заголовком.
#
# Реальний ризик (аудит коду, 2026-08-14): раніше поруч у файлі жила ДРУГА,
# незалежно написана копія цього самого пошуку (_antiseptic_header_row,
# sync_antiseptic_to_excel) - з ІНШИМ діапазоном сканування (обмежена 20
# рядками замість повного листа) і ІНШИМ фолбеком (7 замість 1), якщо
# "Дата" раптом не знайдена. Дві копії пошуку ОДНОГО й того самого рядка
# могли б повернути різні відповіді при зміні структури листа. Тепер один
# спільний пошук, fallback - параметр (кожен викликач лишає свою вже
# перевірену поведінку "якщо не знайдено": generic-виклик нижче безпечно
# трактує рядок 1 як заголовок; sync_antiseptic_to_excel робить
# header_row-4/-3 для зведення зверху листа, тож потребує іншого,
# антисептик-специфічного фолбеку - 7).
def _find_header_row(worksheet, sheet_name, fallback=1):
    if sheet_name != "АНТИСЕПТИРОВАНИЕ":
        return fallback
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == "Дата":
            return row
    return fallback


def ensure_workbook_has_required_sheets():
    workbook = excel_source.open_workbook()
    changed = False
    try:
        # Задача користувача: "якщо приєднати порожній ексель, то має
        # створитись программою красивенька табличка з відповідними
        # вкладками" - кожен із 5 відомих листів, якого взагалі немає у
        # файлі, створюється з нуля з повним набором заголовків.
        for sheet_name, full_headers in _REQUIRED_SHEETS_FULL:
            if sheet_name in workbook.sheetnames:
                continue
            if not changed:
                create_excel_backup()
            worksheet = workbook.create_sheet(sheet_name)
            for column_index, header in enumerate(full_headers, start=1):
                worksheet.cell(row=1, column=column_index, value=header)
            changed = True

        # Той самий принцип самозцілення, тепер для колонки, доданої в лист,
        # що вже існує у реальному файлі (_WRITEOFF_TIME_HEADER) - без цього
        # користувачам зі старим файлом довелось би дописувати колонку
        # вручну.
        if WRITEOFF_SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[WRITEOFF_SHEET_NAME]
            existing_headers = [
                worksheet.cell(row=1, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            if _WRITEOFF_TIME_HEADER not in existing_headers:
                if not changed:
                    create_excel_backup()
                worksheet.insert_cols(1)
                worksheet.cell(row=1, column=1, value=_WRITEOFF_TIME_HEADER)
                changed = True

        for sheet_name, required_header in _REQUIRED_OPERATION_AUTHOR_COLUMNS:
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            header_row = _find_header_row(worksheet, sheet_name)
            existing_headers = [
                worksheet.cell(row=header_row, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            if required_header in existing_headers:
                continue
            if not changed:
                create_excel_backup()
            worksheet.cell(row=header_row, column=worksheet.max_column + 1, value=required_header)
            changed = True

        if changed:
            excel_source.save_workbook(workbook)
        return changed
    finally:
        workbook.close()


# --- "Вирівняти таблицю": стандартизоване візуальне форматування Excel ---
# Задача користувача (2026-08-14): "давай налаштуємо малювання таблиці...
# буде мо всі таблиці самі вирівнювати рівно всі рядки та стовпці, але
# тільки по натисканню на кнопку в налаштуваннях - вирівняти таблицю... а
# потім це ще передамо в налаштування для корегування, зроби мені
# інструменти керування" - параметри (колір/розмір шрифту/ширина стовпців)
# НЕ хардкод у тілі функції нижче, а окремі значення з settings.json
# (TABLE_FORMAT_*_KEY), з дефолтами, що відповідають обраному в чаті
# "Формат 5" (рамка на кожній клітинці, суцільна заливка заголовка) + синій
# #0C447C - щоб майбутній екран Налаштувань міг їх редагувати без зміни
# самої логіки нижче.
TABLE_FORMAT_COLOR_KEY = "table_format_color"
TABLE_FORMAT_FONT_SIZE_KEY = "table_format_font_size"
# Задача користувача (2026-08-14): "немає ще досі... окремої строки" ->
# "не окремої строки, а її налаштування, маю на увазі" - не новий елемент
# інтерфейсу, а МОЖЛИВІСТЬ окремо налаштувати шрифт САМЕ рядка заголовків,
# не обовʼязково той самий розмір, що й у звичайних рядків даних.
# Порожнє значення (типово) = "той самий розмір, що й у TABLE_FORMAT_
# FONT_SIZE_KEY" - без цього перемикача людині, якій узагалі байдуже до
# заголовка окремо, нічого міняти не треба.
TABLE_FORMAT_HEADER_FONT_SIZE_KEY = "table_format_header_font_size"
TABLE_FORMAT_COLUMN_WIDTH_MODE_KEY = "table_format_column_width_mode"
TABLE_FORMAT_COLUMN_WIDTH_KEY = "table_format_column_width"
# Задача користувача (2026-08-14): "мені потрібно не це, мені потрібно
# мати змогу розширювати сам рядок без збільшування тексту в пишину і в
# висоту окремо ОКРЕМО" - розмір шрифту заголовка (вище) НЕ дає більшого
# рядка без більших літер. Ширина - вже окрема від шрифту (колонки,
# TABLE_FORMAT_COLUMN_WIDTH_KEY). Тут - друга, окрема вісь: справжня
# ВИСОТА рядка заголовків (Excel row_dimensions.height, у пунктах), теж
# незалежна від розміру шрифту. Порожньо (типово) = Excel сам підбирає
# висоту під шрифт, як і раніше.
TABLE_FORMAT_HEADER_ROW_HEIGHT_KEY = "table_format_header_row_height"

TABLE_FORMAT_DEFAULT_COLOR = "0C447C"
TABLE_FORMAT_DEFAULT_FONT_SIZE = 12
TABLE_FORMAT_DEFAULT_COLUMN_WIDTH_MODE = "auto"
TABLE_FORMAT_DEFAULT_COLUMN_WIDTH = 16
# Задача користувача (2026-08-15): "немає фільтрів... поверни" - жодних
# структурних втрат насправді нема (Table.ref/headerRowCount/autoFilter
# у живому файлі байт-в-байт ті самі, що й у здоровому оригіналі -
# перевірено прямою інспекцією), але кілька колонок мали авто-ширину
# рівно на межі старого мінімуму (8) - Excel НЕ малює стрілочку-фільтр
# у заголовку, якщо колонці не вистачає місця під саму кнопку поверх
# тексту (добре відомий, задокументований ефект Excel). Мінімум і
# відступ підняті, щоб кнопці завжди було місце.
TABLE_FORMAT_MIN_AUTO_COLUMN_WIDTH = 12
TABLE_FORMAT_MAX_AUTO_COLUMN_WIDTH = 60
TABLE_FORMAT_COLUMN_WIDTH_PADDING = 4


# Форматуються лише 5 листів, якими сам застосунок реально керує (той самий
# перелік, що й ensure_workbook_has_required_sheets створює/підтримує) - НЕ
# будь-які інші листи файлу (аналітика/КАССА тощо, якщо такі є в реальному
# файлі користувача). Їхня структура застосунку невідома (можливо, зведення,
# формули, злиті клітинки) - сліпе форматування могло б їх зіпсувати,
# всупереч самій обіцянці "нічого не буде видалено".
def apply_standard_table_format(settings=None):
    settings = settings or SettingsStore(SETTINGS_PATH)
    color = settings.get(TABLE_FORMAT_COLOR_KEY) or TABLE_FORMAT_DEFAULT_COLOR
    font_size = settings.get(TABLE_FORMAT_FONT_SIZE_KEY) or TABLE_FORMAT_DEFAULT_FONT_SIZE
    header_font_size = settings.get(TABLE_FORMAT_HEADER_FONT_SIZE_KEY) or font_size
    width_mode = settings.get(TABLE_FORMAT_COLUMN_WIDTH_MODE_KEY) or TABLE_FORMAT_DEFAULT_COLUMN_WIDTH_MODE
    fixed_width = settings.get(TABLE_FORMAT_COLUMN_WIDTH_KEY) or TABLE_FORMAT_DEFAULT_COLUMN_WIDTH
    header_row_height = settings.get(TABLE_FORMAT_HEADER_ROW_HEIGHT_KEY) or None

    workbook = excel_source.open_workbook()
    try:
        create_excel_backup()
        for sheet_name, _headers in _REQUIRED_SHEETS_FULL:
            if sheet_name not in workbook.sheetnames:
                continue
            _apply_table_format_to_worksheet(
                workbook[sheet_name], sheet_name, color, font_size, header_font_size,
                width_mode, fixed_width, header_row_height,
            )
        excel_source.save_workbook(workbook)
    finally:
        workbook.close()


def _apply_table_format_to_worksheet(
    worksheet, sheet_name, color, font_size, header_font_size, width_mode, fixed_width, header_row_height=None,
):
    # "зняті всі активні фільтри" - і сам діапазон фільтра (стрілочки
    # заголовка), і рядки, які фільтр міг лишити прихованими - інакше
    # людина бачила б порожні дірки в таблиці без жодного видимого способу
    # їх повернути. Це не суперечить наступній задачі користувача ("у
    # верхніх заголовках мають бути увімкнені фільтри за замовчуванням") -
    # тут прибирається лише СТАН (активний критерій + приховані рядки),
    # сам інструмент фільтра вмикається заново, чистим, наприкінці функції.
    worksheet.auto_filter.ref = None
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].hidden = False

    header_row = _find_header_row(worksheet, sheet_name)
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row < header_row or max_col < 1:
        return

    # Задача користувача (2026-08-14): "зроби для всіх вкладок товсті межі
    # за замовчуванням" -> (2026-08-15) "ніби очі вилазять з голови...
    # внутрішні жирні лінії. зроби їх тоншими. зовнішні залиш" - товста
    # рамка лишається лише по зовнішньому периметру всієї таблиці, а
    # внутрішні лінії між клітинками - тонкі. Border обчислюється ОКРЕМО
    # для кожної клітинки залежно від позиції (перший/останній рядок чи
    # стовпець - товсто, решта - тонко), тож і не один спільний об'єкт,
    # як раніше.
    thick_side = Side(style="thick", color=color)
    thin_side = Side(style="thin", color=color)
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(size=header_font_size, color="FFFFFF", bold=True)
    data_font = Font(size=font_size)
    # Задача користувача (2026-08-14): "текст в заголовках має бути
    # завжди по центру" -> "за замовчуванням у вирівнюванні також у всіх
    # клітинках текст по центру" - тепер ОБИДВА (не лише заголовок) по
    # центру й горизонтально, і вертикально; "завжди"/"за замовчуванням" -
    # фіксовано, не перемикач у Налаштуваннях, як і інші осі вище.
    # "у всіх квадратиках має бути увімкнено 'переносити текст'" -
    # wrap_text=True на КОЖНІЙ клітинці, і заголовку, і даних.
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=header_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = Border(
                left=thick_side if cell.column == 1 else thin_side,
                right=thick_side if cell.column == max_col else thin_side,
                top=thick_side if cell.row == header_row else thin_side,
                bottom=thick_side if cell.row == max_row else thin_side,
            )
            if cell.row == header_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                cell.font = data_font
                cell.alignment = data_alignment

    # Реальний привід (2026-08-15, живий продакшн-файл користувача): "де
    # відразу були рядки ніби сховані - так і не розгорнуло. я із-за цієї
    # вкладки це все і робив" - НЕ hidden=True (те вже знімається вище),
    # а явна висота рядка ~5.85pt (типова - 15pt) - рядки не приховані,
    # а фізично стиснуті майже до нуля, тож виглядають як зниклі. Скидання
    # height=None повертає Excel-у право самому підібрати висоту під
    # вміст+wrap_text (уже увімкнений вище) - природний, "розгорнутий"
    # вигляд, а не довільна фіксована цифра.
    #
    # НЕ обмежено діапазоном header_row..max_row: worksheet.max_row - це
    # межа РЕАЛЬНОГО вмісту (39 у СКЛАД), а вбудована Excel-Таблиця могла
    # мати номінальний діапазон значно ширший (StockTable - A1:U101) -
    # порожні рядки в цьому "хвості" таблиці так само лишались стиснутими
    # й так само виглядали "зниклими" при прокрутці листа. Тому тут -
    # усі рядки з worksheet.row_dimensions на або після header_row,
    # незалежно від max_row (рядки ВИЩЕ header_row, напр. інфоблок
    # АНТИСЕПТИРОВАНИЕ, свідомо не чіпаються, як і скрізь інде тут).
    for row_index in list(worksheet.row_dimensions.keys()):
        if row_index >= header_row:
            worksheet.row_dimensions[row_index].height = None

    # Задача користувача (2026-08-15): "маштаб у всіх вкладках стандарт
    # 85%" - фіксований дефолт (як товсті зовнішні межі/центрування/
    # перенос тексту вище), не окремий TABLE_FORMAT_*_KEY.
    worksheet.sheet_view.zoomScale = 85

    for column_index in range(1, max_col + 1):
        letter = get_column_letter(column_index)
        if width_mode == "fixed":
            worksheet.column_dimensions[letter].width = fixed_width
            continue
        longest = 0
        for row_index in range(header_row, max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None:
                longest = max(longest, len(str(value)))
        worksheet.column_dimensions[letter].width = min(
            TABLE_FORMAT_MAX_AUTO_COLUMN_WIDTH,
            max(TABLE_FORMAT_MIN_AUTO_COLUMN_WIDTH, longest + TABLE_FORMAT_COLUMN_WIDTH_PADDING),
        )

    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1).coordinate

    if header_row_height:
        worksheet.row_dimensions[header_row].height = header_row_height

    # Реальний баг (2026-08-14, живий продакшн-файл користувача): якщо
    # лист має вбудовану Excel-"Таблицю" (ListObject, worksheet.tables -
    # СКЛАД/ПРИХОД/ПРОДАЖА/АНТИСЕПТИРОВАНИЕ мали такі; лише СПИСАНИЕ ні),
    # фільтр-стрілочки заголовка МАЄ давати САМА таблиця (її власний
    # внутрішній autoFilter, окремий XML-вузол /xl/tables/tableN.xml) -
    # worksheet.auto_filter.ref (рівень ЛИСТА) у такому листі має
    # ЗАЛИШАТИСЬ None. Встановлення сюди ще й рівня листа поверх уже
    # наявної Таблиці створює несумісний файл - Excel показав "виявлено
    # проблему із вмістом" і при "відновленні" видалив усі 4 Таблиці
    # цілком. Тому нижче торкаємось auto_filter.ref ЛИШЕ для листів БЕЗ
    # власної Таблиці.
    if not worksheet.tables:
        worksheet.auto_filter.ref = (
            f"{get_column_letter(1)}{header_row}:{get_column_letter(max_col)}{max_row}"
        )
    else:
        # Задача користувача (2026-08-16): "немає фільтрів при
        # вирівнюванні таблиці. верхній рядок їх не отримує" - реальна
        # перевірка живого test_sklad.xlsx показала: table.autoFilter
        # (ВЛАСНИЙ, не рівня листа) був None на КОЖНІЙ вбудованій
        # Таблиці - без цього вузла Excel не малює стрілочки фільтра в
        # заголовку, попри те, що сама Таблиця відформатована. Це саме
        # той механізм, що openpyxl сам застосовує при СТВОРЕННІ нової
        # Таблиці (Table._initialise_columns) - тут відтворюємо його ж
        # для вже наявних Таблиць. НЕ рівень листа (worksheet.auto_filter,
        # той самий, що ламав файл вище) - лише autoFilter самої Таблиці,
        # ref = той самий діапазон, що вже має сама Таблиця.
        for table in worksheet.tables.values():
            table.autoFilter = AutoFilter(ref=table.ref)


# --- Захист app_data.sqlite3: щоденні знімки + безпечне відновлення ---
# SQLite тут — справжні дані (Excel лише регенерується з нього, див.
# docstring модуля), тож саме цей файл і треба захищати знімками, а не
# create_excel_backup вище (той захищає лише мірор для перегляду).


# Задача користувача (2026-08-14): "може паролем краще захистити?" - знімки
# лишаються ЛОКАЛЬНИМИ (та сама папка), але шифруються Fernet-ом (пакет
# cryptography, уже встановлений), якщо заданий пароль. Ключ = sha256(пароль)
# без окремого файлу-солі навмисно: солі ніде більше не зберігаємо (той
# самий пароль завжди дає той самий ключ) - інакше загублений файл солі
# назавжди зробив би ВСІ вже зроблені знімки нечитабельними. Порожній
# пароль = стара поведінка (звичайні .sqlite3 знімки), нічого не змінюється
# для тих, хто його не задав.
#
# Реальний ризик (аудит коду, 2026-08-14): пароль раніше зберігався ПРЯМО в
# settings.json - той самий файл, доступ до якого вже й так дає доступ до
# папки зашифрованих знімків поруч, тож шифрування мало сенс лише проти
# "хтось скопіював ЛИШЕ папку бекапів". Тепер - окремий файл
# (paths.BACKUP_PASSWORD_PATH), той самий принцип, що й telegram_token_file.
# Одноразова міграція нижче переносить уже задане значення зі старого ключа
# settings.json, не ламаючи вже зроблені зашифровані знімки.
def _backup_encryption_password():
    if BACKUP_PASSWORD_PATH.exists():
        try:
            return BACKUP_PASSWORD_PATH.read_text(encoding="utf-8").strip()
        except OSError:
            return ""
    settings = SettingsStore(SETTINGS_PATH)
    legacy_password = (settings.get("backup_encryption_password") or "").strip()
    if legacy_password:
        _set_backup_encryption_password(legacy_password)
        settings.set("backup_encryption_password", "")
    return legacy_password


def _set_backup_encryption_password(password):
    BACKUP_PASSWORD_PATH.parent.mkdir(parents=True, exist_ok=True)
    password = (password or "").strip()
    if not password:
        BACKUP_PASSWORD_PATH.unlink(missing_ok=True)
        return
    tmp_path = BACKUP_PASSWORD_PATH.with_name(BACKUP_PASSWORD_PATH.name + ".tmp")
    tmp_path.write_text(password, encoding="utf-8")
    os.replace(tmp_path, BACKUP_PASSWORD_PATH)


# Реальна знахідка (аудит коду, 2026-08-16): ключ раніше рахувався ОДНИМ
# проходом sha256(пароль) - жодного "розтягування" (key stretching), тож
# хтось із зашифрованим .enc-знімком міг перебирати паролі на повній
# швидкості sha256 (мільярди/сек на GPU). Сіль тут навмисно ФІКСОВАНА
# константа коду (не випадкова й не в окремому файлі) - той самий принцип
# "без солі", що вже обґрунтований коментарем над BACKUP_PASSWORD_PATH
# (жодного файлу-солі, який можна загубити й тим самим зробити нечитабельними
# ВСІ вже зроблені знімки) - PBKDF2 із фіксованою сіллю не захищає від
# райдужних таблиць між РІЗНИМИ інсталяціями цієї самої программи (усі
# ділять ту саму сіль), але й не мав захищати - єдина реальна мета тут
# (стишити грубий перебір) досягається саме кількістю ітерацій, не сіллю.
# 600 000 ітерацій SHA-256 - поточна рекомендація OWASP для PBKDF2-HMAC-
# SHA256 (2023+); викликається лише на кожен знімок/розшифрування, у
# фоновому потоці - не UI-критичний шлях, кількасот мс не відчутні.
_BACKUP_KEY_SALT = b"AI_Automation.backup_encryption.v1"
_BACKUP_KEY_ITERATIONS = 600_000


def _backup_fernet_key(password):
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), _BACKUP_KEY_SALT, _BACKUP_KEY_ITERATIONS, dklen=32,
    )
    return base64.urlsafe_b64encode(digest)


def _encrypt_backup_file(path, password):
    from cryptography.fernet import Fernet

    fernet = Fernet(_backup_fernet_key(password))
    encrypted = fernet.encrypt(path.read_bytes())
    enc_path = path.with_name(path.name + ".enc")
    enc_path.write_bytes(encrypted)
    # Реальний ризик (аудит коду, 2026-08-14): весь сенс шифрування бекапів
    # - щоб той, хто отримав доступ ЛИШЕ до папки backups, не міг прочитати
    # дані без пароля. .enc-файл вище вже надійно записаний на цей момент,
    # тож САМІ ДАНІ в безпеці - але якщо видалення НЕзашифрованого
    # оригіналу (напр. антивірус тримає файл, диск переповнений) впаде без
    # жодного сліду, поруч із .enc лишається читабельний .sqlite3 з тими
    # самими даними - шифрування стає марним, а ніхто про це не дізнається.
    # Явне, конкретне повідомлення замість голого OSError, що вказує САМЕ
    # на файл, який треба прибрати вручну.
    try:
        path.unlink()
    except OSError as exc:
        raise OSError(
            f"Резервну копію зашифровано ({enc_path.name}), але не вдалося видалити "
            f"незашифрований тимчасовий файл {path}: {exc}. "
            "Видаліть цей файл вручну - інакше поруч із зашифрованою копією "
            "лишаються ті самі дані без пароля."
        ) from exc
    return enc_path


def _decrypt_backup_bytes(path, password):
    from cryptography.fernet import Fernet, InvalidToken

    fernet = Fernet(_backup_fernet_key(password))
    try:
        return fernet.decrypt(path.read_bytes())
    except InvalidToken:
        return None
    except (ValueError, TypeError):
        # Реальний ризик (аудит коду, 2026-08-14): пошкоджений/обрізаний
        # .enc-файл (крах/диск повний ПОСЕРЕД шифрування) не завжди дає
        # InvalidToken - невалідний base64 всередині fernet.decrypt кидає
        # binascii.Error (підклас ValueError) чи TypeError РАНІШЕ, ніж
        # дійде до перевірки підпису. Раніше це не ловилось тут і виходило
        # сирим трейсбеком аж до GUI замість "не вдалось розшифрувати" -
        # для викликача (restore_db_snapshot) обидва випадки означають
        # ОДНЕ й те саме: "з цим файлом щось не так", тому лишаємо None.
        return None


def _db_snapshot_paths():
    return list(DB_BACKUP_DIR.glob("app_data_*.sqlite3")) + list(DB_BACKUP_DIR.glob("app_data_*.sqlite3.enc"))


def _db_snapshot_base_name(path):
    """Ім'я знімка без розширення(-нь) - '.sqlite3.enc' знімає ОБИДВА суфікси
    за раз (Path.stem знімає лише останній, тож для .enc-файлів лишив би
    висячий '.sqlite3' і зламав би перевірку "_pre_restore" нижче)."""
    name = path.name
    if name.endswith(".sqlite3.enc"):
        return name[: -len(".sqlite3.enc")]
    if name.endswith(".sqlite3"):
        return name[: -len(".sqlite3")]
    return path.stem


def create_db_snapshot(db_path, label=None):
    """Безпечна копія ЖИВОЇ SQLite-бази через sqlite3 Connection.backup() —
    коректно працює, навіть якщо інше з'єднання (бот/GUI) саме зараз пише в
    ту саму базу. Свідомо НЕ звичайний shutil.copy живого файлу."""
    DB_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    # Мікросекундна точність (як і create_excel_backup вище) - секундної
    # роздільності не вистачило б, якби два знімки (напр. щоденний і
    # pre_restore) стались в межах однієї секунди: другий тихо переписав би
    # файл першого замість створення окремого.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    suffix = f"_{label}" if label else ""
    target_path = DB_BACKUP_DIR / f"app_data_{timestamp}{suffix}.sqlite3"
    source_conn = sqlite3.connect(str(db_path))
    try:
        target_conn = sqlite3.connect(str(target_path))
        try:
            source_conn.backup(target_conn)
        finally:
            target_conn.close()
    finally:
        source_conn.close()
    password = _backup_encryption_password()
    if password:
        target_path = _encrypt_backup_file(target_path, password)
    _rotate_db_snapshots()
    return target_path


def _rotate_db_snapshots(limit=DB_BACKUP_LIMIT):
    snapshots = sorted(
        _db_snapshot_paths(),
        key=lambda path: path.stat().st_mtime,
    )
    excess = len(snapshots) - limit
    if excess <= 0:
        return
    for path in snapshots[:excess]:
        path.unlink(missing_ok=True)


def list_db_snapshots():
    """Метадані знімків для GUI, найновіші перші."""
    if not DB_BACKUP_DIR.exists():
        return []
    snapshots = sorted(
        _db_snapshot_paths(),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return [
        {
            "path": path,
            "mtime": path.stat().st_mtime,
            "size": path.stat().st_size,
            "is_pre_restore": _db_snapshot_base_name(path).endswith("_pre_restore"),
            "is_encrypted": path.suffix == ".enc",
        }
        for path in snapshots
    ]


def maybe_create_scheduled_snapshot(db_path):
    """Ідемпотентний дешевий чек (один glob, коли робити нічого не треба):
    якщо знімок за СЬОГОДНІШНЮ дату вже є — нічого не робити. Інакше —
    create_db_snapshot. Викликається і з gui.py (старт + періодичний тик),
    і з main.py (кожна ітерація поллінгу бота) — обидва виклики безпечно
    надлишкові один для одного, коли працюють одночасно."""
    today = datetime.now().strftime("%Y%m%d")
    # Свіжий пере-аудит (2026-08-02, New-Important #6): pre_restore-знімок
    # (створюється restore_db_snapshot ПЕРЕД відновленням) мав ту саму
    # дату-в-імені, тож "*" у glob нижче раніше збігався і з ним —
    # відновлення бекапу в будь-який момент дня хибно "гасило" реальний
    # щоденний знімок на решту того самого дня. Виключаємо pre_restore
    # окремо (той самий суфікс, що вже перевіряє list_db_snapshots вище).
    if DB_BACKUP_DIR.exists() and any(
        not _db_snapshot_base_name(path).endswith("_pre_restore")
        for path in list(DB_BACKUP_DIR.glob(f"app_data_{today}_*.sqlite3"))
        + list(DB_BACKUP_DIR.glob(f"app_data_{today}_*.sqlite3.enc"))
    ):
        return None
    return create_db_snapshot(db_path)


# Реальна знахідка (аудит коду, 2026-08-16): відновлення НЕЗАШИФРОВАНОГО
# знімка раніше було сирим shutil.copyfile без жодної перевірки, що джерело
# взагалі є коректною SQLite-базою - пошкоджений файл (напр. конфлікт
# синхронізації OneDrive, обірваний диск) прийнявся б так само, як
# справжній, і живу базу було б уже перезаписано ДО того, як хтось помітив
# проблему. Зашифровані знімки отримують цю перевірку безкоштовно через
# власний HMAC Fernet (_decrypt_backup_bytes падає на biту/підроблену копію
# ще ДО запису) - незашифровані такого захисту не мали.
def _looks_like_valid_sqlite_bytes(data):
    if data[:16] != b"SQLite format 3\x00":
        return False
    tmp_path = Path(tempfile.mkdtemp()) / "integrity_check.sqlite3"
    try:
        tmp_path.write_bytes(data)
        conn = sqlite3.connect(str(tmp_path))
        try:
            result = conn.execute("PRAGMA integrity_check").fetchone()
            return bool(result) and result[0] == "ok"
        finally:
            conn.close()
    except (OSError, sqlite3.Error):
        return False
    finally:
        shutil.rmtree(tmp_path.parent, ignore_errors=True)


def restore_db_snapshot(db_path, snapshot_path):
    """Відновлює db_path зі snapshot_path. Викликач (gui.py) МАЄ вже закрити
    всі живі з'єднання до db_path і зупинити бота ДО цього виклику — тут це
    не перевіряється повторно (відповідальність викликача, як і перевірка
    ролі в _is_real_telegram_user не повторюється в кожній apply_*_operation).

    Крок 1 (безпека, задача користувача "можна спокійно повернутись до
    новішого стану"): знімок ПОТОЧНОГО, ще не заміненого стану з міткою
    pre_restore — так відновлення ніколи не є точкою неповернення: щоб
    повернутись до новішого, досить відновити САМЕ ЦЕЙ знімок.
    Крок 2: заміна файлу — безпечна тут, бо джерело (snapshot_path) вже
    закритий статичний файл, а не жива база (на відміну від create_db_
    snapshot вище, де джерело — жива база, тому там і потрібен .backup()."""
    create_db_snapshot(db_path, label="pre_restore")
    snapshot_path = Path(snapshot_path)
    if snapshot_path.suffix == ".enc":
        password = _backup_encryption_password()
        if not password:
            raise ValueError(
                "Этот снимок зашифрован, а пароль резервных копий сейчас не задан в настройках."
            )
        data = _decrypt_backup_bytes(snapshot_path, password)
        if data is None:
            raise ValueError(
                "Не удалось расшифровать снимок. Либо пароль резервных копий не совпадает с тем, "
                "которым снимок был зашифрован, либо сам файл снимка повреждён — попробуйте другой снимок."
            )
        if not _looks_like_valid_sqlite_bytes(data):
            raise ValueError(
                "Снимок расшифровался, но получившийся файл не является корректной базой данных — "
                "снимок повреждён. Живая база не тронута, попробуйте другой снимок."
            )
        Path(db_path).write_bytes(data)
    else:
        if not _looks_like_valid_sqlite_bytes(snapshot_path.read_bytes()):
            raise ValueError(
                "Файл снимка повреждён и не является корректной базой данных. "
                "Живая база не тронута, попробуйте другой снимок."
            )
        shutil.copyfile(snapshot_path, db_path)


def regenerate_excel_after_restore(db_path):
    """Перебудовує Excel-мірор із щойно відновленої бази. ОБОВ'ЯЗКОВИЙ крок
    одразу після restore_db_snapshot: gui.py реімпортує дані З Excel при
    кожному запуску — без цього наступний запуск GUI тихо перезаписав би
    щойно відновлені дані назад старим Excel-вмістом, зводячи відновлення
    нанівець."""
    store = ExcelSqliteStore(db_path)
    try:
        sync_sheets_to_excel(store, ["СКЛАД", SALES_SHEET_NAME])
        sync_antiseptic_to_excel(store)
    finally:
        store.close()


# Шар даних на SQLite: тут живуть усі таблиці (sheet_rows, bot_users,
# bot_pending_operations, stock_movements тощо) і CRUD-методи над ними.
# І TelegramBotWorker, і ExcelViewerApp відкривають свій власний
# ExcelSqliteStore(DB_PATH) — обидва пишуть в один і той самий файл БД.
class ExcelSqliteStore:
    # --- Життєвий цикл + створення схеми БД (CREATE TABLE) ---
    # Реальний ризик (аудит коду, 2026-08-14): бот, GUI і client_app.py
    # кожен тримає СВОЄ власне з'єднання до того самого файлу БД
    # одночасно - gui.py вже має коментар, що визнає "database is locked"
    # реальним ризиком саме через це. sqlite3.connect() без явного timeout
    # мовчки покладається на власний дефолт Python (5 секунд) - ніде в
    # цьому файлі, дуже уважному до одночасного доступу (BEGIN IMMEDIATE
    # усюди), це не було свідомим вибором. 15с - явний, навмисний запас
    # (звичайна операція приходу/продажу триває мілісекунди; довший
    # утримувач блокування - це масове редагування в GUI чи повний
    # реімпорт Excel при старті) замість неозначеного дефолту.
    def __init__(self, db_path):
        self.db_path = Path(db_path)
        self.conn = sqlite3.connect(self.db_path, timeout=15)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._init_db()

    def close(self):
        self.conn.close()

    def _init_db(self):
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sheet_meta (
                sheet_name TEXT PRIMARY KEY,
                headers_json TEXT NOT NULL,
                read_only INTEGER NOT NULL DEFAULT 0,
                imported_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sheet_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_name TEXT NOT NULL,
                position INTEGER NOT NULL,
                values_json TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (sheet_name) REFERENCES sheet_meta(sheet_name)
                    ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_sheet_rows_sheet_position
                ON sheet_rows(sheet_name, position);

            CREATE TABLE IF NOT EXISTS bot_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE,
                username TEXT,
                full_name TEXT,
                role TEXT NOT NULL DEFAULT 'user',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS bot_commands (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                title TEXT NOT NULL,
                description TEXT,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS bot_command_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                command_id INTEGER NOT NULL,
                phrase TEXT NOT NULL UNIQUE,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (command_id) REFERENCES bot_commands(id)
                    ON DELETE CASCADE,
                FOREIGN KEY (created_by_user_id) REFERENCES bot_users(id)
                    ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS bot_unknown_phrases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phrase TEXT NOT NULL,
                telegram_user_id INTEGER,
                chat_id INTEGER,
                status TEXT NOT NULL DEFAULT 'new',
                created_at TEXT NOT NULL,
                resolved_at TEXT
            );

            CREATE TABLE IF NOT EXISTS bot_user_preferences (
                telegram_user_id TEXT PRIMARY KEY,
                chat_id TEXT,
                username TEXT,
                full_name TEXT,
                bot_mode TEXT NOT NULL DEFAULT 'no_ai',
                language TEXT NOT NULL DEFAULT 'ru',
                claude_api_key TEXT,
                claude_key_updated_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS action_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                actor_user_id INTEGER,
                action_type TEXT NOT NULL,
                details_json TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (actor_user_id) REFERENCES bot_users(id)
                    ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                movement_type TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT 'telegram',
                telegram_user_id TEXT,
                username TEXT,
                full_name TEXT,
                product TEXT,
                breed TEXT,
                condition TEXT,
                thickness REAL,
                width REAL,
                length REAL,
                quantity REAL,
                volume REAL,
                area REAL,
                linear REAL,
                sheet_row_id INTEGER,
                original_text TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_stock_movements_created_at
                ON stock_movements(created_at);

            CREATE INDEX IF NOT EXISTS idx_stock_movements_type_created_at
                ON stock_movements(movement_type, created_at);

            CREATE TABLE IF NOT EXISTS bot_pending_operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id TEXT NOT NULL,
                telegram_user_id TEXT NOT NULL,
                operation_type TEXT NOT NULL,
                status TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_bot_pending_operations_active
                ON bot_pending_operations(chat_id, telegram_user_id);

            -- Кастомні кнопки бота (Редактор кнопок, GUI) — дерево довільної
            -- глибини: parent_id NULL = коренева кнопка (додається до
            -- головного меню бота), інакше — дитина іншої кастомної кнопки.
            -- section обов'язковий лише для кореня (визначає, які дії з
            -- CUSTOM_BUTTON_ACTIONS дозволені для всієї цієї гілки); дочірні
            -- вузли успадковують секцію свого кореня через обхід parent_id.
            -- Перший робочий зріз (Задача користувача, поетапно) підтримує
            -- лише кореневі кнопки БЕЗ дітей — action_code/message_text
            -- використовуються, дерево (parent_id у інших рядків) буде
            -- задіяне пізніше, коли існуючі пункти меню почнуть переносити
            -- сюди по одному.
            CREATE TABLE IF NOT EXISTS custom_menu_buttons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                parent_id INTEGER,
                section TEXT,
                label TEXT NOT NULL,
                message_text TEXT,
                action_code TEXT,
                position INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (parent_id) REFERENCES custom_menu_buttons(id)
                    ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_custom_menu_buttons_parent
                ON custom_menu_buttons(parent_id, position);

            -- Задача користувача: "видаляю кнопку - вона потім знову
            -- з'являється". _seed_builtin_migrated_custom_buttons (нижче)
            -- звіряє лише "чи є ЗАРАЗ рядок із цим migration_key" - після
            -- реального DELETE рядка (delete_custom_button) цей рядок
            -- зникає, тож НАСТУПНИЙ запуск програми бачив "ще не засіяно" й
            -- мовчки вставляв кнопку знову. Ця таблиця - надгробок
            -- ("цей migration_key адміністратор видалив НАВМИСНО, більше
            -- ніколи не досіювати") - заповнюється в delete_custom_button,
            -- звіряється в _seed_builtin_migrated_custom_buttons.
            CREATE TABLE IF NOT EXISTS deleted_builtin_migrations (
                migration_key TEXT PRIMARY KEY,
                deleted_at TEXT NOT NULL
            );

            -- Задача користувача: адміністратор може перевизначити ТЕКСТ,
            -- який бот показує в чаті на вході в конкретну дію (напр.
            -- "Приход. Выберите категорию товара:") — без правки коду.
            -- key = action_code (BOT_MESSAGE_DEFAULTS нижче); відсутність
            -- рядка = типовий текст ще не змінювали (get_message_template
            -- повертає default, переданий викликачем).
            CREATE TABLE IF NOT EXISTS bot_message_templates (
                key TEXT PRIMARY KEY,
                text TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            -- Задача користувача (Крок 3+ "Дії"): "дія" = редагована категорія
            -- на кшталт "ДОСКА AD" (лист, що бот показує під ПРИХОД/
            -- РЕАЛИЗАЦИЯ) — на відміну від самих ПРИХОД/РЕАЛИЗАЦИЯ (то чиста
            -- навігація без власного запису в таблицю). Кожна дія має
            -- впорядкований список полів-запитів (bot_operation_fields), і
            -- кожне поле — список прив'язок "вкладка+колонка++/-/показ"
            -- (bot_operation_field_columns). Розпізнавання ТЕКСТУ користувача
            -- (як бот розуміє, що слово "сосна" — це порода) тут НЕ
            -- описується і не змінюється — лише куди йде вже розпізнане
            -- значення. builtin_key — той самий одноразовий-сідінг ідіом, що
            -- й custom_menu_buttons.migration_key: після першого запуску це
            -- звичайний рядок, подальші правки адміністратора (включно з
            -- видаленням) ніколи не відкочуються повторним стартом застосунку.
            CREATE TABLE IF NOT EXISTS bot_operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                kind TEXT NOT NULL,
                requires_row_identity INTEGER NOT NULL DEFAULT 1,
                label TEXT NOT NULL,
                parent_action_code TEXT NOT NULL,
                prefill_json TEXT,
                position INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                builtin_key TEXT UNIQUE,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_bot_operations_parent
                ON bot_operations(parent_action_code, position);

            -- is_identity=1 захищає поле-запит від видалення/вимкнення
            -- (Порода/Товар/Товщина/Ширина/Довжина, коли в операції
            -- requires_row_identity=1) — _warehouse_row_matches
            -- (telegram_dialog.py) шукає потрібний рядок складу саме за
            -- ними, без жодного винятку "якщо порожньо — пропустити": якщо
            -- прибрати запит цього поля з чек-листа, бот перестане питати
            -- значення, і пошук рядка почне ламатись. Захист програмний
            -- (у store-методах нижче), не лише в GUI.
            CREATE TABLE IF NOT EXISTS bot_operation_fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_id INTEGER NOT NULL REFERENCES bot_operations(id) ON DELETE CASCADE,
                field_key TEXT NOT NULL,
                label TEXT NOT NULL,
                is_identity INTEGER NOT NULL DEFAULT 0,
                position INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                builtin_key TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE (operation_id, field_key)
            );

            CREATE INDEX IF NOT EXISTS idx_bot_operation_fields_operation
                ON bot_operation_fields(operation_id, position);

            -- write_mode='generic' — прив'язку виконує execute_operation_write
            -- (просте +/- на вже знайденому рядку СКЛАД, через наявний
            -- add_to_row_value); write_mode='ledger' — лише показ у
            -- редакторі (реальний запис і далі робить sale_sheet_values/
            -- antiseptic_sheet_values, бо там є бізнес-логіка — сума,
            -- розподіл готівка/банк, нумерація документа тощо, а не проста
            -- мапа поле->колонка).
            CREATE TABLE IF NOT EXISTS bot_operation_field_columns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_field_id INTEGER NOT NULL REFERENCES bot_operation_fields(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                column_key TEXT NOT NULL,
                marker TEXT NOT NULL,
                write_mode TEXT NOT NULL DEFAULT 'generic',
                position INTEGER NOT NULL DEFAULT 0,
                builtin_key TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_bot_operation_field_columns_field
                ON bot_operation_field_columns(operation_field_id, position);

            -- Крок 4.4 "Дії": реальний редагований перелік способів оплати
            -- (раніше — жорсткий PAYMENT_METHODS = 3 назви). label — те, що
            -- бот показує на кнопці й пише в нові записи ЗАРАЗ; змінюється
            -- при перейменуванні. payment_method_synonyms зберігає СТАРІ
            -- назви (і початковий латинський "alt") — розпізнавання
            -- вільного тексту й далі приймає їх, навіть після перейменування
            -- відповідного варіанту, назавжди (Задача користувача: "щоб бот
            -- уже це сприймав у запиті автоматично", жодна стара звичка
            -- набору не має "загубитись"). Видалення варіанту (не
            -- перейменування!) каскадно прибирає й усі його синоніми —
            -- саме тому й тільки тому воно дійсно прибирає розпізнавання.
            CREATE TABLE IF NOT EXISTS payment_method_options (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                label TEXT NOT NULL,
                position INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS payment_method_synonyms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                option_id INTEGER NOT NULL REFERENCES payment_method_options(id) ON DELETE CASCADE,
                phrase TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_payment_method_synonyms_option
                ON payment_method_synonyms(option_id);

            -- Задача користувача: 3 окремі можливості створювати свої
            -- шаблони для Приход/Реализация/Списание (webapp-мега-форми) —
            -- готовий набір розмір+порода(+клієнт+спосіб оплати для продажу),
            -- обраний зі списку одним тапом замість повторного набору вручну.
            -- Одна спільна таблиця на всі 3 kind (структура ідентична,
            -- client/payment_method просто NULL для income/writeoff) —
            -- дешевше за 3 майже однакові таблиці. Цілеспрямовано БЕЗ ціни
            -- за одиницю/кількості (Задача користувача: "лишається дописати
            -- ціну штуки і т д" — це завжди вводиться заново, шаблон лише
            -- прискорює вибір розміру/породи).
            CREATE TABLE IF NOT EXISTS operation_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kind TEXT NOT NULL,
                category_operation_id INTEGER,
                breed TEXT,
                thickness REAL,
                width REAL,
                length REAL,
                client TEXT,
                address TEXT,
                payment_method TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_operation_templates_kind
                ON operation_templates(kind, created_at);

            -- "А праворуч, таким же списком, 5 останніх створених" — та сама
            -- форма даних, але наповнюється АВТОМАТИЧНО (не користувачем
            -- явно) на кожне реальне подання мега-форми, щоб показати живу
            -- історію без жодної ручної дії.
            CREATE TABLE IF NOT EXISTS operation_recent_uses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kind TEXT NOT NULL,
                category_operation_id INTEGER,
                breed TEXT,
                thickness REAL,
                width REAL,
                length REAL,
                client TEXT,
                address TEXT,
                payment_method TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_operation_recent_uses_kind
                ON operation_recent_uses(kind, created_at);

            -- Задача користувача: категорії товару (ДОСКА AD і т.д. —
            -- bot_operations, kind income/sale/service) мають реальний
            -- CRUD додати/перейменувати/видалити, з тим самим принципом
            -- "перейменування зберігає стару назву як синонім", що й у
            -- payment_method_synonyms вище. Раніше _category_keyboard/
            -- _category_from_text (telegram_dialog.py) були повністю
            -- хардкоджені й НЕ мали жодного зв'язку з bot_operations —
            -- тепер обидва store-driven, а ця таблиця — та сама роль,
            -- що й payment_method_synonyms, лише для категорій.
            CREATE TABLE IF NOT EXISTS bot_operation_synonyms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_id INTEGER NOT NULL REFERENCES bot_operations(id) ON DELETE CASCADE,
                phrase TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_bot_operation_synonyms_operation
                ON bot_operation_synonyms(operation_id);

            CREATE TABLE IF NOT EXISTS dev_work_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                summary TEXT NOT NULL,
                benefit TEXT,
                future_impact TEXT,
                created_at TEXT NOT NULL
            );

            -- Пам'ять виправлень імені клієнта: коли продавець підтверджує
            -- "Прийняти і запам'ятати" для одруку ("jospeh" -> "Joseph"),
            -- наступного разу той самий одрук вирішується мовчки, без
            -- повторного питання. normalized_typo — саме те, що ввів
            -- продавець, приведене до єдиного вигляду (_normalize_phrase),
            -- щоб "Jospeh"/"jospeh"/"ДЖОСПЕХ" збігались з одним записом.
            CREATE TABLE IF NOT EXISTS client_name_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                normalized_typo TEXT NOT NULL UNIQUE,
                canonical_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS warehouse_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_row_id INTEGER UNIQUE NOT NULL,
                sku TEXT,
                product TEXT,
                breed TEXT,
                condition TEXT,
                thickness REAL,
                width REAL,
                length REAL,
                unit TEXT,
                income_qty REAL,
                income_volume REAL,
                income_area REAL,
                income_linear REAL,
                sold_qty REAL,
                sold_volume REAL,
                sold_area REAL,
                sold_linear REAL,
                balance_qty REAL,
                balance_volume REAL,
                balance_area REAL,
                balance_linear REAL,
                FOREIGN KEY (sheet_row_id) REFERENCES sheet_rows(id)
                    ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_warehouse_items_product
                ON warehouse_items(product, breed, condition);

            CREATE INDEX IF NOT EXISTS idx_warehouse_items_dims
                ON warehouse_items(thickness, width, length);

            CREATE INDEX IF NOT EXISTS idx_warehouse_items_balance
                ON warehouse_items(balance_qty, balance_volume, balance_area);

            -- Важлива знахідка нового аудиту (28.07.2026, #4): номер
            -- документа (document_number/service_number) раніше рахувався
            -- як len(fetch_rows(...))+1 — вразливо до видалення рядка
            -- (delete_row у GUI чи повторний імпорт з Excel): наступний
            -- запис міг порахувати номер, що вже присвоєний рядку, який НЕ
            -- видалили. Персистентний лічильник (per sheet_name) імунний і
            -- до видалення, і до ручного document_type-перезапису.
            CREATE TABLE IF NOT EXISTS document_counters (
                sheet_name TEXT PRIMARY KEY,
                next_number INTEGER NOT NULL
            );

            -- Задача користувача (2026-08-14): "прибери запам'ятовування.
            -- криво працює" - фільтри "Данные" більше НЕ читаються й НЕ
            -- пишуться сюди (store.get_webapp_data_browser_prefs/save_
            -- webapp_data_browser_prefs видалені) - таблиця лишена
            -- порожньою заготовкою без міграції видалення, щоб не чіпати
            -- схему заради вже й так неактивної фічі.
            CREATE TABLE IF NOT EXISTS webapp_data_browser_prefs (
                telegram_user_id TEXT PRIMARY KEY,
                remember_enabled INTEGER NOT NULL DEFAULT 1,
                prefs_json TEXT,
                updated_at TEXT NOT NULL
            );
            """
        )
        self.conn.execute(
            """
            INSERT INTO app_meta (key, value)
            VALUES ('schema_version', '10')
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """
        )
        self.conn.commit()
        self._ensure_column("bot_user_preferences", "claude_api_key", "TEXT")
        self._ensure_column("bot_user_preferences", "claude_key_updated_at", "TEXT")
        self._ensure_column("warehouse_items", "income_area", "REAL")
        self._ensure_column("warehouse_items", "sold_area", "REAL")
        self._ensure_column("warehouse_items", "balance_area", "REAL")
        # мп (погонні метри, розміри 25x50/30x50/50x50) — той самий підхід,
        # що й для area вище: додаткові колонки для вже існуючих БД, які
        # не мали цих полів на момент створення.
        self._ensure_column("warehouse_items", "income_linear", "REAL")
        self._ensure_column("warehouse_items", "sold_linear", "REAL")
        self._ensure_column("warehouse_items", "balance_linear", "REAL")
        # stock_movements раніше взагалі не мав "area" — Вагонка (і будь-
        # який інший площинний товар) в історії приходу/продажу завжди
        # показувала "0 м3" (реальний, давній баг, знайдений при роботі
        # над мп). Виправляємо разом з додаванням "linear".
        self._ensure_column("stock_movements", "area", "REAL")
        self._ensure_column("stock_movements", "linear", "REAL")
        # Списання (writeoff): необов'язкова вільнотекстова причина - на
        # відміну від приходу/продажу, тут немає власного листа-довідника
        # (стор._sync_warehouse_item/apply_writeoff_operation), тож єдиний
        # запис аудиту - саме цей рядок stock_movements.
        self._ensure_column("stock_movements", "reason", "TEXT")
        # Шаблони/недавні мега-форми (Задача користувача: "в історії
        # зберігається все... окрім ціни, штук") спершу забули адресу
        # вивантаження - вона теж мала зберігатись разом з клієнтом/оплатою.
        self._ensure_column("operation_templates", "address", "TEXT")
        self._ensure_column("operation_recent_uses", "address", "TEXT")
        # Розмір кнопки в клавіатурі Telegram (Задача користувача): "full" —
        # суцільна, на весь рядок; "half" — вдвічі менша, спарюється із
        # СУСІДНЬОЮ (за порядком position) половинною кнопкою в ОДИН рядок
        # клавіатури (менший position — зліва, більший — справа; окремого
        # "боку" немає). NULL у вже існуючих рядків (до цього стовпця)
        # трактується як "full" на рівні коду (_custom_button_layout), не тут.
        self._ensure_column("custom_menu_buttons", "layout", "TEXT")
        # migration_key — позначає кнопки, перенесені з хардкоду (ПРИХОД
        # і т.д. по одній, Задача користувача) — див. BUILTIN_MIGRATED_
        # CUSTOM_BUTTONS і _seed_builtin_migrated_custom_buttons нижче.
        self._ensure_column("custom_menu_buttons", "migration_key", "TEXT")
        # Крок 4.3 (Задача користувача, "пряме посилання на дію з 'Дії'"):
        # NULL = вузол, як і раніше, працює через children/action_code;
        # заповнене значення — ID рядка bot_operations (ДОСКА AD і т.д.),
        # вузол ОДРАЗУ запускає саме цю дію (_start_operation_leaf,
        # telegram_dialog.py), в обхід зношеного action_code+_category_
        # from_text механізму. Взаємовиключне з action_code на рівні GUI-
        # форми (радіо-вибір), не на рівні схеми — SQLite FK тут навмисно
        # відсутній (bot_operations.id не завжди стабільний — категорію чи
        # дію можна видалити, delete_operation/delete_operation_category
        # самі очищають "осиротілі" custom_menu_buttons.operation_id перед
        # видаленням, аудит коду знайшов цю прогалину — див. коментарі там).
        self._ensure_column("custom_menu_buttons", "operation_id", "INTEGER")
        # Задача користувача: "час останнього відправленого повідомлення
        # користувачем в чат" у "Персонал" - NULL для рядків, доданих
        # вручну (ще ніколи не писали боту), оновлюється щоразу через
        # ensure_bot_user_seen.
        self._ensure_column("bot_users", "last_seen_at", "TEXT")
        # Аудит коду: antiseptic_sheet_values раніше визначала "це банківський
        # переказ?" порівнянням із жорстким рядком "ЕФАКТУРА Б/Н" — перейменування
        # варіанту оплати (штатна дія через "Способи оплати") тихо ламало
        # розподіл готівка/банк. kind='bank' на РІВНО ОДНОМУ рядку — стабільний
        # прапорець, що переживає перейменування (NULL для решти = готівка,
        # як і зараз). НЕ вгадуємо, який із наявних варіантів позначити —
        # лишається NULL для всіх, поки адмін сам не позначить через новий
        # перемикач у "Способи оплати".
        self._ensure_column("payment_method_options", "kind", "TEXT")
        # Половинна кнопка більше не має "боку" — зліва/справа тепер визначає
        # сама позиція (менший position = ліва половина пари, більший = права,
        # див. _pack_custom_button_rows у telegram_dialog.py), тож старі
        # значення half_left/half_right зводяться до єдиного "half". Ідемпотентно
        # (безпечно виконувати щоразу), не потребує зміни schema_version.
        with self.conn:
            self.conn.execute(
                "UPDATE custom_menu_buttons SET layout = 'half' WHERE layout IN ('half_left', 'half_right')"
            )
        self._seed_builtin_bot_commands()
        self._seed_builtin_migrated_custom_buttons()
        self._apply_standard_menu_policy()
        self._backfill_writeoff_root_action_code()
        self._backfill_writeoff_form_root_label()
        self._seed_builtin_operations()
        self._seed_report_operations()
        self._migrate_add_address_field()
        self._migrate_osb_quantity_only()
        self._fix_operation_field_language()
        self._simplify_measure_field_label()
        self._relabel_vagonka_measure_field()
        self._remove_noop_identity_bindings()
        self._restructure_report_fields()
        self._clear_message_text_for_builtin_actions()
        self._seed_payment_method_options()
        self._seed_operation_category_synonyms()
        self._backfill_bot_user_names()
        self._seed_known_personnel()

    def _ensure_column(self, table_name, column_name, column_sql):
        columns = {
            row[1]
            for row in self.conn.execute(f"PRAGMA table_info({table_name})").fetchall()
        }
        if column_name in columns:
            return
        self.conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}")
        self.conn.commit()

    def _seed_builtin_bot_commands(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            for command in BUILTIN_BOT_COMMANDS:
                self.conn.execute(
                    """
                    INSERT INTO bot_commands (code, title, description, enabled, created_at, updated_at)
                    VALUES (?, ?, ?, 1, ?, ?)
                    ON CONFLICT(code) DO UPDATE SET
                        title = excluded.title,
                        description = excluded.description,
                        enabled = 1,
                        updated_at = excluded.updated_at
                    """,
                    (
                        command["code"],
                        command["title"],
                        command["description"],
                        now,
                        now,
                    ),
                )
                command_id = self.conn.execute(
                    "SELECT id FROM bot_commands WHERE code = ?",
                    (command["code"],),
                ).fetchone()[0]
                for alias in command["aliases"]:
                    self.conn.execute(
                        """
                        INSERT OR IGNORE INTO bot_command_aliases (command_id, phrase, created_at)
                        VALUES (?, ?, ?)
                        """,
                        (command_id, _normalize_phrase(alias), now),
                    )

    # На відміну від _seed_builtin_bot_commands (upsert — оновлює
    # title/description при кожному запуску), тут НЕМАЄ ON CONFLICT DO
    # UPDATE: рядок вставляється РІВНО ОДИН РАЗ (перевірка за migration_key
    # перед INSERT) — далі це звичайна кнопка, і правки адміністратора
    # (назва/текст/розмір, і навіть видалення) ніколи не відкочуються
    # повторним запуском застосунку.
    def _seed_builtin_migrated_custom_buttons(self):
        now = datetime.now().isoformat(timespec="seconds")
        for entry in BUILTIN_MIGRATED_CUSTOM_BUTTONS:
            exists = self.conn.execute(
                "SELECT 1 FROM custom_menu_buttons WHERE migration_key = ?",
                (entry["migration_key"],),
            ).fetchone()
            if exists:
                continue
            # Реальний баг з аудиту: перевірка вище бачить лише "чи є рядок
            # ЗАРАЗ" - адміністратор, що видалив кнопку через "x", отримував
            # її НАЗАД на наступному запуску, бо рядок (і його migration_key)
            # реально зникав разом із DELETE. deleted_builtin_migrations -
            # надгробок, який delete_custom_button заповнює саме для цього.
            tombstoned = self.conn.execute(
                "SELECT 1 FROM deleted_builtin_migrations WHERE migration_key = ?",
                (entry["migration_key"],),
            ).fetchone()
            if tombstoned:
                continue
            parent_migration_key = entry.get("parent_migration_key")
            parent_id = None
            if parent_migration_key is not None:
                parent_row = self.conn.execute(
                    "SELECT id FROM custom_menu_buttons WHERE migration_key = ?",
                    (parent_migration_key,),
                ).fetchone()
                if parent_row is None:
                    continue
                parent_id = parent_row[0]
            with self.conn:
                position = self.conn.execute(
                    "SELECT COALESCE(MAX(position), -1) + 1 FROM custom_menu_buttons WHERE parent_id IS ?",
                    (parent_id,),
                ).fetchone()[0]
                self.conn.execute(
                    """
                    INSERT INTO custom_menu_buttons
                        (parent_id, section, label, message_text, action_code, position, layout,
                         enabled, migration_key, created_at, updated_at)
                    VALUES (?, NULL, ?, '', ?, ?, ?, 1, ?, ?, ?)
                    """,
                    (
                        parent_id,
                        entry["label"],
                        entry["action_code"],
                        position,
                        entry["layout"],
                        entry["migration_key"],
                        now,
                        now,
                    ),
                )

    # "Стандартне меню" — Задача користувача (2026-08-18, після повторного
    # "калькулятор виліз знову, ти ж казав що приберешь"): попередній
    # механізм (_hide_legacy_non_form_operation_buttons_once, ОДИН
    # глобальний прапорець 'legacy_non_form_buttons_hidden' в app_meta)
    # ховав рівно 3 конкретні migration_key РІВНО ОДИН РАЗ — коли пізніше
    # знадобилось сховати ще й "data_menu"/"calculator"/"help" (просто
    # забуті в першому списку), дописування їх у той старий список НІЧОГО
    # не дало б на вже існуючих продакшн-базах: прапорець уже стояв '1',
    # функція виходила одразу, до нового ключа код не доходив. Це РЕАЛЬНА
    # причина, чому ДАННЫЕ/Калькулятор/Помощь "поверталися" — вони
    # ніколи насправді не ховались на живій базі, не "воскресали".
    #
    # Фікс — ПЕР-КЛЮЧОВА мітка замість одного спільного прапорця:
    # app_meta['standard_menu_resolved:<migration_key>'] окремо на кожен
    # migration_key. Будь-який КОРІННИЙ (parent_migration_key=None) пункт
    # з BUILTIN_MIGRATED_CUSTOM_BUTTONS, якого немає в
    # _STANDARD_MENU_ROOT_MIGRATION_KEYS, ховається (enabled=0) РІВНО ОДИН
    # РАЗ на своєму власному ключі — додавання будь-якого нового пункту в
    # майбутньому більше ніколи не залежить від того, чи вже "спрацював"
    # якийсь інший прапорець (той самий клас багу тут просто не повториться).
    # Дітей схованого батька (СКЛАД/ПРОДАЖИ/... під ДАННЫЕ) не чіпаємо —
    # вони й так стають недосяжними без кнопки-батька, яка б до них вела.
    #
    # Рішення адміністратора ЗАВЖДИ важливіше за цю політику: якщо адмін
    # пізніше вручну увімкне (enabled=1) приховану кнопку через Редактор
    # кнопок — мітка 'resolved' вже стоїть, тож жоден майбутній запуск
    # застосунку більше НІКОЛИ не поверне її назад у enabled=0.
    _STANDARD_MENU_ROOT_MIGRATION_KEYS = frozenset(
        {"income_form", "sale_form", "antiseptic_form", "writeoff_form", "data_browser_form"}
    )

    def _apply_standard_menu_policy(self):
        with self.conn:
            # Міграція старого стану: якщо старий глобальний прапорець уже
            # спрацював раніше (income/sale/writeoff вже сховані, можливо
            # адмін уже й повернув котрийсь назад вручну) — переносимо ці 3
            # ключі під нові пер-ключові мітки БЕЗ повторного enabled=0,
            # щоб не затерти можливе ручне рішення адміністратора.
            legacy_flag_applied = self.conn.execute(
                "SELECT 1 FROM app_meta WHERE key = 'legacy_non_form_buttons_hidden'"
            ).fetchone()
            if legacy_flag_applied:
                for migration_key in ("income", "sale", "writeoff"):
                    self.conn.execute(
                        "INSERT OR IGNORE INTO app_meta (key, value) VALUES (?, '1')",
                        (f"standard_menu_resolved:{migration_key}",),
                    )
            for entry in BUILTIN_MIGRATED_CUSTOM_BUTTONS:
                if entry.get("parent_migration_key") is not None:
                    continue
                migration_key = entry["migration_key"]
                if migration_key in self._STANDARD_MENU_ROOT_MIGRATION_KEYS:
                    continue
                meta_key = f"standard_menu_resolved:{migration_key}"
                already_resolved = self.conn.execute(
                    "SELECT 1 FROM app_meta WHERE key = ?", (meta_key,)
                ).fetchone()
                if already_resolved:
                    continue
                self.conn.execute(
                    "UPDATE custom_menu_buttons SET enabled = 0 WHERE migration_key = ?",
                    (migration_key,),
                )
                self.conn.execute(
                    "INSERT INTO app_meta (key, value) VALUES (?, '1')", (meta_key,)
                )

    # "Хмарна істина" (standard_menu_cloud.py) — Задача користувача
    # (2026-08-18): звірити локальний enabled-стан 11 кореневих мігрованих
    # кнопок з хмарою (OneDrive) при старті client_app.py, і оновлювати
    # хмару при будь-якій локальній зміні. Ці два методи — ЛИШЕ читання/
    # запис ЛОКАЛЬНОГО стану, самé рішення "коли звіряти з хмарою" (і сама
    # хмара) навмисно НЕ тут — див. коментар на початку standard_menu_cloud.py
    # чому це має викликатись виключно з client_app.py, не з __init__ вище.
    def get_standard_menu_root_keys(self):
        return [
            entry["migration_key"]
            for entry in BUILTIN_MIGRATED_CUSTOM_BUTTONS
            if entry.get("parent_migration_key") is None
        ]

    def get_standard_menu_state(self):
        keys = self.get_standard_menu_root_keys()
        placeholders = ",".join("?" for _ in keys)
        rows = self.conn.execute(
            f"SELECT migration_key, enabled FROM custom_menu_buttons WHERE migration_key IN ({placeholders})",
            keys,
        ).fetchall()
        return {migration_key: bool(enabled) for migration_key, enabled in rows}

    def apply_standard_menu_state(self, state):
        with self.conn:
            for migration_key, enabled in state.items():
                self.conn.execute(
                    "UPDATE custom_menu_buttons SET enabled = ? WHERE migration_key = ?",
                    (1 if enabled else 0, migration_key),
                )

    # --- Імпорт з Excel + CRUD над рядками листів (sheet_rows) ---
    def import_workbook(self, workbook, read_only_sheets):
        self._reset_file_scoped_state_if_source_changed()
        for worksheet in workbook.worksheets:
            self.import_sheet(worksheet, worksheet.title in read_only_sheets)

    # Задача користувача (2026-08-14): "ніяких перенесень. всі таблиці і
    # дані з таблиць ЛИШЕ ПЕРСОНАЛЬНІ і не мають ЖОДНІ дані бути
    # пов'язаними між таблицями, жодні" — приводом стало те, що номер
    # документа ("Приход №28") діставався з наскрізного лічильника
    # (document_counters), який не знав, що підключений файл змінився на
    # зовсім інший, порожній. Тут — єдина точка (import_workbook, спільна
    # для gui.py й client_app.py), де перевіряється, чи excel_source
    # вказує на ІНШИЙ файл, ніж минулого разу; якщо так — стан, похідний
    # від вмісту ПОПЕРЕДНЬОГО файлу, скидається, щоб не протікав у новий:
    #   - document_counters (нумерація Приход/Продажа/Списание/Услуга)
    #   - operation_recent_uses ("останні використані" підказки форми)
    #   - client_name_aliases (вивчені виправлення одруків клієнтів)
    #   - stock_movements (журнал приход/продажа/списание/антисептирование,
    #     на якому побудований репорт бота "скільки прийшло за період") —
    #     спершу лишав окремо (реальна історія, не кеш), але Задача
    #     користувача (2026-08-14, одразу після пояснення що це таке):
    #     "а, так, його скидаємо" — підтверджено явно, тож теж входить.
    # Перший запуск (ще немає збереженого excel_source_identity) НІЧОГО
    # не скидає — лише запам'ятовує поточний файл, щоб не знищити вже
    # накопичені реальні дані існуючих встановлень одразу після оновлення.
    def _reset_file_scoped_state_if_source_changed(self):
        current_identity = excel_source.current_source_identity()
        row = self.conn.execute(
            "SELECT value FROM app_meta WHERE key = 'excel_source_identity'"
        ).fetchone()
        stored_identity = row[0] if row else None
        if stored_identity is not None and stored_identity != current_identity:
            with self.conn:
                self.conn.execute("DELETE FROM document_counters")
                self.conn.execute("DELETE FROM operation_recent_uses")
                self.conn.execute("DELETE FROM client_name_aliases")
                self.conn.execute("DELETE FROM stock_movements")
        if stored_identity != current_identity:
            with self.conn:
                self.conn.execute(
                    """
                    INSERT INTO app_meta (key, value) VALUES ('excel_source_identity', ?)
                    ON CONFLICT(key) DO UPDATE SET value = excluded.value
                    """,
                    (current_identity,),
                )

    # Деякі листи (АНТИСЕПТИРОВАНИЕ) мають зверху інформаційний блок
    # (назва/зведення KPI/примітка) ПЕРЕД реальними заголовками колонок —
    # звичайний імпорт (заголовок завжди перший рядок) для них зробив би
    # заголовком просто текст назви. Для таких листів шукаємо реальний
    # заголовок за вмістом першої клітинки ("Дата"), а не за індексом 0.
    # Для всіх інших листів поведінка НЕ змінюється (header_index=0, як і
    # раніше).
    _SHEETS_WITH_HEADER_BLOCK = {"АНТИСЕПТИРОВАНИЕ"}

    def import_sheet(self, worksheet, read_only=False):
        rows = self._read_worksheet_rows(worksheet)
        header_index = 0
        if worksheet.title in self._SHEETS_WITH_HEADER_BLOCK:
            for index, row in enumerate(rows):
                if row and row[0] == "Дата":
                    header_index = index
                    break
        headers = list(rows[header_index]) if rows else []
        data_rows = [list(row) for row in rows[header_index + 1 :]]
        now = datetime.now().isoformat(timespec="seconds")

        with self.conn:
            self.conn.execute("DELETE FROM sheet_rows WHERE sheet_name = ?", (worksheet.title,))
            self.conn.execute(
                """
                INSERT INTO sheet_meta (sheet_name, headers_json, read_only, imported_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(sheet_name) DO UPDATE SET
                    headers_json = excluded.headers_json,
                    read_only = excluded.read_only,
                    imported_at = excluded.imported_at
                """,
                (
                    worksheet.title,
                    _serialize_row(headers),
                    1 if read_only else 0,
                    now,
                ),
            )
            if worksheet.title == "СКЛАД":
                # Окремі INSERT (не executemany), щоб отримати lastrowid і
                # одразу синхронізувати warehouse_items для кожного рядка.
                for index, row in enumerate(data_rows, start=1):
                    values = list(row)
                    cursor = self.conn.execute(
                        """
                        INSERT INTO sheet_rows (sheet_name, position, values_json, updated_at)
                        VALUES (?, ?, ?, ?)
                        """,
                        (worksheet.title, index, _serialize_row(values), now),
                    )
                    self._sync_warehouse_item(cursor.lastrowid, worksheet.title, values)
            else:
                self.conn.executemany(
                    """
                    INSERT INTO sheet_rows (sheet_name, position, values_json, updated_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    [
                        (worksheet.title, index, _serialize_row(row), now)
                        for index, row in enumerate(data_rows, start=1)
                    ],
                )

    def _read_worksheet_rows(self, worksheet):
        max_col = worksheet.max_column
        last_row = 0

        for row_number in range(worksheet.max_row, 0, -1):
            has_data = any(
                worksheet.cell(row=row_number, column=col).value is not None
                for col in range(1, max_col + 1)
            )
            if has_data:
                last_row = row_number
                break

        if not last_row:
            return []

        return [
            row
            for row in worksheet.iter_rows(min_row=1, max_row=last_row, values_only=True)
            if not all(cell is None for cell in row)
        ]

    def sheet_names(self):
        cursor = self.conn.execute(
            "SELECT sheet_name FROM sheet_meta ORDER BY rowid"
        )
        return [row[0] for row in cursor.fetchall()]

    def get_headers(self, sheet_name):
        cursor = self.conn.execute(
            "SELECT headers_json FROM sheet_meta WHERE sheet_name = ?",
            (sheet_name,),
        )
        row = cursor.fetchone()
        return _deserialize_row(row[0]) if row else []

    def is_read_only(self, sheet_name):
        cursor = self.conn.execute(
            "SELECT read_only FROM sheet_meta WHERE sheet_name = ?",
            (sheet_name,),
        )
        row = cursor.fetchone()
        return bool(row[0]) if row else True

    def count_rows(self, sheet_name):
        cursor = self.conn.execute(
            "SELECT COUNT(*) FROM sheet_rows WHERE sheet_name = ?",
            (sheet_name,),
        )
        return cursor.fetchone()[0]

    def fetch_rows(self, sheet_name, limit, offset=0):
        cursor = self.conn.execute(
            """
            SELECT id, values_json
            FROM sheet_rows
            WHERE sheet_name = ?
            ORDER BY position, id
            LIMIT ? OFFSET ?
            """,
            (sheet_name, limit, offset),
        )
        return [(row_id, _deserialize_row(values_json)) for row_id, values_json in cursor.fetchall()]

    def fetch_all_rows(self, sheet_name):
        cursor = self.conn.execute(
            """
            SELECT values_json
            FROM sheet_rows
            WHERE sheet_name = ?
            ORDER BY position, id
            """,
            (sheet_name,),
        )
        return [_deserialize_row(row[0]) for row in cursor.fetchall()]

    # Задача користувача (2026-08-14): "всі ці колонки мають бути з
    # фільтрами в таблиці" - фільтрація gui.py-таблиці по стовпцю
    # відбувається в Python (структура листа динамічна, values_json - не
    # окремі SQL-колонки), тож їй потрібен ПОВНИЙ набір рядків ІЗ row_id
    # (для Treeview iid/подальшого редагування), на відміну від
    # fetch_all_rows вище (уже має 4 виклики, які чекають лише значення,
    # без id - тут окремий метод, а не зміна існуючого).
    def fetch_all_rows_with_ids(self, sheet_name):
        cursor = self.conn.execute(
            """
            SELECT id, values_json
            FROM sheet_rows
            WHERE sheet_name = ?
            ORDER BY position, id
            """,
            (sheet_name,),
        )
        return [(row_id, _deserialize_row(values_json)) for row_id, values_json in cursor.fetchall()]

    def get_row(self, row_id):
        cursor = self.conn.execute(
            "SELECT values_json FROM sheet_rows WHERE id = ?",
            (row_id,),
        )
        row = cursor.fetchone()
        return _deserialize_row(row[0]) if row else []

    def update_row(self, row_id, values):
        # was_in_transaction: той самий прийом, що й у add_stock_movement —
        # якщо цей метод викликано з середини вже відкритої self.conn: (як у
        # apply_sale_operation/apply_income_operation), не комітимо самі,
        # інакше передчасний коміт розірве атомарність всієї операції.
        was_in_transaction = self.conn.in_transaction
        row = self.conn.execute(
            "SELECT sheet_name FROM sheet_rows WHERE id = ?", (row_id,)
        ).fetchone()
        self.conn.execute(
            """
            UPDATE sheet_rows
            SET values_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (_serialize_row(values), datetime.now().isoformat(timespec="seconds"), row_id),
        )
        if row:
            self._sync_warehouse_item(row_id, row[0], values)
        if not was_in_transaction:
            self.conn.commit()

    def add_row(self, sheet_name, values):
        was_in_transaction = self.conn.in_transaction
        cursor = self.conn.execute(
            "SELECT COALESCE(MAX(position), 0) + 1 FROM sheet_rows WHERE sheet_name = ?",
            (sheet_name,),
        )
        position = cursor.fetchone()[0]
        now = datetime.now().isoformat(timespec="seconds")
        insert_cursor = self.conn.execute(
            """
            INSERT INTO sheet_rows (sheet_name, position, values_json, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (sheet_name, position, _serialize_row(values), now),
        )
        row_id = insert_cursor.lastrowid
        self._sync_warehouse_item(row_id, sheet_name, values)
        if not was_in_transaction:
            self.conn.commit()
        return row_id

    # Тримає warehouse_items (структуровані, індексовані колонки) синхронними
    # з sheet_rows (JSON, для GUI/Excel) для листа СКЛАД — єдина точка, де
    # це робиться, щоб жоден шлях запису (бот чи ручне редагування в GUI) не
    # забув оновити один із двох. Для будь-якого іншого листа — нічого не робить.
    def _sync_warehouse_item(self, sheet_row_id, sheet_name, values):
        if sheet_name != "СКЛАД":
            return
        columns = warehouse_columns(self.get_headers(sheet_name))

        def text_value(key):
            index = columns.get(key)
            if index is None or index >= len(values):
                return None
            value = values[index]
            return None if value in (None, "") else str(value)

        def number_value(key):
            index = columns.get(key)
            if index is None or index >= len(values):
                return None
            value = values[index]
            return None if value in (None, "") else _number_value(value)

        if not text_value("product"):
            # Порожній/шаблонний рядок складу (напр. службовий маркер типу
            # "ПередЗнімком" у "Порода" без жодного товару) - не заводимо
            # фіктивну позицію в warehouse_items і прибираємо застарілий
            # запис, якщо рядок раніше мав реальні дані і потім спорожнів.
            # Той самий клас багу, що вже зламав звіт антисептирования
            # (порожні рядки листа трактувались як реальні позиції).
            self.conn.execute("DELETE FROM warehouse_items WHERE sheet_row_id = ?", (sheet_row_id,))
            return

        self.conn.execute(
            """
            INSERT INTO warehouse_items (
                sheet_row_id, sku, product, breed, condition,
                thickness, width, length, unit,
                income_qty, income_volume, income_area, income_linear,
                sold_qty, sold_volume, sold_area, sold_linear,
                balance_qty, balance_volume, balance_area, balance_linear
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet_row_id) DO UPDATE SET
                sku = excluded.sku,
                product = excluded.product,
                breed = excluded.breed,
                condition = excluded.condition,
                thickness = excluded.thickness,
                width = excluded.width,
                length = excluded.length,
                unit = excluded.unit,
                income_qty = excluded.income_qty,
                income_volume = excluded.income_volume,
                income_area = excluded.income_area,
                income_linear = excluded.income_linear,
                sold_qty = excluded.sold_qty,
                sold_volume = excluded.sold_volume,
                sold_area = excluded.sold_area,
                sold_linear = excluded.sold_linear,
                balance_qty = excluded.balance_qty,
                balance_volume = excluded.balance_volume,
                balance_area = excluded.balance_area,
                balance_linear = excluded.balance_linear
            """,
            (
                sheet_row_id,
                text_value("sku"),
                # Нормалізуємо категорію товару лише в цьому індексі для пошуку
                # (не чіпаємо оригінальний запис у sheet_rows/Excel) — "доска AD"
                # і "Доска AD" в реальних даних мають зіставлятись в одну категорію.
                normalize_product_category(text_value("product")),
                text_value("breed"),
                text_value("condition"),
                number_value("thickness"),
                number_value("width"),
                number_value("length"),
                text_value("unit"),
                number_value("income_qty"),
                number_value("income_volume"),
                number_value("income_area"),
                number_value("income_linear"),
                number_value("sold_qty"),
                number_value("sold_volume"),
                number_value("sold_area"),
                number_value("sold_linear"),
                number_value("balance_qty"),
                number_value("balance_volume"),
                number_value("balance_area"),
                number_value("balance_linear"),
            ),
        )

    def delete_rows(self, row_ids):
        if not row_ids:
            return
        placeholders = ",".join("?" for _ in row_ids)
        with self.conn:
            self.conn.execute(
                f"DELETE FROM sheet_rows WHERE id IN ({placeholders})",
                tuple(row_ids),
            )

    # --- Команди бота, користувачі, журнал дій, вподобання Telegram-користувача ---
    def list_commands(self):
        cursor = self.conn.execute(
            """
            SELECT id, code, title, COALESCE(description, ''), enabled
            FROM bot_commands
            ORDER BY title COLLATE NOCASE, id
            """
        )
        return cursor.fetchall()

    def find_command_code_by_phrase(self, phrase):
        normalized_phrase = _normalize_phrase(phrase)
        if not normalized_phrase:
            return None

        cursor = self.conn.execute(
            """
            SELECT c.code
            FROM bot_commands c
            LEFT JOIN bot_command_aliases a ON a.command_id = c.id
            WHERE c.enabled = 1
              AND (
                  lower(c.code) = ?
                  OR lower(c.title) = ?
                  OR a.phrase = ?
              )
            LIMIT 1
            """,
            (normalized_phrase, normalized_phrase, normalized_phrase),
        )
        row = cursor.fetchone()
        return row[0] if row else None

    def find_command_code_in_text(self, text):
        exact = self.find_command_code_by_phrase(text)
        if exact:
            return exact

        normalized_text = f" {_normalize_phrase(text)} "
        if not normalized_text.strip():
            return None

        cursor = self.conn.execute(
            """
            SELECT c.code, a.phrase
            FROM bot_commands c
            JOIN bot_command_aliases a ON a.command_id = c.id
            WHERE c.enabled = 1
            ORDER BY length(a.phrase) DESC
            """
        )
        for code, phrase in cursor.fetchall():
            if f" {phrase} " in normalized_text:
                return code
        return None

    def add_command(self, code, title, description=""):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_commands (code, title, description, enabled, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (code, title, description, now, now),
            )

    def delete_command(self, command_id):
        with self.conn:
            self.conn.execute("DELETE FROM bot_commands WHERE id = ?", (command_id,))

    def list_command_aliases(self, command_id):
        cursor = self.conn.execute(
            """
            SELECT id, phrase
            FROM bot_command_aliases
            WHERE command_id = ?
            ORDER BY phrase COLLATE NOCASE, id
            """,
            (command_id,),
        )
        return cursor.fetchall()

    def add_command_alias(self, command_id, phrase):
        normalized_phrase = _normalize_phrase(phrase)
        if not normalized_phrase:
            return
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_command_aliases (command_id, phrase, created_at)
                VALUES (?, ?, ?)
                """,
                (command_id, normalized_phrase, now),
            )

    def update_command_alias(self, alias_id, phrase):
        normalized_phrase = _normalize_phrase(phrase)
        if not normalized_phrase:
            return
        with self.conn:
            self.conn.execute(
                "UPDATE bot_command_aliases SET phrase = ? WHERE id = ?",
                (normalized_phrase, alias_id),
            )

    def delete_command_alias(self, alias_id):
        with self.conn:
            self.conn.execute("DELETE FROM bot_command_aliases WHERE id = ?", (alias_id,))

    # --- Кастомні кнопки бота (Редактор кнопок, GUI) ---
    # Дерево довільної глибини й ширини (Задача користувача: до батьківської
    # кнопки можна додати скільки завгодно дочірніх, і так само далі по
    # кожній гілці). parent_id=None -> кореневі кнопки (додаються до
    # головного меню бота). include_disabled=True — для GUI (показує ВСІ
    # кнопки незалежно від enabled); бот викликає з include_disabled=False
    # (замовчування) — вимкнені кнопки в чаті не показуються.
    def list_custom_buttons(self, parent_id=None, include_disabled=False):
        query = """
            SELECT id, label, message_text, action_code, section, enabled, COALESCE(layout, 'full'), operation_id
            FROM custom_menu_buttons
            WHERE parent_id IS ?
        """
        if not include_disabled:
            query += " AND enabled = 1"
        query += " ORDER BY position, id"
        cursor = self.conn.execute(query, (parent_id,))
        return cursor.fetchall()

    # Задача користувача (2026-08-17): "редактор кнопок зроби синхронним" -
    # gui.py.  тепер редагує ЖИВЕ дерево client_app.py через тунель
    # (webapp_server._handle_remote_custom_buttons), а не власну окрему й
    # порожню локальну копію (той самий факт, що вже пояснено для Персонал/
    # Журналів). Один запит за ВЕСЬ пласким деревом одразу (parent_id
    # включно) - клієнт сам групує в ієрархію в пам'яті, замість N
    # мережевих походів через тунель на кожен рівень вкладеності (як був би
    # результат простого повторного виклику list_custom_buttons по кожному
    # parent_id окремо).
    def list_all_custom_buttons(self, include_disabled=True):
        query = """
            SELECT id, parent_id, label, message_text, action_code, section, enabled,
                   COALESCE(layout, 'full'), operation_id
            FROM custom_menu_buttons
        """
        if not include_disabled:
            query += " WHERE enabled = 1"
        query += " ORDER BY parent_id, position, id"
        cursor = self.conn.execute(query)
        return cursor.fetchall()

    def get_custom_button(self, node_id):
        cursor = self.conn.execute(
            """
            SELECT id, parent_id, label, message_text, action_code, section, enabled,
                   COALESCE(layout, 'full'), operation_id
            FROM custom_menu_buttons
            WHERE id = ?
            """,
            (node_id,),
        )
        return cursor.fetchone()

    # Надійний пошук мігрованої кнопки за migration_key (не за міткою — та
    # могла бути перейменована адміністратором). Використовується ботом,
    # щоб "Назад"/легасі-фрази-хардкод (напр. _is_data_menu_request) заходили
    # у ПОТОЧНИЙ стан вузла в дереві (з урахуванням перейменувань/розміру/
    # порядку), а не в застарілу хардкоджену клавіатуру.
    def get_custom_button_by_migration_key(self, migration_key):
        cursor = self.conn.execute(
            """
            SELECT id, parent_id, label, message_text, action_code, section, enabled,
                   COALESCE(layout, 'full'), operation_id
            FROM custom_menu_buttons
            WHERE migration_key = ?
            """,
            (migration_key,),
        )
        return cursor.fetchone()

    # Задача користувача: поки старі (не-"форма") кнопки приховані
    # (_hide_legacy_non_form_operation_buttons_once), набірний текст
    # ("приход ...", "продажа ...") теж не повинен заводити в старий
    # покроковий флоу — інакше "не має бути шляху" лишається неправдою для
    # тих, хто просто друкує команду замість натискання кнопки.
    def is_legacy_non_form_button_hidden(self, migration_key):
        row = self.get_custom_button_by_migration_key(migration_key)
        return bool(row) and not row[6]

    def custom_button_label_collides(self, label):
        return _normalize_phrase(label) in _RESERVED_BUTTON_LABELS or bool(
            self.find_command_code_by_phrase(label)
        )

    # Задача користувача: перевизначити текст, який бот пише в чат на вході
    # в дію. default передає ВИКЛИКАЧ (telegram_dialog.py) — тут немає
    # жодного хардкодженого тексту, лише зберігання; якщо рядка для key ще
    # немає (адміністратор нічого не міняв), повертається саме default.
    def get_message_template(self, key, default):
        row = self.conn.execute(
            "SELECT text FROM bot_message_templates WHERE key = ?", (key,)
        ).fetchone()
        return row[0] if row else default

    def set_message_template(self, key, text):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_message_templates (key, text, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET text = excluded.text, updated_at = excluded.updated_at
                """,
                (key, text, now),
            )

    # "Скинути до типового" (Задача користувача, GUI) — прибирає
    # перевизначення, get_message_template знову повертатиме default.
    def reset_message_template(self, key):
        with self.conn:
            self.conn.execute("DELETE FROM bot_message_templates WHERE key = ?", (key,))

    def is_message_template_customized(self, key):
        row = self.conn.execute(
            "SELECT 1 FROM bot_message_templates WHERE key = ?", (key,)
        ).fetchone()
        return bool(row)

    # Кнопка, перенесена з хардкоду (BUILTIN_MIGRATED_CUSTOM_BUTTONS) —
    # тести, що чистять custom_menu_buttons для перевірки "порожнього"
    # головного меню, мають ЗБЕРЕГТИ такі кнопки (вони тепер частина
    # звичайного меню, а не тестові дані).
    def is_custom_button_migrated(self, node_id):
        row = self.conn.execute(
            "SELECT migration_key FROM custom_menu_buttons WHERE id = ?", (node_id,)
        ).fetchone()
        return bool(row and row[0])

    def add_custom_button(
        self, label, message_text="", action_code=None, section=None, parent_id=None, layout="full",
        operation_id=None,
    ):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            position = self.conn.execute(
                """
                SELECT COALESCE(MAX(position), -1) + 1
                FROM custom_menu_buttons
                WHERE parent_id IS ?
                """,
                (parent_id,),
            ).fetchone()[0]
            cursor = self.conn.execute(
                """
                INSERT INTO custom_menu_buttons
                    (parent_id, section, label, message_text, action_code, position, layout, operation_id,
                     enabled, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (parent_id, section, label, message_text, action_code, position, layout, operation_id, now, now),
            )
            return cursor.lastrowid

    # action_code й operation_id — взаємовиключні на рівні GUI-форми (радіо-
    # вибір "Дія" / "Пряме посилання на дію з 'Дії'"), тут просто пишуться
    # обидва як прийшли — виконавець (_enter_custom_button_node,
    # telegram_dialog.py) перевіряє operation_id ПЕРШИМ, тож навіть якщо
    # обидва колись опиняться заповненими, поведінка лишається однозначною.
    def update_custom_button(self, node_id, label, message_text, action_code, layout="full", operation_id=None):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE custom_menu_buttons
                SET label = ?, message_text = ?, action_code = ?, layout = ?, operation_id = ?, updated_at = ?
                WHERE id = ?
                """,
                (label, message_text, action_code, layout, operation_id, now, node_id),
            )

    # Скільки кнопок (усіх нащадків, не лише прямих дітей) видалиться разом
    # із node_id — показується у попередженні перед каскадним видаленням
    # гілки (Задача користувача: підтвердження видалення завжди, і явне
    # попередження, якщо видаляється ціла гілка). Обхід у ширину в Python,
    # а не рекурсивний SQL — дерево завжди невелике.
    def count_custom_button_descendants(self, node_id):
        total = 0
        frontier = [node_id]
        while frontier:
            next_frontier = []
            for current_id in frontier:
                children = self.conn.execute(
                    "SELECT id FROM custom_menu_buttons WHERE parent_id = ?",
                    (current_id,),
                ).fetchall()
                next_frontier.extend(row[0] for row in children)
            total += len(next_frontier)
            frontier = next_frontier
        return total

    # Переставити кнопку на позицію сусіда серед "братів" (той самий
    # parent_id) — Задача користувача: змога поставити кнопку у будь-який
    # слот, не лише в кінець. Просто міняє значення position місцями (не
    # арифметично рахує position+-1), тож працює навіть якщо позиції колись
    # стали не суцільними. На межі списку (перший/останній) — нічого не
    # робить, а не падає.
    def move_custom_button(self, node_id, direction):
        row = self.get_custom_button(node_id)
        if not row:
            return
        parent_id = row[1]
        siblings = self.list_custom_buttons(parent_id, include_disabled=True)
        ids_in_order = [sibling_row[0] for sibling_row in siblings]
        if node_id not in ids_in_order:
            return
        index = ids_in_order.index(node_id)
        if direction == "up" and index > 0:
            swap_with_id = ids_in_order[index - 1]
        elif direction == "down" and index < len(ids_in_order) - 1:
            swap_with_id = ids_in_order[index + 1]
        else:
            return
        with self.conn:
            position_a = self.conn.execute(
                "SELECT position FROM custom_menu_buttons WHERE id = ?", (node_id,)
            ).fetchone()[0]
            position_b = self.conn.execute(
                "SELECT position FROM custom_menu_buttons WHERE id = ?", (swap_with_id,)
            ).fetchone()[0]
            self.conn.execute(
                "UPDATE custom_menu_buttons SET position = ? WHERE id = ?", (position_b, node_id)
            )
            self.conn.execute(
                "UPDATE custom_menu_buttons SET position = ? WHERE id = ?", (position_a, swap_with_id)
            )

    # Прямий вибір слота (Задача користувача: випадаючий список позиції у
    # формі додавання/редагування замість стрілочок ↑/↓). new_index — 0-based
    # позиція серед "братів" (parent_id), рахуючи СЕРЕД РЕШТИ (без самого
    # node_id) — те саме, що бачить користувач у випадаючому списку форми.
    # Перезаписує position УСІХ братів підряд 0..N-1, а не лише двох (як
    # move_custom_button), тому дає перейти в будь-який слот за один виклик.
    def set_custom_button_position(self, node_id, new_index):
        row = self.get_custom_button(node_id)
        if not row:
            return
        parent_id = row[1]
        siblings = self.list_custom_buttons(parent_id, include_disabled=True)
        ids_in_order = [sibling_row[0] for sibling_row in siblings]
        if node_id in ids_in_order:
            ids_in_order.remove(node_id)
        new_index = max(0, min(new_index, len(ids_in_order)))
        ids_in_order.insert(new_index, node_id)
        with self.conn:
            for position, sibling_id in enumerate(ids_in_order):
                self.conn.execute(
                    "UPDATE custom_menu_buttons SET position = ? WHERE id = ?",
                    (position, sibling_id),
                )

    # Задача користувача: видалена кнопка (особливо вбудована/мігрована -
    # ПРИХОД (форма) і т.д.) не має воскресати на наступному запуску
    # програми. Рядок migration_key ЩЕЗАЄ разом із самим видаленням (тут
    # реальний DELETE, FOREIGN KEY ON DELETE CASCADE забирає й дітей), тож
    # єдиний спосіб "запам'ятати" намір - надгробок в окремій таблиці
    # (deleted_builtin_migrations), звіряється в
    # _seed_builtin_migrated_custom_buttons нижче. Кастомні (не вбудовані)
    # кнопки не мають migration_key - для них цей блок просто нічого не
    # робить.
    def delete_custom_button(self, node_id):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            row = self.conn.execute(
                "SELECT migration_key FROM custom_menu_buttons WHERE id = ?", (node_id,)
            ).fetchone()
            if row and row[0]:
                self.conn.execute(
                    "INSERT OR IGNORE INTO deleted_builtin_migrations (migration_key, deleted_at) VALUES (?, ?)",
                    (row[0], now),
                )
            self.conn.execute("DELETE FROM custom_menu_buttons WHERE id = ?", (node_id,))

    # --- Крок 3+ "Дії": bot_operations / bot_operation_fields /
    # bot_operation_field_columns (дія-категорія -> поля-запити -> прив'язки
    # вкладка+колонка) ---

    def list_operations(self, parent_action_code=None, include_disabled=False):
        query = (
            "SELECT id, code, kind, requires_row_identity, label, parent_action_code, "
            "prefill_json, position, enabled, builtin_key FROM bot_operations"
        )
        params = ()
        clauses = []
        if parent_action_code is not None:
            clauses.append("parent_action_code = ?")
            params = (parent_action_code,)
        if not include_disabled:
            clauses.append("enabled = 1")
        if clauses:
            query += " WHERE " + " AND ".join(clauses)
        query += " ORDER BY position, id"
        return self.conn.execute(query, params).fetchall()

    def get_operation(self, operation_id):
        return self.conn.execute(
            "SELECT id, code, kind, requires_row_identity, label, parent_action_code, "
            "prefill_json, position, enabled, builtin_key FROM bot_operations WHERE id = ?",
            (operation_id,),
        ).fetchone()

    def get_operation_by_code(self, code):
        return self.conn.execute(
            "SELECT id, code, kind, requires_row_identity, label, parent_action_code, "
            "prefill_json, position, enabled, builtin_key FROM bot_operations WHERE code = ?",
            (code,),
        ).fetchone()

    def delete_operation(self, operation_id):
        with self.conn:
            # Аудит коду: custom_menu_buttons.operation_id немає SQLite FK
            # (навмисно, схема це документує) — без цього очищення видалена
            # дія лишала б "осиротіле" посилання: кнопка з прямим посиланням
            # мовчки перетворювалась би на no-op (_start_operation_leaf
            # повертає None для operation_id, що не існує).
            self.conn.execute(
                "UPDATE custom_menu_buttons SET operation_id = NULL WHERE operation_id = ?",
                (operation_id,),
            )
            self.conn.execute("DELETE FROM bot_operations WHERE id = ?", (operation_id,))

    # Крок "Дії" remote-sync (2026-08-18): той самий bulk-read принцип, що
    # вже й list_all_custom_buttons — client_app.py віддає ВЕСЬ дерево
    # (операції+поля-запити+прив'язки) ОДНИМ запитом, gui.py сам групує в
    # пам'яті (той самий "домашня программа лише ТЯГНЕ живі дані" принцип,
    # що вже й у Редакторі кнопок/Способах оплати).
    def list_all_operations_tree(self):
        operations = self.conn.execute(
            "SELECT id, code, kind, requires_row_identity, label, parent_action_code, "
            "prefill_json, position, enabled, builtin_key FROM bot_operations "
            "ORDER BY parent_action_code, position, id"
        ).fetchall()
        fields = self.conn.execute(
            "SELECT id, operation_id, field_key, label, is_identity, position, enabled, builtin_key "
            "FROM bot_operation_fields ORDER BY operation_id, position, id"
        ).fetchall()
        columns = self.conn.execute(
            "SELECT id, operation_field_id, sheet, column_key, marker, write_mode, position, builtin_key "
            "FROM bot_operation_field_columns ORDER BY operation_field_id, position, id"
        ).fetchall()
        return {"operations": operations, "fields": fields, "columns": columns}

    # --- Шаблони мега-форм (Приход/Реализация/Списание) ---

    # Задача користувача: "5 занадто багато місця займає" - 3 рядки на
    # панель (шаблони і недавні окремо).
    _OPERATION_TEMPLATE_LIMIT = 3

    def add_operation_template(
        self, kind, category_operation_id, breed=None, thickness=None, width=None,
        length=None, client=None, address=None, payment_method=None,
    ):
        # Задача користувача: "мають бути лише унікальні шаблони. однакових
        # там не має бути" - унікальність рахується за тим, що РЕАЛЬНО
        # показано в рядку панелі (категорія/порода/розмір/спосіб оплати),
        # а НЕ за клієнтом/адресою - вони в списку взагалі не відображаються,
        # тож два шаблони, що відрізняються лише невидимим клієнтом,
        # виглядали б як дублікат. "IS" (не "=") коректно порівнює й NULL.
        with self.conn:
            existing = self.conn.execute(
                """
                SELECT id FROM operation_templates
                WHERE kind = ? AND category_operation_id = ?
                  AND breed IS ? AND thickness IS ? AND width IS ? AND length IS ?
                  AND payment_method IS ?
                """,
                (kind, category_operation_id, breed, thickness, width, length, payment_method),
            ).fetchone()
            if existing is not None:
                return
            now = datetime.now().isoformat(timespec="seconds")
            self.conn.execute(
                """
                INSERT INTO operation_templates
                    (kind, category_operation_id, breed, thickness, width, length,
                     client, address, payment_method, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (kind, category_operation_id, breed, thickness, width, length, client, address, payment_method, now),
            )
            # Задача користувача: рівно 5 рядків на панель — найстаріший
            # шаблон цього kind мовчки поступається місцем новому 6-му,
            # без окремого видалення користувачем (той самий "FIFO-вітрина"
            # принцип, що й у db-знімків/резервних копій).
            excess_ids = [
                row[0]
                for row in self.conn.execute(
                    "SELECT id FROM operation_templates WHERE kind = ? ORDER BY created_at DESC, id DESC",
                    (kind,),
                ).fetchall()[self._OPERATION_TEMPLATE_LIMIT:]
            ]
            if excess_ids:
                placeholders = ",".join("?" for _ in excess_ids)
                self.conn.execute(f"DELETE FROM operation_templates WHERE id IN ({placeholders})", excess_ids)

    def list_operation_templates(self, kind, limit=_OPERATION_TEMPLATE_LIMIT):
        return self.conn.execute(
            """
            SELECT id, category_operation_id, breed, thickness, width, length, client, address, payment_method
            FROM operation_templates WHERE kind = ? ORDER BY created_at DESC, id DESC LIMIT ?
            """,
            (kind, limit),
        ).fetchall()

    def delete_operation_template(self, template_id):
        with self.conn:
            self.conn.execute("DELETE FROM operation_templates WHERE id = ?", (template_id,))

    # "5 останніх створених" — жива історія подань мега-форми, наповнюється
    # автоматично (record_operation_use), не користувачем. Дедуплікація —
    # у Python (recent_operation_uses), не в SQL: ключ поєднує кілька NULL-
    # придатних полів (client/payment_method відсутні для income/writeoff),
    # а SQLite DISTINCT трактує NULL непередбачувано для такого випадку.
    def record_operation_use(
        self, kind, category_operation_id, breed=None, thickness=None, width=None,
        length=None, client=None, address=None, payment_method=None,
    ):
        if category_operation_id is None:
            return
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO operation_recent_uses
                    (kind, category_operation_id, breed, thickness, width, length,
                     client, address, payment_method, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (kind, category_operation_id, breed, thickness, width, length, client, address, payment_method, now),
            )
            # Дешева гігієна диска — лишаємо із запасом (10x показуваного
            # ліміту) на дедуплікацію в recent_operation_uses, не тримаємо
            # необмежену історію.
            excess_ids = [
                row[0]
                for row in self.conn.execute(
                    "SELECT id FROM operation_recent_uses WHERE kind = ? ORDER BY created_at DESC, id DESC",
                    (kind,),
                ).fetchall()[self._OPERATION_TEMPLATE_LIMIT * 10:]
            ]
            if excess_ids:
                placeholders = ",".join("?" for _ in excess_ids)
                self.conn.execute(f"DELETE FROM operation_recent_uses WHERE id IN ({placeholders})", excess_ids)

    def recent_operation_uses(self, kind, limit=_OPERATION_TEMPLATE_LIMIT):
        rows = self.conn.execute(
            """
            SELECT id, category_operation_id, breed, thickness, width, length, client, address, payment_method
            FROM operation_recent_uses WHERE kind = ? ORDER BY created_at DESC, id DESC LIMIT ?
            """,
            (kind, self._OPERATION_TEMPLATE_LIMIT * 10),
        ).fetchall()
        # Дедуплікація за ВМІСТОМ (без id, який завжди унікальний) - id
        # лишається в поверненому рядку, щоб GUI/webapp могли видалити
        # САМЕ цей конкретний запис (найновіший серед дублікатів, бо
        # ORDER BY created_at DESC).
        seen = set()
        unique = []
        for row in rows:
            signature = row[1:]
            if signature in seen:
                continue
            seen.add(signature)
            unique.append(row)
            if len(unique) >= limit:
                break
        return unique

    def get_operation_template(self, template_id):
        return self.conn.execute(
            "SELECT id, kind, category_operation_id FROM operation_templates WHERE id = ?",
            (template_id,),
        ).fetchone()

    def get_operation_recent_use(self, use_id):
        return self.conn.execute(
            "SELECT id, kind, category_operation_id FROM operation_recent_uses WHERE id = ?",
            (use_id,),
        ).fetchone()

    def delete_operation_recent_use(self, use_id):
        with self.conn:
            self.conn.execute("DELETE FROM operation_recent_uses WHERE id = ?", (use_id,))

    def list_operation_fields(self, operation_id, include_disabled=False):
        query = (
            "SELECT id, operation_id, field_key, label, is_identity, position, enabled, builtin_key "
            "FROM bot_operation_fields WHERE operation_id = ?"
        )
        if not include_disabled:
            query += " AND enabled = 1"
        query += " ORDER BY position, id"
        return self.conn.execute(query, (operation_id,)).fetchall()

    def get_operation_field(self, field_id):
        return self.conn.execute(
            "SELECT id, operation_id, field_key, label, is_identity, position, enabled, builtin_key "
            "FROM bot_operation_fields WHERE id = ?",
            (field_id,),
        ).fetchone()

    def add_operation_field(self, operation_id, field_key, label, is_identity=False):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            position = self.conn.execute(
                "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operation_fields WHERE operation_id = ?",
                (operation_id,),
            ).fetchone()[0]
            cursor = self.conn.execute(
                """
                INSERT INTO bot_operation_fields
                    (operation_id, field_key, label, is_identity, position, enabled, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (operation_id, field_key, label, 1 if is_identity else 0, position, now, now),
            )
            return cursor.lastrowid

    # Критична знахідка аудиту 28.07.2026 (#1): видалення/вимкнення
    # ОСТАННЬОЇ generic-прив'язки до "СКЛАД" серед усіх полів операції
    # мовчки зупиняє облік залишку для цілої категорії —
    # execute_operation_write просто нічого не записує, без помилки й без
    # сигналу (is_identity нижче захищає ЛИШЕ Порода/Товщина/Ширина/
    # Довжина, а не quantity/measure). Перевіряємо лише для
    # requires_row_identity=1 (звичайний прихід/продаж, що веде облік
    # складу) — сервіс-дії (антисептирование) і звіти (kind='report')
    # мають requires_row_identity=0 і ЖОДНОЇ generic/СКЛАД-прив'язки
    # взагалі, перевірка їх не стосується. Інваріант — "лишиться БОДАЙ
    # ОДНА" (не "лишаться quantity І measure одночасно"): ОСБ-подібна
    # категорія (лише quantity, без measure) лишається валідною, видалення
    # measure там, де лишається quantity, і далі дозволено.
    def _operation_would_lose_balance_tracking(
        self, operation_id, exclude_field_id=None, exclude_column_id=None, owning_field_id=None
    ):
        operation = self.get_operation(operation_id)
        if operation is None or not operation[3]:
            return False
        if exclude_column_id is not None:
            # Свіжий пере-аудит (2026-08-02, New-Important #5): видаляється
            # ОДНА прив'язка, а САМЕ ПОЛЕ лишається активним (і далі
            # проситиме користувача цю величину) — тож перевіряємо ЛИШЕ
            # прив'язки ЦЬОГО САМОГО поля, а НЕ всієї операції. Приклад
            # реальної діри, яку це закриває: поле "quantity" з окремими
            # income_qty/balance_qty — видалення САМЕ balance_qty, коли
            # income_qty лишається, раніше хибно дозволялось, якщо
            # десь-інде в операції (напр. поле "measure") лишалась ІНША
            # balance_-прив'язка — хоча ця, зовсім інша величина (об'єм),
            # ніяк не рятує "Остаток, шт", який після цього назавжди
            # застигає, хоча бот і далі питає й пише "Приход, шт".
            for binding in self.list_operation_field_columns(owning_field_id):
                column_id, _operation_field_id, sheet, column_key, _marker, write_mode, _position, _builtin_key = binding
                if column_id == exclude_column_id:
                    continue
                if write_mode == "generic" and sheet == "СКЛАД" and column_key.startswith("balance_"):
                    return False
            return True
        # exclude_field_id (ціле поле йде геть) — лишається операційно-
        # широкий інваріант "лишиться БОДАЙ ОДНА (в ІНШОМУ полі)", той
        # самий, що вже дозволяє ОСБ-подібний стан (видалення "measure" з
        # доски, коли "quantity" лишається, лишається валідним) — на
        # відміну від видалення ОДНІЄЇ прив'язки вище, тут саме ПОЛЕ (і всі
        # його прив'язки) зникає разом, тож нічого "осиротілого" не лишається.
        for field in self.list_operation_fields(operation_id):
            if field[0] == exclude_field_id:
                continue
            for binding in self.list_operation_field_columns(field[0]):
                column_id, _operation_field_id, sheet, column_key, _marker, write_mode, _position, _builtin_key = binding
                if write_mode == "generic" and sheet == "СКЛАД" and column_key.startswith("balance_"):
                    return False
        return True

    # is_identity=1 (Порода/Товар/Товщина/Ширина/Довжина, коли операція
    # шукає рядок складу) — захищене поле: enabled=0 і видалення заборонені
    # програмно (не лише в GUI), бо _warehouse_row_matches (telegram_dialog.py)
    # порівнює саме ці значення БЕЗ винятку "якщо порожньо — пропустити".
    def update_operation_field(self, field_id, label, enabled=True):
        row = self.get_operation_field(field_id)
        if row is None:
            return
        is_identity = bool(row[4])
        if is_identity and not enabled:
            raise ValueError(
                "Це поле не можна вимкнути: бот використовує його, щоб знайти потрібний рядок на складі."
            )
        if not enabled and self._operation_would_lose_balance_tracking(row[1], exclude_field_id=field_id):
            raise ValueError(
                "Це поле не можна вимкнути: без нього бот перестане оновлювати залишок складу для цієї категорії."
            )
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                "UPDATE bot_operation_fields SET label = ?, enabled = ?, updated_at = ? WHERE id = ?",
                (label, 1 if enabled else 0, now, field_id),
            )

    def delete_operation_field(self, field_id):
        row = self.get_operation_field(field_id)
        if row is None:
            return
        if row[4]:
            raise ValueError(
                "Це поле не можна видалити: бот використовує його, щоб знайти потрібний рядок на складі."
            )
        if self._operation_would_lose_balance_tracking(row[1], exclude_field_id=field_id):
            raise ValueError(
                "Це поле не можна видалити: без нього бот перестане оновлювати залишок складу для цієї категорії."
            )
        with self.conn:
            self.conn.execute("DELETE FROM bot_operation_fields WHERE id = ?", (field_id,))

    def list_operation_field_columns(self, operation_field_id):
        return self.conn.execute(
            "SELECT id, operation_field_id, sheet, column_key, marker, write_mode, position, builtin_key "
            "FROM bot_operation_field_columns WHERE operation_field_id = ? ORDER BY position, id",
            (operation_field_id,),
        ).fetchall()

    # Захист від подвійного запису (Крок 3+ "Дії", Етап 3): execute_operation_
    # write резолвить значення ЗА column_key (warehouse_data.py, _OPERATION_
    # FIELD_ITEM_KEYS), не за тим, ЯКЕ поле власник прив'язки — тож якщо
    # ДВІ прив'язки (навіть із різних полів-запитів) вказують на ОДНУ й ту ж
    # генеровану (write_mode='generic') колонку, значення застосується
    # ДВІЧІ (напр. "Остаток, шт" зменшиться вдвічі). Перевіряємо це саме тут
    # (не лише в GUI), щоб жоден шлях додавання прив'язки не міг створити
    # такий дубль.
    # exclude_column_id — при РЕДАГУВАННІ вже наявної прив'язки (не
    # створенні нової) виключає саму цю прив'язку з перевірки дублю, інакше
    # збереження без зміни колонки (лише маркера) завжди хибно "конфліктувало"
    # б само із собою.
    def add_operation_field_column(
        self, operation_field_id, sheet, column_key, marker, write_mode="generic", exclude_column_id=None,
    ):
        if write_mode == "generic":
            field_row = self.get_operation_field(operation_field_id)
            operation_id = field_row[1] if field_row else None
            query = """
                SELECT 1 FROM bot_operation_field_columns c
                JOIN bot_operation_fields f ON f.id = c.operation_field_id
                WHERE f.operation_id = ? AND c.sheet = ? AND c.column_key = ? AND c.write_mode = 'generic'
            """
            params = [operation_id, sheet, column_key]
            if exclude_column_id is not None:
                query += " AND c.id != ?"
                params.append(exclude_column_id)
            conflict = self.conn.execute(query, params).fetchone()
            if conflict:
                raise ValueError(
                    "Ця колонка вже отримує значення від іншого поля цієї дії — "
                    "спершу приберіть стару прив'язку."
                )
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            position = self.conn.execute(
                "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operation_field_columns "
                "WHERE operation_field_id = ?",
                (operation_field_id,),
            ).fetchone()[0]
            cursor = self.conn.execute(
                """
                INSERT INTO bot_operation_field_columns
                    (operation_field_id, sheet, column_key, marker, write_mode, position, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (operation_field_id, sheet, column_key, marker, write_mode, position, now, now),
            )
            return cursor.lastrowid

    def delete_operation_field_column(self, column_id):
        binding = self.conn.execute(
            "SELECT operation_field_id FROM bot_operation_field_columns WHERE id = ?",
            (column_id,),
        ).fetchone()
        if binding is not None:
            field_row = self.get_operation_field(binding[0])
            operation_id = field_row[1] if field_row else None
            if operation_id is not None and self._operation_would_lose_balance_tracking(
                operation_id, exclude_column_id=column_id, owning_field_id=binding[0]
            ):
                raise ValueError(
                    "Цю прив'язку не можна видалити: без неї бот перестане оновлювати залишок складу для цієї категорії."
                )
        with self.conn:
            self.conn.execute("DELETE FROM bot_operation_field_columns WHERE id = ?", (column_id,))

    def _insert_operation_field(self, operation_id, field_key, label, is_identity, builtin_key, now):
        position = self.conn.execute(
            "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operation_fields WHERE operation_id = ?",
            (operation_id,),
        ).fetchone()[0]
        cursor = self.conn.execute(
            """
            INSERT INTO bot_operation_fields
                (operation_id, field_key, label, is_identity, position, enabled, builtin_key, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
            """,
            (operation_id, field_key, label, 1 if is_identity else 0, position, builtin_key, now, now),
        )
        return cursor.lastrowid

    def _insert_operation_field_column(self, field_id, sheet, column_key, marker, write_mode, builtin_key, now):
        position = self.conn.execute(
            "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operation_field_columns "
            "WHERE operation_field_id = ?",
            (field_id,),
        ).fetchone()[0]
        self.conn.execute(
            """
            INSERT INTO bot_operation_field_columns
                (operation_field_id, sheet, column_key, marker, write_mode, position, builtin_key, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (field_id, sheet, column_key, marker, write_mode, position, builtin_key, now, now),
        )

    # Ідентифікаційні поля складу (товар/порода/товщина/ширина/довжина +
    # тип, якщо condition_identity) — спільні для приходу й продажу. БЕЗ
    # прив'язок-колонок (Задача користувача: "потрібно прибрати з дій те,
    # що просто порівнює — не бачу логіки в тому, щоб бачити... це мені
    # ні про що не каже") — для приходу/продажу ці поля використовуються
    # ЛИШЕ для внутрішнього пошуку/звірки потрібного рядка складу
    # (_warehouse_row_matches, telegram_dialog.py), а не для показу
    # якогось значення користувачу (на відміну від тих самих полів у
    # звітах "Остаток"/"Отчёт по продажам", де вони справді ПОКАЗУЮТЬ
    # значення — там прив'язки лишаються). Показ порожньої "□"-прив'язки,
    # яка нічого не виконує, лише плутав адміністратора.
    def _seed_warehouse_identity_fields(self, operation_id, condition_identity, now):
        field_ids = {}
        for field_key, label in _WAREHOUSE_IDENTITY_FIELD_LABELS:
            is_identity = condition_identity if field_key == "condition" else True
            field_id = self._insert_operation_field(operation_id, field_key, label, is_identity, field_key, now)
            field_ids[field_key] = field_id
        return field_ids

    # "quantity" (шт) і "measure" (м3/м2/мп — рівно один з трьох застосується
    # до конкретного рядка, залежно від _row_measure_kind; execute_operation_
    # write просто пропускає прив'язку, якщо відповідне значення в item
    # відсутнє, тож зайві прив'язки безпечні) — саме ті два поля, чиї
    # прив'язки на СКЛАД є write_mode='generic' (просте +/-).
    def _seed_quantity_measure_fields(self, operation_id, kind, now):
        qty_field_id = self._insert_operation_field(operation_id, "quantity", "Количество, шт", False, "quantity", now)
        measure_field_id = self._insert_operation_field(
            operation_id, "measure", "Количество, м3", False, "measure", now
        )
        if kind == "income":
            self._insert_operation_field_column(qty_field_id, "СКЛАД", "income_qty", "add", "generic", "income_qty", now)
            self._insert_operation_field_column(qty_field_id, "СКЛАД", "balance_qty", "add", "generic", "balance_qty", now)
            for suffix in ("volume", "area", "linear"):
                self._insert_operation_field_column(
                    measure_field_id, "СКЛАД", f"income_{suffix}", "add", "generic", f"income_{suffix}", now
                )
                self._insert_operation_field_column(
                    measure_field_id, "СКЛАД", f"balance_{suffix}", "add", "generic", f"balance_{suffix}", now
                )
        elif kind == "sale":
            self._insert_operation_field_column(qty_field_id, "СКЛАД", "sold_qty", "add", "generic", "sold_qty", now)
            self._insert_operation_field_column(
                qty_field_id, "СКЛАД", "balance_qty", "subtract", "generic", "balance_qty", now
            )
            self._insert_operation_field_column(
                qty_field_id, SALES_SHEET_NAME, "quantity", "add", "ledger", "sales_quantity", now
            )
            for suffix in ("volume", "area", "linear"):
                self._insert_operation_field_column(
                    measure_field_id, "СКЛАД", f"sold_{suffix}", "add", "generic", f"sold_{suffix}", now
                )
                self._insert_operation_field_column(
                    measure_field_id, "СКЛАД", f"balance_{suffix}", "subtract", "generic", f"balance_{suffix}", now
                )
                self._insert_operation_field_column(
                    measure_field_id,
                    SALES_SHEET_NAME,
                    f"total_{suffix}",
                    "add",
                    "ledger",
                    f"sales_total_{suffix}",
                    now,
                )
        else:
            # "writeoff" (списання) і будь-який майбутній подібний вид -
            # лише відняти з залишку складу, без ledger-запису (немає
            # окремого листа списання) і без власної кумулятивної
            # "списано"-колонки (на відміну від income_qty/sold_qty).
            self._insert_operation_field_column(
                qty_field_id, "СКЛАД", "balance_qty", "subtract", "generic", "balance_qty", now
            )
            for suffix in ("volume", "area", "linear"):
                self._insert_operation_field_column(
                    measure_field_id, "СКЛАД", f"balance_{suffix}", "subtract", "generic", f"balance_{suffix}", now
                )

    # Клиент/цена/сумма/оплата — лише на аркуші ПРОДАЖА МАТЕРИАЛА, лише для
    # продажу; усі write_mode='ledger' (їх фактично записує sale_sheet_values,
    # разом з обчисленою математикою суми/готівки/банку).
    def _seed_sale_ledger_fields(self, operation_id, now):
        client_id = self._insert_operation_field(operation_id, "client", "Клиент", False, "client", now)
        self._insert_operation_field_column(client_id, SALES_SHEET_NAME, "client", "info", "ledger", "client", now)

        # ТЗ: "Адрес выгрузки" — вводиться один раз на всю продажу (як і
        # client), обов'язкове поле (_flat_checklist_missing_fields).
        address_id = self._insert_operation_field(operation_id, "address", "Адрес выгрузки", False, "address", now)
        self._insert_operation_field_column(address_id, SALES_SHEET_NAME, "address", "info", "ledger", "address", now)

        # Базове слово "Цена" — " за {одиниця}" дописується динамічно
        # (_sale_mandatory_fields_missing, telegram_dialog.py), бо одиниця
        # залежить від товару (шт/м3/м2/мп).
        price_id = self._insert_operation_field(operation_id, "price_per_unit", "Цена", False, "price_per_unit", now)
        self._insert_operation_field_column(
            price_id, SALES_SHEET_NAME, "price_per_unit", "info", "ledger", "price_per_unit", now
        )

        amount_id = self._insert_operation_field(operation_id, "total_amount", "Сумма", False, "total_amount", now)
        self._insert_operation_field_column(
            amount_id, SALES_SHEET_NAME, "total_amount", "add", "ledger", "total_amount", now
        )

        payment_id = self._insert_operation_field(operation_id, "payment_method", "Способ оплаты", False, "payment_method", now)
        self._insert_operation_field_column(
            payment_id, SALES_SHEET_NAME, "payment_method", "info", "ledger", "payment_method", now
        )

    def _seed_antiseptic_fields(self, operation_id, now):
        for field_key, label, marker in _ANTISEPTIC_FIELD_DEFS:
            field_id = self._insert_operation_field(operation_id, field_key, label, False, field_key, now)
            self._insert_operation_field_column(
                field_id, ANTISEPTIC_SHEET_NAME, field_key, marker, "ledger", field_key, now
            )

    def _seed_builtin_operations(self):
        now = datetime.now().isoformat(timespec="seconds")
        for entry in BUILTIN_OPERATIONS:
            exists = self.conn.execute(
                "SELECT 1 FROM bot_operations WHERE builtin_key = ?", (entry["builtin_key"],)
            ).fetchone()
            if exists:
                continue
            with self.conn:
                position = self.conn.execute(
                    "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operations WHERE parent_action_code = ?",
                    (entry["parent_action_code"],),
                ).fetchone()[0]
                cursor = self.conn.execute(
                    """
                    INSERT INTO bot_operations
                        (code, kind, requires_row_identity, label, parent_action_code, prefill_json,
                         position, enabled, builtin_key, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?, ?, ?, 1, ?, ?, ?)
                    """,
                    (
                        entry["builtin_key"],
                        entry["kind"],
                        entry["label"],
                        entry["parent_action_code"],
                        json.dumps(entry["prefill"], ensure_ascii=False),
                        position,
                        entry["builtin_key"],
                        now,
                        now,
                    ),
                )
                operation_id = cursor.lastrowid
                self._seed_warehouse_identity_fields(operation_id, entry["condition_identity"], now)
                self._seed_quantity_measure_fields(operation_id, entry["kind"], now)
                if entry["kind"] == "sale":
                    self._seed_sale_ledger_fields(operation_id, now)

        for entry in BUILTIN_SERVICE_OPERATIONS:
            exists = self.conn.execute(
                "SELECT 1 FROM bot_operations WHERE builtin_key = ?", (entry["builtin_key"],)
            ).fetchone()
            if exists:
                continue
            with self.conn:
                position = self.conn.execute(
                    "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operations WHERE parent_action_code = ?",
                    (entry["parent_action_code"],),
                ).fetchone()[0]
                cursor = self.conn.execute(
                    """
                    INSERT INTO bot_operations
                        (code, kind, requires_row_identity, label, parent_action_code, prefill_json,
                         position, enabled, builtin_key, created_at, updated_at)
                    VALUES (?, ?, 0, ?, ?, NULL, ?, 1, ?, ?, ?)
                    """,
                    (
                        entry["builtin_key"],
                        entry["kind"],
                        entry["label"],
                        entry["parent_action_code"],
                        position,
                        entry["builtin_key"],
                        now,
                        now,
                    ),
                )
                operation_id = cursor.lastrowid
                self._seed_antiseptic_fields(operation_id, now)

    # Задача користувача: "Показать остаток склада"/"Показать отчёт по
    # продажам" теж мають бути редаговані (додавання/прибирання внутрішніх
    # дій), як і категорії приходу/продажу. kind='report', requires_row_
    # identity=0 (звіт показує ВСІ рядки, не шукає один) — жодне поле не
    # is_identity, усе вільно редагується/видаляється.
    def _seed_report_operations(self):
        now = datetime.now().isoformat(timespec="seconds")
        for entry in BUILTIN_REPORT_OPERATIONS:
            exists = self.conn.execute(
                "SELECT 1 FROM bot_operations WHERE builtin_key = ?", (entry["builtin_key"],)
            ).fetchone()
            if exists:
                continue
            with self.conn:
                position = self.conn.execute(
                    "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operations WHERE parent_action_code = ?",
                    (entry["parent_action_code"],),
                ).fetchone()[0]
                cursor = self.conn.execute(
                    """
                    INSERT INTO bot_operations
                        (code, kind, requires_row_identity, label, parent_action_code, prefill_json,
                         position, enabled, builtin_key, created_at, updated_at)
                    VALUES (?, 'report', 0, ?, ?, NULL, ?, 1, ?, ?, ?)
                    """,
                    (
                        entry["builtin_key"],
                        entry["label"],
                        entry["parent_action_code"],
                        position,
                        entry["builtin_key"],
                        now,
                        now,
                    ),
                )
                operation_id = cursor.lastrowid
                for field_key, label, column_key in entry["fields"]:
                    field_id = self._insert_operation_field(operation_id, field_key, label, False, field_key, now)
                    if column_key is not None:
                        self._insert_operation_field_column(
                            field_id, entry["sheet"], column_key, "info", "ledger", field_key, now
                        )

    # ТЗ: "Адрес выгрузки" — нове обов'язкове поле продажу, додане ПІСЛЯ
    # того, як sale_*/sale_antiseptic/sales_report уже давно засіяні в
    # продакшн-БД — звичайний _seed_builtin_operations/_seed_report_
    # operations шлях ("if exists: continue" за builtin_key ОПЕРАЦІЇ) їх
    # більше не торкається. Окрема одноразова міграція: перевіряє для
    # кожної вже існуючої дії, чи вже є в неї ПОЛЕ "address" (а не сама
    # операція), і додає, якщо ще нема — той самий ідемпотентний ідіом.
    def _migrate_add_address_field(self):
        now = datetime.now().isoformat(timespec="seconds")

        def ensure_address_field(operation_id, sheet):
            exists = self.conn.execute(
                "SELECT 1 FROM bot_operation_fields WHERE operation_id = ? AND field_key = 'address'",
                (operation_id,),
            ).fetchone()
            if exists:
                return
            with self.conn:
                field_id = self._insert_operation_field(operation_id, "address", "Адрес выгрузки", False, "address", now)
                self._insert_operation_field_column(field_id, sheet, "address", "info", "ledger", "address", now)

        for row in self.conn.execute("SELECT id FROM bot_operations WHERE kind = 'sale'").fetchall():
            ensure_address_field(row[0], SALES_SHEET_NAME)

        antiseptic_row = self.conn.execute(
            "SELECT id FROM bot_operations WHERE builtin_key = 'sale_antiseptic'"
        ).fetchone()
        if antiseptic_row:
            ensure_address_field(antiseptic_row[0], ANTISEPTIC_SHEET_NAME)

        sales_report_row = self.conn.execute(
            "SELECT id FROM bot_operations WHERE builtin_key = 'sales_report'"
        ).fetchone()
        if sales_report_row:
            ensure_address_field(sales_report_row[0], SALES_SHEET_NAME)

    # ТЗ gap-аналіз, ОСБ: реальна практика — рахувати ЛИСТАМИ і платити за
    # лист, а не за фіктивний м3 (товщина x ширина x довжина x кількість) —
    # сьогодні income_osb/sale_osb засіяні СТРУКТУРНО ІДЕНТИЧНО до ДОСКА
    # AD/KD (поле "measure" з прив'язками до income_volume/balance_volume
    # тощо). Прибираємо поле-запит "measure" для цих двох дій — телеграм-
    # діалог (_row_measure_kind) уже трактує "нема поля measure" як
    # "рахувати напряму по кількості" для товару "ОСБ". delete_operation_field
    # каскадно прибирає й усі прив'язки-колонки цього поля (FK ON DELETE
    # CASCADE) — жодного окремого SQL для bot_operation_field_columns не
    # треба. Ідемпотентно (як і сусідні одноразові міграції вище): якщо
    # поле вже відсутнє (попереднім запуском ЦІЄЇ міграції) — просто
    # пропускаємо.
    def _migrate_osb_quantity_only(self):
        for code in ("income_osb", "sale_osb"):
            operation_row = self.conn.execute(
                "SELECT id FROM bot_operations WHERE builtin_key = ?", (code,)
            ).fetchone()
            if operation_row is None:
                continue
            field_row = self.conn.execute(
                "SELECT id, builtin_key FROM bot_operation_fields WHERE operation_id = ? AND field_key = 'measure'",
                (operation_row[0],),
            ).fetchone()
            if field_row is None:
                continue
            # Свіжий пере-аудит (New-Notable #7): без цієї перевірки міграція
            # (виконується на КОЖЕН старт програми) видаляла б і поле, яке
            # адміністратор навмисно ЗАНОВО додав через "Дії" - той самий
            # provenance-гвард, що вже застосовує сусідня _remove_noop_
            # identity_bindings (перевіряє builtin_key перед автовидаленням
            # вручну доданих рядків). Видаляємо ЛИШЕ оригінально засіяне поле.
            if field_row[1] != "measure":
                continue
            self.delete_operation_field(field_row[0])

    # Одноразове виправлення мови (Задача користувача, знайдено при
    # підключенні чек-листа до конфігурації): перший сідінг (Крок 3+
    # Етапи 1-3) помилково заповнив мітки полів-запитів УКРАЇНСЬКОЮ — не
    # проблема, поки це був лише перегляд у GUI, але критично, коли ці самі
    # мітки стають буквальним текстом у РОСІЙСЬКОМОВНОМУ чат-повідомленні
    # бота. UPDATE зіставляє (field_key, СТАРА мітка) -> НОВА — зачіпає
    # ЛИШЕ рядки, де мітка досі ТОЧНО збігається зі старою (помилковою)
    # — якщо адміністратор уже сам перейменував поле на щось інше, цей
    # рядок відповідності не знайде і НЕ зачепить чиєсь ручне редагування.
    # Ідемпотентно (безпечно виконувати щоразу): після першого виправлення
    # WHERE-умова більше нічого не знаходить.
    _LABEL_LANGUAGE_FIXES = [
        ("condition", "Тип (стан)", "Тип продукта"),
        ("thickness", "Товщина", "Толщина"),
        ("length", "Довжина", "Длина"),
        ("quantity", "Кількість, шт", "Количество, шт"),
        ("measure", "Кількість, м3/м2/мп", "Количество, м3/м2/мп"),
        ("client", "Клієнт", "Клиент"),
        # price_per_unit у ПРОДАЖУ — БАЗОВЕ слово "Цена" (не повна фраза):
        # " за {одиниця}" дописується динамічно (_sale_mandatory_fields_
        # missing), бо одиниця залежить від товару (шт/м3/м2/мп), на
        # відміну від антисептирования нижче, де одиниця завжди фіксована
        # "м3" — там повна готова фраза, без дописування.
        ("price_per_unit", "Ціна за одиницю", "Цена"),
        ("price_per_unit", "Цена за единицу", "Цена"),  # власна попередня помилка цієї ж міграції — коригуємо
        ("price_per_unit", "Ціна за м3", "Цена за м3"),
        ("total_amount", "Сума", "Сумма"),
        ("total_amount", "Вартість", "Стоимость"),
        ("payment_method", "Спосіб оплати", "Способ оплаты"),
        # payment_method (антисептирование) — мітка МАЄ бути та сама фраза,
        # що й у чек-листі (_antiseptic_mandatory_fields_missing), не сирий
        # заголовок колонки аркуша "Тип расчета" (той залишається окремо
        # видимим у рядку прив'язки, а не як заголовок поля).
        ("payment_method", "Тип розрахунку", "Способ оплаты"),
        ("payment_method", "Тип расчета", "Способ оплаты"),  # власна попередня помилка цієї ж міграції — коригуємо
        ("service_number", "№ послуги", "№ услуги"),
        ("service", "Послуга", "Услуга"),
        ("unit", "Од. виміру", "Ед. изм."),
        # volume (антисептирование) — та сама причина, що й payment_method
        # вище: мітка має збігатись із чек-листом ("Объем услуги, м3"), не
        # із сирою назвою колонки аркуша.
        ("volume", "Об'єм, м3", "Объем услуги, м3"),
        ("volume", "Объем, м3", "Объем услуги, м3"),  # власна попередня помилка цієї ж міграції — коригуємо
        ("total_volume", "Об'єм, м3", "Объем, м3"),
        ("payment_status", "Статус оплати", "Статус оплаты"),
        ("cash_amount", "Прихід готівкою", "Приход наличных"),
        ("bank_amount", "Прихід по банку", "Приход по банку"),
        ("reflection", "Відображення в розрахунках", "Отражение в расчетах"),
        ("manager", "Відповідальний", "Ответственный"),
        ("comment", "Коментар", "Комментарий"),
        ("balance_qty", "Залишок, шт", "Остаток, шт"),
        ("balance_volume", "Залишок, м3", "Остаток, м3"),
        ("balance_area", "Залишок, м2", "Остаток, м2"),
        ("balance_linear", "Залишок, мп", "Остаток, мп"),
    ]

    # Реальний баг, знайдений при підключенні звітів до реальної логіки:
    # field_key тут НЕ унікальний по всій системі — "volume" одночасно
    # належить і антисептированию (сервіс), і sales_report (звіт), кожен
    # зі своєю ПРАВИЛЬНОЮ, але РІЗНОЮ міткою. Без обмеження за kind цей
    # UPDATE зачіпав би поле sales_report лише тому, що воно випадково
    # мало той самий (field_key, old_label), що й антисептированиевський
    # рядок, який ця міграція мала виправити — звіти сіються ВЖЕ
    # правильною російською міткою (BUILTIN_REPORT_OPERATIONS), їм ця
    # міграція взагалі не потрібна.
    def _fix_operation_field_language(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            for field_key, old_label, new_label in self._LABEL_LANGUAGE_FIXES:
                self.conn.execute(
                    """
                    UPDATE bot_operation_fields SET label = ?, updated_at = ?
                    WHERE field_key = ? AND label = ?
                      AND operation_id NOT IN (SELECT id FROM bot_operations WHERE kind = 'report')
                    """,
                    (new_label, now, field_key, old_label),
                )

    # Одноразове спрощення (Задача користувача, після живого тестування:
    # "прошу вже не перший раз ЗАМІНИ на Количество м3, все інше, м2, мп -
    # лише показується в умовах") — сирий заголовок поля-запиту "measure"
    # показував усі 3 одиниці одразу ("Количество, м3/м2/мп"), хоча м3 —
    # базовий випадок для більшості товарів, а м2 (Вагонка)/мп (певні
    # розміри) — рідкісні винятки, які й так окремо показані під умовами
    # (_MEASURE_CONDITION_LABELS, gui.py). Той самий UPDATE-де-мітка-досі-
    # старий-текст ідіом, що й вище: не чіпає ручний адмінський перейменування.
    def _simplify_measure_field_label(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE bot_operation_fields SET label = ?, updated_at = ?
                WHERE field_key = 'measure' AND label = ?
                  AND operation_id NOT IN (SELECT id FROM bot_operations WHERE kind = 'report')
                """,
                ("Количество, м3", now, "Количество, м3/м2/мп"),
            )

    # Задача користувача: "Вагонка? значить і інформація за умовчуванням
    # для м2 потрібна, яке м3 у вагонці?" — Вагонка РЕАЛЬНО ніколи не
    # рахується в м3 (_is_area_based_product, telegram_dialog.py — площа
    # перевіряється ПЕРШОЮ, до розміру-винятку, тож м3 для цього товару
    # структурно недосяжний), тож заголовок поля-запиту для двох вагонка-
    # категорій одразу каже "м2", а не спільний дефолт усіх інших 6
    # категорій. Якщо колись до area-based товарів додасться ще один
    # (напр. ОСБ) — відповідні коди теж треба буде дописати сюди.
    def _relabel_vagonka_measure_field(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE bot_operation_fields SET label = ?, updated_at = ?
                WHERE field_key = 'measure' AND label = ?
                  AND operation_id IN (
                      SELECT id FROM bot_operations WHERE code IN ('income_vagonka', 'sale_vagonka')
                  )
                """,
                ("Количество, м2", now, "Количество, м3"),
            )

    # Крок 4.2 (Задача користувача, "єдиний текст на вузол замість двох
    # накладених"): РІВНО для 7 уже мігрованих кореневих вузлів (позначені
    # migration_key — ПРИХОД/РЕАЛИЗАЦИЯ/ДАННЫЕ/СКЛАД/ПРОДАЖИ/Калькулятор/
    # Помощь) текст завжди йде з bot_message_templates —
    # _enter_custom_button_node (telegram_dialog.py) тепер програмно
    # ІГНОРУЄ message_text саме для таких вузлів, тож це очищення суто
    # захисне (прибирає вже введений, але мертвий текст, щоб не плутати
    # адміна в GUI-формі). НЕ чіпає звичайні кастомні вузли з action_code
    # (напр. ярлик на "start_sale" з власним вступним текстом) — для них
    # message_text і далі реально показується, і migration_key в них
    # завжди NULL. У продакшні на момент цього фікса message_text був
    # порожній для всіх 7 мігрованих вузлів (перевірено напряму в БД) —
    # тобто цей запуск є no-op сьогодні. Ідемпотентно, безпечно виконувати
    # щоразу.
    def _clear_message_text_for_builtin_actions(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE custom_menu_buttons SET message_text = '', updated_at = ?
                WHERE migration_key IS NOT NULL
                  AND message_text IS NOT NULL AND message_text != ''
                """,
                (now,),
            )

    # Реальний баг живого тестування (2026-08-08): тап/введення "СПИСАНИЕ"
    # давало мертвий кінець "Ок." замість "Списание товара. Выберите
    # категорию:". Корінь: _seed_builtin_migrated_custom_buttons - чисто
    # insert-if-missing (ніколи не оновлює вже наявний рядок за migration_key)
    # - тож коли BUILTIN_MIGRATED_CUSTOM_BUTTONS отримав action_code=
    # "start_writeoff" для migration_key="writeoff", уже наявний, засіяний
    # РАНІШЕ рядок "СПИСАНИЕ" (з action_code=NULL з тих часів, коли
    # списання ще не мало власної дії) так і лишився з NULL - і
    # _custom_root_button_by_label/_enter_custom_button_node (telegram_
    # dialog_core.py/_menu.py) для НЬОГО завжди перехоплюють натиснуту
    # кнопку РАНІШЕ будь-якого іншого розпізнавання, тож "СПИСАНИЕ"
    # назавжди впирався в "лист без дітей/дії" -> голе "Ок.". Одноразовий,
    # ідемпотентний backfill - оновлює ЛИШЕ цей конкретний, відомий рядок,
    # і ЛИШЕ якщо він досі NULL (не чіпає жодного кастомного вузла, який
    # адмін міг навмисно лишити без дії).
    def _backfill_writeoff_root_action_code(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE custom_menu_buttons SET action_code = 'start_writeoff', updated_at = ?
                WHERE migration_key = 'writeoff' AND action_code IS NULL
                """,
                (now,),
            )

    # Реальний баг живого тестування (2026-08-08, знайдений НЕ бот-скріншотом,
    # а повним регресійним прогоном — GUI/бот-тести, що звіряють точну
    # множину кнопок головного меню, почали ловити зайву "[test]" серед
    # них): рядок "СПИСАНИЕ (форма)" (migration_key="writeoff_form") у
    # продакшн-БД мав мітку "[test]" замість правильної — судячи з усього,
    # плейсхолдер лишився в ЖИВІЙ базі після ручного тестування мега-форми
    # списання в цій-таки сесії, і той самий insert-if-missing seed
    # (_seed_builtin_migrated_custom_buttons) ніколи не виправляє вже
    # наявний рядок. Одноразовий, ідемпотентний backfill — виправляє ЛИШЕ
    # цей конкретний, відомий випадок (мітка буквально "[test]"), не чіпає
    # жодну реальну кастомізацію адміністратора.
    def _backfill_writeoff_form_root_label(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE custom_menu_buttons SET label = 'СПИСАНИЕ (форма)', updated_at = ?
                WHERE migration_key = 'writeoff_form' AND label = '[test]'
                """,
                (now,),
            )

    # Одноразовий backfill (Задача користувача: "в тих що є зараз - поправ
    # імена скрізь де їх не підтянуло") - той самий фолбек "Гость {id}", що
    # тепер отримують НОВІ рядки (ensure_bot_user_seen/_message_context),
    # застосовується заднім числом до вже наявних рядків без імені й без
    # username (продакшн мав рівно такі - користувача додавали вручну
    # раніше, коли ці поля ще не читались/не вводились). Ідемпотентно -
    # WHERE ловить лише досі порожні рядки, повторний запуск нічого не робить.
    def _backfill_bot_user_names(self):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE bot_users
                SET full_name = 'Гость ' || telegram_id, updated_at = ?
                WHERE (full_name IS NULL OR full_name = '')
                  AND (username IS NULL OR username = '')
                """,
                (now,),
            )

    # Одноразове прибирання (Задача користувача, після живого тестування:
    # "потрібно прибрати з дій те, що просто порівнює — не бачу логіки в
    # тому, щоб бачити... це мені ні про що не каже") — інформаційні "□"
    # прив'язки товар/порода/тип/товщина/ширина/довжина для ПРИХОДУ й
    # ПРОДАЖУ нічого не виконують (write_mode='ledger', лише показ) і лише
    # плутали, не показуючи жодної реальної дії. Захоплює лише прив'язки,
    # заведені builtin_key ('product'..'length', і '..._sales' — точно ті,
    # що сіяв _seed_warehouse_identity_fields/_attach_sale_identity_
    # bindings), і ЛИШЕ для дій kind IN ('income','sale') — ЗВІТИ
    # (stock_report/sales_report) мають ті самі поля/builtin_key, але там
    # ці прив'язки РЕАЛЬНО показують значення в звіті, тому не чіпаються.
    # Якщо адмін уже додав власну змістовну прив'язку до цих полів — вона
    # має ІНШИЙ builtin_key (None), тож ця чистка її не зачепить.
    def _remove_noop_identity_bindings(self):
        with self.conn:
            self.conn.execute(
                """
                DELETE FROM bot_operation_field_columns
                WHERE id IN (
                    SELECT c.id FROM bot_operation_field_columns c
                    JOIN bot_operation_fields f ON f.id = c.operation_field_id
                    JOIN bot_operations o ON o.id = f.operation_id
                    WHERE o.kind IN ('income', 'sale')
                      AND c.builtin_key IN (
                        'product', 'breed', 'condition', 'thickness', 'width', 'length',
                        'product_sales', 'breed_sales', 'condition_sales',
                        'thickness_sales', 'width_sales', 'length_sales'
                      )
                )
                """
            )

    # Одноразова міграція (Задача користувача: "звіти теж мають реально
    # показувати те, що редагується, а не бути бутафорією"). Перше
    # розширення "Дії" на stock_report/sales_report сіяло поля-запити, що
    # НЕ відповідали реальним колонкам справжнього звіту (окремі thickness/
    # width/length замість комбінованого "Размер"/"Товар", мітки типу
    # "Товар"/"Тип продукта", яких немає в реальному друкованому звіті) —
    # бо тоді _sales_report_spec/_stock_report_spec ще не читали цю
    # таблицю взагалі. Тепер, коли telegram_dialog.py дійсно будує колонки
    # звіту з bot_operation_fields, стару, невідповідну структуру треба
    # замінити на нову (BUILTIN_REPORT_OPERATIONS), що відповідає 1-в-1.
    # Сигнал "ще стара структура" — наявність поля з builtin_key, якого в
    # НОВІЙ структурі вже немає (thickness/width/length для stock_report;
    # breed/condition/thickness/width/length для sales_report) — після
    # міграції цих ключів більше нема, тож повторний запуск — no-op.
    # Перевірено напряму (2026-07-17): продакшн ще не мав ЖОДНИХ реальних
    # адмінських правок цих двох дій (усі builtin_key й досі збігаються з
    # початковим сідінгом) — тому повне перестворення полів безпечне.
    _OBSOLETE_REPORT_FIELD_BUILTIN_KEYS = {
        "stock_report": ("thickness", "width", "length"),
        "sales_report": ("breed", "condition", "thickness", "width", "length"),
    }

    def _restructure_report_fields(self):
        now = datetime.now().isoformat(timespec="seconds")
        for entry in BUILTIN_REPORT_OPERATIONS:
            operation = self.conn.execute(
                "SELECT id FROM bot_operations WHERE builtin_key = ?", (entry["builtin_key"],)
            ).fetchone()
            if operation is None:
                continue
            operation_id = operation[0]
            obsolete_keys = self._OBSOLETE_REPORT_FIELD_BUILTIN_KEYS.get(entry["builtin_key"], ())
            if not obsolete_keys:
                continue
            placeholders = ",".join("?" for _ in obsolete_keys)
            has_old_structure = self.conn.execute(
                f"SELECT 1 FROM bot_operation_fields WHERE operation_id = ? AND builtin_key IN ({placeholders})",
                (operation_id, *obsolete_keys),
            ).fetchone()
            if not has_old_structure:
                continue
            with self.conn:
                self.conn.execute("DELETE FROM bot_operation_fields WHERE operation_id = ?", (operation_id,))
                for field_key, label, column_key in entry["fields"]:
                    field_id = self._insert_operation_field(operation_id, field_key, label, False, field_key, now)
                    if column_key is not None:
                        self._insert_operation_field_column(
                            field_id, entry["sheet"], column_key, "info", "ledger", field_key, now
                        )

    def list_users(self):
        cursor = self.conn.execute(
            """
            SELECT id, telegram_id, COALESCE(username, ''), COALESCE(full_name, ''), role, last_seen_at
            FROM bot_users
            ORDER BY role, full_name COLLATE NOCASE, username COLLATE NOCASE, id
            """
        )
        return cursor.fetchall()

    def get_user(self, user_id):
        cursor = self.conn.execute(
            """
            SELECT id, telegram_id, COALESCE(username, ''), COALESCE(full_name, ''), role, last_seen_at
            FROM bot_users
            WHERE id = ?
            """,
            (user_id,),
        )
        return cursor.fetchone()

    # role навмисно без дефолту (був "admin", що розходилось зі схемою -
    # DEFAULT 'user' - і жоден виклик на нього насправді не покладався;
    # мовчазний дефолт на адміна - небезпечний landmine для майбутнього коду).
    def add_user(self, telegram_id, username, full_name, role):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_users (telegram_id, username, full_name, role, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (telegram_id, username, full_name, role, now, now),
            )

    def update_user(self, user_id, username, full_name, role):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                UPDATE bot_users
                SET username = ?, full_name = ?, role = ?, updated_at = ?
                WHERE id = ?
                """,
                (username, full_name, role, now, user_id),
            )

    def get_user_role(self, telegram_id):
        cursor = self.conn.execute(
            "SELECT role FROM bot_users WHERE telegram_id = ?",
            (str(telegram_id),),
        )
        row = cursor.fetchone()
        return row[0] if row else None

    # Викликається на кожне повідомлення від реального Telegram-користувача
    # (telegram_dialog.py._build_reply_pipeline) - Задача користувача: "щоб
    # програма бачила користувачів, які вже хоча б натиснули кнопку
    # розпочати чи старт бота", без ручного введення ID адміністратором.
    # Один атомарний запит замість SELECT-потім-гілка: новий telegram_id -
    # вставляється з роллю GUEST; уже наявний - username/full_name
    # оновлюються (щоб нік читався так, як він РЕАЛЬНО підписаний у
    # Telegram зараз), АЛЕ роль ніколи не чіпається тут - вже призначена
    # роль недоторканна. WHERE у DO UPDATE - щоб не рухати updated_at на
    # кожне повідомлення, якщо нік/ім'я насправді не змінились.
    # last_seen_at - Задача користувача: "час останнього відправленого
    # повідомлення користувачем в чат" у "Персонал". На відміну від
    # username/full_name/updated_at (оновлюються лише коли РЕАЛЬНО
    # змінились - щоб не смикати updated_at на кожне повідомлення),
    # last_seen_at МАЄ рухатись на кожен виклик - це і є ціле призначення
    # поля. Тому - CASE замість WHERE на весь запит: WHERE заблокував би
    # ВЕСЬ UPDATE (включно з last_seen_at), якщо нік/ім'я не змінились;
    # CASE лишає той самий "не чіпати дарма" фільтр ЛИШЕ на updated_at.
    def ensure_bot_user_seen(self, telegram_id, username, full_name):
        if telegram_id is None:
            return
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_users (telegram_id, username, full_name, role, last_seen_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(telegram_id) DO UPDATE SET
                    username = excluded.username,
                    full_name = excluded.full_name,
                    last_seen_at = excluded.last_seen_at,
                    updated_at = CASE
                        WHEN bot_users.username IS NOT excluded.username
                          OR bot_users.full_name IS NOT excluded.full_name
                        THEN excluded.updated_at
                        ELSE bot_users.updated_at
                    END
                """,
                (telegram_id, username, full_name, permissions.GUEST, now, now, now),
            )

    def delete_user(self, user_id):
        with self.conn:
            self.conn.execute("DELETE FROM bot_users WHERE id = ?", (user_id,))

    def add_action_log(self, action_type, details, actor_user_id=None):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO action_log (actor_user_id, action_type, details_json, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    actor_user_id,
                    action_type,
                    json.dumps(details or {}, ensure_ascii=False),
                    now,
                ),
            )

    def add_stock_movement(self, movement):
        was_in_transaction = self.conn.in_transaction
        self.conn.execute(
            """
            INSERT INTO stock_movements (
                movement_type, source, telegram_user_id, username, full_name,
                product, breed, condition, thickness, width, length,
                quantity, volume, area, linear, reason, sheet_row_id, original_text, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                movement.get("movement_type", "income"),
                movement.get("source", "telegram"),
                movement.get("telegram_user_id"),
                movement.get("username"),
                movement.get("full_name"),
                movement.get("product"),
                movement.get("breed"),
                movement.get("condition"),
                movement.get("thickness"),
                movement.get("width"),
                movement.get("length"),
                movement.get("quantity"),
                movement.get("volume"),
                movement.get("area"),
                movement.get("linear"),
                movement.get("reason"),
                movement.get("sheet_row_id"),
                movement.get("original_text"),
                movement.get("created_at", datetime.now().isoformat(timespec="seconds")),
            ),
        )
        if not was_in_transaction:
            self.conn.commit()

    def get_user_preference(self, telegram_user_id):
        cursor = self.conn.execute(
            """
            SELECT telegram_user_id, chat_id, COALESCE(username, ''), COALESCE(full_name, ''),
                   bot_mode, language, COALESCE(claude_api_key, ''), COALESCE(claude_key_updated_at, '')
            FROM bot_user_preferences
            WHERE telegram_user_id = ?
            """,
            (str(telegram_user_id),),
        )
        row = cursor.fetchone()
        if not row:
            return None
        return {
            "telegram_user_id": row[0],
            "chat_id": row[1],
            "username": row[2],
            "full_name": row[3],
            "bot_mode": row[4],
            "language": row[5],
            "claude_api_key": row[6],
            "claude_key_updated_at": row[7],
        }

    def save_user_preference(self, context, bot_mode, language="ru"):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_user_preferences (
                    telegram_user_id, chat_id, username, full_name,
                    bot_mode, language, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(telegram_user_id) DO UPDATE SET
                    chat_id = excluded.chat_id,
                    username = excluded.username,
                    full_name = excluded.full_name,
                    bot_mode = excluded.bot_mode,
                    language = excluded.language,
                    updated_at = excluded.updated_at
                """,
                (
                    str(context.get("user_id")),
                    str(context.get("chat_id")),
                    context.get("username", ""),
                    context.get("full_name", ""),
                    bot_mode,
                    language,
                    now,
                    now,
                ),
            )

    def save_user_claude_api_key(self, context, claude_api_key):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_user_preferences (
                    telegram_user_id, chat_id, username, full_name,
                    bot_mode, language, claude_api_key, claude_key_updated_at,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, 'no_ai', 'ru', ?, ?, ?, ?)
                ON CONFLICT(telegram_user_id) DO UPDATE SET
                    chat_id = excluded.chat_id,
                    username = excluded.username,
                    full_name = excluded.full_name,
                    claude_api_key = excluded.claude_api_key,
                    claude_key_updated_at = excluded.claude_key_updated_at,
                    updated_at = excluded.updated_at
                """,
                (
                    str(context.get("user_id")),
                    str(context.get("chat_id")),
                    context.get("username", ""),
                    context.get("full_name", ""),
                    claude_api_key.strip(),
                    now,
                    now,
                    now,
                ),
            )

    # --- Історія руху товару та стан діалогу (bot_pending_operations) ---
    def list_stock_movements(self, movement_type="income", start_date=None, end_date=None, limit=500):
        conditions = ["movement_type = ?"]
        params = [movement_type]
        if start_date:
            conditions.append("date(created_at) >= date(?)")
            params.append(start_date.isoformat() if isinstance(start_date, date) else str(start_date))
        if end_date:
            conditions.append("date(created_at) <= date(?)")
            params.append(end_date.isoformat() if isinstance(end_date, date) else str(end_date))
        params.append(limit)
        cursor = self.conn.execute(
            f"""
            SELECT
                movement_type, source, telegram_user_id, username, full_name,
                product, breed, condition, thickness, width, length,
                quantity, volume, area, linear, sheet_row_id, original_text, created_at
            FROM stock_movements
            WHERE {" AND ".join(conditions)}
            ORDER BY created_at DESC, id DESC
            LIMIT ?
            """,
            params,
        )
        keys = [
            "movement_type",
            "source",
            "telegram_user_id",
            "username",
            "full_name",
            "product",
            "breed",
            "condition",
            "thickness",
            "width",
            "length",
            "quantity",
            "volume",
            "area",
            "linear",
            "sheet_row_id",
            "original_text",
            "created_at",
        ]
        return [dict(zip(keys, row)) for row in cursor.fetchall()]

    def list_action_log(self, limit=100):
        cursor = self.conn.execute(
            """
            SELECT id, action_type, COALESCE(details_json, ''), created_at
            FROM action_log
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        )
        return cursor.fetchall()

    def get_action_log(self, log_id):
        cursor = self.conn.execute(
            """
            SELECT id, action_type, COALESCE(details_json, ''), created_at
            FROM action_log
            WHERE id = ?
            """,
            (log_id,),
        )
        return cursor.fetchone()

    def delete_action_log(self, log_id):
        with self.conn:
            self.conn.execute("DELETE FROM action_log WHERE id = ?", (log_id,))

    def clear_action_log(self):
        with self.conn:
            self.conn.execute("DELETE FROM action_log")

    def get_pending_operation(self, chat_id, telegram_user_id):
        cursor = self.conn.execute(
            """
            SELECT id, operation_type, status, payload_json
            FROM bot_pending_operations
            WHERE chat_id = ? AND telegram_user_id = ?
            """,
            (str(chat_id), str(telegram_user_id)),
        )
        row = cursor.fetchone()
        if not row:
            return None
        return {
            "id": row[0],
            "operation_type": row[1],
            "status": row[2],
            "payload": json.loads(row[3]),
        }

    def save_pending_operation(self, chat_id, telegram_user_id, operation_type, status, payload):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO bot_pending_operations (
                    chat_id, telegram_user_id, operation_type, status,
                    payload_json, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(chat_id, telegram_user_id) DO UPDATE SET
                    operation_type = excluded.operation_type,
                    status = excluded.status,
                    payload_json = excluded.payload_json,
                    updated_at = excluded.updated_at
                """,
                (
                    str(chat_id),
                    str(telegram_user_id),
                    operation_type,
                    status,
                    json.dumps(payload, ensure_ascii=False),
                    now,
                    now,
                ),
            )

    def delete_pending_operation(self, chat_id, telegram_user_id):
        with self.conn:
            self.conn.execute(
                "DELETE FROM bot_pending_operations WHERE chat_id = ? AND telegram_user_id = ?",
                (str(chat_id), str(telegram_user_id)),
            )

    # --- Журнал виконаних робіт (записи про роботу асистента над проєктом) ---
    def add_work_log_entry(self, title, summary, benefit="", future_impact=""):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO dev_work_log (title, summary, benefit, future_impact, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (title, summary, benefit, future_impact, now),
            )

    def list_work_log(self, limit=200):
        cursor = self.conn.execute(
            """
            SELECT id, title, summary, COALESCE(benefit, ''), COALESCE(future_impact, ''), created_at
            FROM dev_work_log
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        )
        return cursor.fetchall()

    def get_work_log_entry(self, log_id):
        cursor = self.conn.execute(
            """
            SELECT id, title, summary, COALESCE(benefit, ''), COALESCE(future_impact, ''), created_at
            FROM dev_work_log
            WHERE id = ?
            """,
            (log_id,),
        )
        return cursor.fetchone()

    def delete_work_log_entry(self, log_id):
        with self.conn:
            self.conn.execute("DELETE FROM dev_work_log WHERE id = ?", (log_id,))

    def clear_work_log(self):
        with self.conn:
            self.conn.execute("DELETE FROM dev_work_log")

    # --- Пам'ять виправлень імені клієнта (client_name_aliases) ---
    def remember_client_alias(self, typed_text, canonical_name):
        normalized = _normalize_phrase(typed_text)
        if not normalized or not canonical_name:
            return
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO client_name_aliases (normalized_typo, canonical_name, created_at)
                VALUES (?, ?, ?)
                ON CONFLICT(normalized_typo) DO UPDATE SET
                    canonical_name = excluded.canonical_name,
                    created_at = excluded.created_at
                """,
                (normalized, canonical_name, now),
            )

    def get_client_alias(self, typed_text):
        normalized = _normalize_phrase(typed_text)
        if not normalized:
            return None
        cursor = self.conn.execute(
            "SELECT canonical_name FROM client_name_aliases WHERE normalized_typo = ?",
            (normalized,),
        )
        row = cursor.fetchone()
        return row[0] if row else None

    # --- Індексований пошук по складу (warehouse_items) — для майбутньої заміни
    # лінійних Python-сканів у _stock_balance_rows/_similar_sale_rows/фільтрах ---
    def find_warehouse_items(
        self,
        product=None,
        breed=None,
        condition=None,
        thickness=None,
        width=None,
        length=None,
        only_in_stock=False,
        dimension_tolerance=0.01,
    ):
        conditions = []
        params = []
        if product is not None:
            conditions.append("product = ?")
            params.append(normalize_product_category(product))
        if breed is not None:
            conditions.append("breed = ?")
            params.append(breed)
        if condition is not None:
            conditions.append("condition = ?")
            params.append(condition)
        for field, value in (("thickness", thickness), ("width", width), ("length", length)):
            if value is not None:
                conditions.append(f"ABS({field} - ?) < ?")
                params.extend([value, dimension_tolerance])
        if only_in_stock:
            conditions.append("(balance_qty > 0 OR balance_volume > 0 OR balance_area > 0)")

        where = " AND ".join(conditions) if conditions else "1=1"
        cursor = self.conn.execute(
            f"""
            SELECT sheet_row_id, sku, product, breed, condition,
                   thickness, width, length, unit,
                   income_qty, income_volume, income_area,
                   sold_qty, sold_volume, sold_area,
                   balance_qty, balance_volume, balance_area
            FROM warehouse_items
            WHERE {where}
            ORDER BY product, breed, thickness, width, length
            """,
            params,
        )
        keys = (
            "sheet_row_id", "sku", "product", "breed", "condition",
            "thickness", "width", "length", "unit",
            "income_qty", "income_volume", "income_area",
            "sold_qty", "sold_volume", "sold_area",
            "balance_qty", "balance_volume", "balance_area",
        )
        return [dict(zip(keys, row)) for row in cursor.fetchall()]

    # Звіт "Низкий остаток" — перший реальний споживач find_warehouse_items-
    # інфраструктури (warehouse_items + idx_warehouse_items_balance). Фільтр
    # виключно за balance_qty ("Кількість, шт") - єдиний вимір, що присутній
    # у КОЖНІЙ категорії без винятку (навіть ОСБ, де це взагалі єдиний
    # вимір) - тому не потрібно визначати "який вимір головний" для рядка,
    # як це робить _stock_balance_rows (is_area_based/is_linear_based/
    # is_quantity_only). balance_volume/area показуються лише як
    # інформаційна колонка, у фільтрації/сортуванні участі не беруть.
    def low_stock_warehouse_items(self, threshold):
        cursor = self.conn.execute(
            """
            SELECT sheet_row_id, sku, product, breed, condition,
                   thickness, width, length, unit,
                   balance_qty, balance_volume, balance_area
            FROM warehouse_items
            WHERE balance_qty IS NOT NULL AND balance_qty <= ?
            ORDER BY balance_qty ASC, product, breed, thickness, width, length
            """,
            (threshold,),
        )
        keys = (
            "sheet_row_id", "sku", "product", "breed", "condition",
            "thickness", "width", "length", "unit",
            "balance_qty", "balance_volume", "balance_area",
        )
        return [dict(zip(keys, row)) for row in cursor.fetchall()]

    # Задача користувача (2026-08-08, реальний баг живого тестування —
    # рядок "доска AD|Ель|KD|25x150x6000": "Остаток, шт"=7428, "Остаток,
    # м3"=58.265, хоча за розміром рядка 7428 шт фізично зайняли б 167.13 м3):
    # "потрібно щоб в середині програми був у користувача при такій різниці -
    # вибір що саме зберегти" — САМЕ в GUI-програмі (не в боті), бо це
    # адміністративне виправлення вже наявних даних складу, а не щоденний
    # чат-флоу. Скануємо СКЛАД напряму (той самий `warehouse_rows`, яким
    # користується решта складської логіки — не вторинну, необов'язково
    # синхронну таблицю `warehouse_items`), для кожного рядка з фізичним
    # виміром (не ОСБ) рахуємо, скільки штук РЕАЛЬНО відповідає збереженому
    # м3/м2/мп (`_shared_row_measure_kind`/`_shared_piece_measure` — та сама
    # класифікація товару, що вже й бот використовує, перенесена в utils.py,
    # щоб не дублювати правила площинний/безрозмірний/погонний товар у двох
    # місцях). `tolerance_pieces` — навмисно НЕ 0: звичайні операції бота й
    # так тримають шт/м3 у синхроні з точністю до похибки округлення, тому
    # поріг ловить лише СПРАВЖНІ розбіжності (типово — ручне/імпортне
    # введення), а не шум від багаторазового +/-.
    def find_quantity_measure_mismatches(self, tolerance_pieces=1.0):
        headers, columns, rows = warehouse_rows(self)
        product_idx = columns.get("product")
        breed_idx = columns.get("breed")
        thickness_idx = columns.get("thickness")
        width_idx = columns.get("width")
        length_idx = columns.get("length")
        sku_idx = columns.get("sku")
        balance_qty_idx = columns.get("balance_qty")
        if None in (product_idx, thickness_idx, width_idx, length_idx, balance_qty_idx):
            return []
        measure_balance_columns = {
            "volume": columns.get("balance_volume"),
            "area": columns.get("balance_area"),
            "linear": columns.get("balance_linear"),
        }
        mismatches = []
        for row_id, row in rows:
            product = row_value(row, product_idx)
            thickness = row_value(row, thickness_idx)
            width = row_value(row, width_idx)
            length = row_value(row, length_idx)
            if thickness in (None, "") or width in (None, "") or length in (None, ""):
                continue
            measure_kind = _shared_row_measure_kind(product, thickness, width)
            if measure_kind is None:
                continue
            measure_idx = measure_balance_columns.get(measure_kind)
            if measure_idx is None:
                continue
            piece = _shared_piece_measure(thickness, width, length, measure_kind)
            if piece <= 0:
                continue
            balance_qty = _number_value(row_value(row, balance_qty_idx))
            balance_measure = _number_value(row_value(row, measure_idx))
            implied_qty = balance_measure / piece
            if abs(implied_qty - balance_qty) <= tolerance_pieces:
                continue
            mismatches.append({
                "row_id": row_id,
                "sku": row_value(row, sku_idx) if sku_idx is not None else None,
                "product": product,
                "breed": row_value(row, breed_idx) if breed_idx is not None else None,
                "thickness": thickness,
                "width": width,
                "length": length,
                "measure_kind": measure_kind,
                "measure_unit": ITEM_MEASURE_UNIT[measure_kind],
                "measure_column_index": measure_idx,
                "balance_qty_column_index": balance_qty_idx,
                "balance_qty": balance_qty,
                "balance_measure": balance_measure,
                "implied_qty_from_measure": implied_qty,
                "implied_measure_from_qty": round(balance_qty * piece, 6),
                "piece_measure": piece,
            })
        return mismatches

    # Застосовує рішення адміністратора з діалогу реконсиляції — рівно ОДНЕ
    # з двох полів (кількість чи вимір) лишається джерелом істини, друге
    # перераховується з нього. `new_qty`/`new_measure` — уже готові, обрані
    # користувачем числа (округлення дробової кількості — GUI-рівнева
    # відповідальність, не тут); функція лише атомарно записує обидва
    # значення в один і той самий рядок (той самий BEGIN IMMEDIATE-патерн,
    # що вже захищає ручне редагування клітинки в gui.py).
    def resolve_quantity_measure_mismatch(self, row_id, balance_qty_column_index, measure_column_index, new_qty, new_measure):
        with self.conn:
            self.conn.execute("BEGIN IMMEDIATE")
            values = self.get_row(row_id)
            needed_length = max(balance_qty_column_index, measure_column_index) + 1
            while len(values) < needed_length:
                values.append("")
            values[balance_qty_column_index] = new_qty
            values[measure_column_index] = new_measure
            self.update_row(row_id, values)

    # --- Крок 4.4 "Дії": реальний CRUD над способами оплати ---
    def list_payment_method_options(self):
        return self.conn.execute(
            "SELECT id, label, kind FROM payment_method_options ORDER BY position, id"
        ).fetchall()

    def add_payment_method_option(self, label):
        label = label.strip()
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            position = self.conn.execute(
                "SELECT COALESCE(MAX(position), -1) + 1 FROM payment_method_options"
            ).fetchone()[0]
            cursor = self.conn.execute(
                "INSERT INTO payment_method_options (label, position, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (label, position, now, now),
            )
            return cursor.lastrowid

    # Перейменування ЗБЕРІГАЄ старе слово назавжди як розпізнаваний синонім
    # (Задача користувача: "щоб бот уже це сприймав у запиті автоматично") —
    # принципово інакше, ніж delete_payment_method_option, яка прибирає
    # розпізнавання повністю. No-op (жодного нового синоніма), якщо новий
    # текст після нормалізації фактично збігається зі старим.
    def update_payment_method_option(self, option_id, new_label):
        new_label = new_label.strip()
        row = self.conn.execute(
            "SELECT label FROM payment_method_options WHERE id = ?", (option_id,)
        ).fetchone()
        if not row:
            return
        old_label = row[0]
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            if _normalize_phrase(old_label) != _normalize_phrase(new_label):
                self.conn.execute(
                    "INSERT INTO payment_method_synonyms (option_id, phrase, created_at) VALUES (?, ?, ?)",
                    (option_id, old_label, now),
                )
            self.conn.execute(
                "UPDATE payment_method_options SET label = ?, updated_at = ? WHERE id = ?",
                (new_label, now, option_id),
            )

    # Видалення (на відміну від перейменування) прибирає варіант ПОВНІСТЮ:
    # каскадно й усі його синоніми (FK ON DELETE CASCADE) — тож розпізнавання
    # вільного тексту теж перестає бачити і поточну мітку, і всі старі.
    def delete_payment_method_option(self, option_id):
        with self.conn:
            self.conn.execute("DELETE FROM payment_method_options WHERE id = ?", (option_id,))

    def payment_method_label_collides(self, label, exclude_id=None):
        normalized = _normalize_phrase(label)
        for option_id, option_label, _kind in self.list_payment_method_options():
            if exclude_id is not None and option_id == exclude_id:
                continue
            if _normalize_phrase(option_label) == normalized:
                return True
        return False

    # Реальний баг з аудиту: antiseptic_sheet_values звіряла спосіб оплати з
    # ЖОРСТКИМ рядком "ЕФАКТУРА Б/Н" — перейменування (штатна дія в "Способи
    # оплати") тихо ламало розподіл готівка/банк. kind — стабільний прапорець,
    # що переживає перейменування. Лише ОДИН варіант може бути банком одразу
    # (сам розподіл бінарний) — встановлення 'bank' скидає його з усіх інших.
    def set_payment_method_kind(self, option_id, kind):
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            if kind == "bank":
                self.conn.execute(
                    "UPDATE payment_method_options SET kind = NULL, updated_at = ? WHERE id != ? AND kind = 'bank'",
                    (now, option_id),
                )
            self.conn.execute(
                "UPDATE payment_method_options SET kind = ?, updated_at = ? WHERE id = ?",
                (kind, now, option_id),
            )

    # Та сама логіка пошуку, що й resolve_payment_method (поточні мітки +
    # синоніми), але повертає kind замість мітки — щоб antiseptic_sheet_values
    # могла звірити "це банк?" за стабільним ідентифікатором, не текстом.
    def get_payment_method_kind(self, text):
        normalized = _normalize_phrase(text)
        if not normalized:
            return None
        for option_id, label, kind in self.list_payment_method_options():
            if _normalize_phrase(label) == normalized:
                return kind
        for option_id, phrase in self.conn.execute(
            "SELECT option_id, phrase FROM payment_method_synonyms"
        ).fetchall():
            if _normalize_phrase(phrase) == normalized:
                row = self.conn.execute(
                    "SELECT kind FROM payment_method_options WHERE id = ?", (option_id,)
                ).fetchone()
                if row:
                    return row[0]
        return None

    # Єдина точка "чи це спосіб оплати" — звіряє І з поточними мітками, І з
    # усіма збереженими синонімами (старі перейменовані назви + початковий
    # латинський "alt" для АЛЬТ), повертає САМЕ ПОТОЧНУ (канонічну) мітку.
    def resolve_payment_method(self, text):
        normalized = _normalize_phrase(text)
        if not normalized:
            return None
        for _option_id, label, _kind in self.list_payment_method_options():
            if _normalize_phrase(label) == normalized:
                return label
        for option_id, phrase in self.conn.execute(
            "SELECT option_id, phrase FROM payment_method_synonyms"
        ).fetchall():
            if _normalize_phrase(phrase) == normalized:
                option_row = self.conn.execute(
                    "SELECT label FROM payment_method_options WHERE id = ?", (option_id,)
                ).fetchone()
                if option_row:
                    return option_row[0]
        return None

    # Усі фрази, які МАЮТЬ розпізнаватись у вільному тексті (поточні мітки +
    # синоніми), кожна прив'язана до СВОЄЇ поточної (канонічної) мітки —
    # telegram_dialog.py будує з цього regex-патерни для багатослівних фраз
    # ("ЕФАКТУРА Б/Н") так само, як і раніше з PAYMENT_METHODS.
    def payment_method_recognized_phrases(self):
        options = self.list_payment_method_options()
        pairs = [(label, label) for _option_id, label, _kind in options]
        labels_by_id = {option_id: label for option_id, label, _kind in options}
        for option_id, phrase in self.conn.execute(
            "SELECT option_id, phrase FROM payment_method_synonyms"
        ).fetchall():
            current_label = labels_by_id.get(option_id)
            if current_label:
                pairs.append((phrase, current_label))
        return pairs

    # Одноразовий сідінг — точно сьогоднішні 3 назви (_DEFAULT_PAYMENT_
    # METHOD_OPTIONS) + латинський "alt" як синонім АЛЬТ (був хардкодженим
    # _PAYMENT_METHOD_ALIASES). Ідемпотентно: якщо таблиця вже має бодай
    # один рядок, нічого не робить (адмін уже міг додати/видалити/
    # перейменувати варіанти, повторний запуск не повинен це відкочувати).
    def _seed_payment_method_options(self):
        if self.conn.execute("SELECT 1 FROM payment_method_options LIMIT 1").fetchone():
            return
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            for position, label in enumerate(_DEFAULT_PAYMENT_METHOD_OPTIONS):
                cursor = self.conn.execute(
                    "INSERT INTO payment_method_options (label, position, created_at, updated_at) VALUES (?, ?, ?, ?)",
                    (label, position, now, now),
                )
                if label == "АЛЬТ":
                    self.conn.execute(
                        "INSERT INTO payment_method_synonyms (option_id, phrase, created_at) VALUES (?, ?, ?)",
                        (cursor.lastrowid, "alt", now),
                    )

    # Задача користувача (2026-08-16): "ці 6 користувачів уже увімкни в
    # дану версію, і обов'язково вони мають бути присутні при релізі" -
    # відомий реальний персонал (звірено по УСІХ доступних резервних
    # копіях на цій машині - один і той самий набір скрізь, жодних інших
    # ролей ніде не знайдено). Мета - щоб на щойно встановленій/скинутій
    # БД ці люди НЕ реєструвались як "Гость" при першому зверненні до
    # бота. INSERT OR IGNORE - ідемпотентно, ніколи не чіпає роль/дані
    # вже існуючого рядка (напр. якщо адміністратор уже вручну поміняв
    # комусь роль через клікабельний бейдж) - лише додає відсутніх.
    #
    # Задача користувача (2026-08-17): "в фінальний реліз 2 користувача
    # Ліна не мають входити. а поки тестимо - Ліна має входити своїми 2ма
    # користувачами" - обидва акаунти Ліни (2 telegram-акаунти однієї
    # людини) винесені в ОКРЕМИЙ список, що вмикається лише прапорцем
    # нижче. ПЕРЕД ЗБІРКОЮ ФІНАЛЬНОГО PRODUCTION-РЕЛІЗУ (не тестового) -
    # поставити _INCLUDE_TESTING_PERSONNEL = False.
    _INCLUDE_TESTING_PERSONNEL = True

    _KNOWN_PERSONNEL_SEED_PRODUCTION = (
        (8608586896, "", "Володимир", "admin"),
        (6228766189, "alexp_87", "Alex P", "admin"),
        (1158918079, "", "Володимир", "admin"),
        (7785125398, "", "Belpas Склад", "admin"),
    )
    _KNOWN_PERSONNEL_SEED_TESTING_ONLY = (
        (469711288, "", "Lina", "admin"),
        (1447000264, "", "Lina Kozachenko", "admin"),
    )

    def _seed_known_personnel(self):
        seed = self._KNOWN_PERSONNEL_SEED_PRODUCTION
        if self._INCLUDE_TESTING_PERSONNEL:
            seed = seed + self._KNOWN_PERSONNEL_SEED_TESTING_ONLY
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            for telegram_id, username, full_name, role in seed:
                self.conn.execute(
                    """
                    INSERT OR IGNORE INTO bot_users
                        (telegram_id, username, full_name, role, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (telegram_id, username, full_name, role, now, now),
                )

    # --- Категорії товару (bot_operations) — реальний CRUD ---

    _CYRILLIC_TO_LATIN_SLUG = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "i", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "",
        "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya", "і": "i", "ї": "i",
        "є": "e", "ґ": "g",
    }

    def _slugify_operation_code(self, text):
        chars = []
        for ch in text.lower():
            if ch in self._CYRILLIC_TO_LATIN_SLUG:
                chars.append(self._CYRILLIC_TO_LATIN_SLUG[ch])
            elif ch.isalnum() and ch.isascii():
                chars.append(ch)
            else:
                chars.append("_")
        slug = re.sub(r"_+", "_", "".join(chars)).strip("_")
        return slug or "category"

    def _unique_operation_code(self, parent_action_code, label):
        base = f"{parent_action_code.replace('start_', '')}_{self._slugify_operation_code(label)}"
        candidate = base
        suffix = 2
        while self.conn.execute("SELECT 1 FROM bot_operations WHERE code = ?", (candidate,)).fetchone():
            candidate = f"{base}_{suffix}"
            suffix += 1
        return candidate

    # Задача користувача: "додай можливість додавати... ці параметри
    # (кнопки)... для гнучкості налаштувань на майбутнє" — нова категорія
    # придатна до роботи ТОЧНО так само, як і 8 вбудованих (ДОСКА AD і
    # т.д.): той самий набір полів-запитів (товар/порода/товщина/ширина/
    # довжина + кількість/вимір, і для продажу — клієнт/ціна/сума/оплата).
    # requires_row_identity=1 завжди — нова категорія це такий самий
    # фізичний товар зі складу, як і всі існуючі; дії БЕЗ пошуку рядка
    # складу (kind='service', на кшталт антисептирования) — окремий,
    # спеціалізований сценарій, не пропонується як шлях додавання тут.
    def add_operation_category(self, parent_action_code, kind, label, product, condition=None):
        label = label.strip()
        product = product.strip()
        condition = condition.strip() if condition else None
        now = datetime.now().isoformat(timespec="seconds")
        code = self._unique_operation_code(parent_action_code, label)
        prefill = {"product": product}
        if condition:
            prefill["condition"] = condition
        with self.conn:
            position = self.conn.execute(
                "SELECT COALESCE(MAX(position), -1) + 1 FROM bot_operations WHERE parent_action_code = ?",
                (parent_action_code,),
            ).fetchone()[0]
            cursor = self.conn.execute(
                """
                INSERT INTO bot_operations
                    (code, kind, requires_row_identity, label, parent_action_code, prefill_json,
                     position, enabled, created_at, updated_at)
                VALUES (?, ?, 1, ?, ?, ?, ?, 1, ?, ?)
                """,
                (code, kind, label, parent_action_code, json.dumps(prefill, ensure_ascii=False), position, now, now),
            )
            operation_id = cursor.lastrowid
            self._seed_warehouse_identity_fields(operation_id, bool(condition), now)
            self._seed_quantity_measure_fields(operation_id, kind, now)
            if kind == "sale":
                self._seed_sale_ledger_fields(operation_id, now)
            return operation_id

    # Перейменування ЗБЕРІГАЄ стару назву як розпізнаваний синонім
    # назавжди (той самий принцип, що й update_payment_method_option) —
    # клієнт, який звик натискати/писати стару назву кнопки, не
    # "загубиться" після перейменування.
    def rename_operation_category(self, operation_id, new_label):
        new_label = new_label.strip()
        row = self.conn.execute("SELECT label FROM bot_operations WHERE id = ?", (operation_id,)).fetchone()
        if not row:
            return
        old_label = row[0]
        now = datetime.now().isoformat(timespec="seconds")
        with self.conn:
            if _normalize_phrase(old_label) != _normalize_phrase(new_label):
                self.conn.execute(
                    "INSERT INTO bot_operation_synonyms (operation_id, phrase, created_at) VALUES (?, ?, ?)",
                    (operation_id, old_label, now),
                )
            self.conn.execute(
                "UPDATE bot_operations SET label = ?, updated_at = ? WHERE id = ?",
                (new_label, now, operation_id),
            )

    # Видалення (на відміну від перейменування) прибирає категорію
    # ПОВНІСТЮ: каскадно й усі її поля-запити/прив'язки/синоніми (FK ON
    # DELETE CASCADE) — кнопка й розпізнавання зникають назавжди.
    def delete_operation_category(self, operation_id):
        with self.conn:
            # Той самий "осиротілий" ризик, що й у delete_operation вище —
            # категорія теж є рядком bot_operations, на який може вести
            # пряме посилання кастомної кнопки.
            self.conn.execute(
                "UPDATE custom_menu_buttons SET operation_id = NULL WHERE operation_id = ?",
                (operation_id,),
            )
            self.conn.execute("DELETE FROM bot_operations WHERE id = ?", (operation_id,))

    def operation_category_label_collides(self, parent_action_code, label, exclude_id=None):
        normalized = _normalize_phrase(label)
        for operation in self.list_operations(parent_action_code, include_disabled=True):
            operation_id = operation[0]
            op_label = operation[4]
            if exclude_id is not None and operation_id == exclude_id:
                continue
            if _normalize_phrase(op_label) == normalized:
                return True
        return False

    # Єдина точка "яка це категорія" — звіряє і з поточними мітками, і з
    # усіма збереженими синонімами (старі перейменовані назви + "osb"
    # латиницею для ОСБ, збережене від початкового хардкодженого
    # словника), повертає ID операції (або None, якщо нічого не збіглось).
    def resolve_operation_category(self, parent_action_code, text):
        normalized = _normalize_phrase(text)
        if not normalized:
            return None
        operations = self.list_operations(parent_action_code)
        for operation in operations:
            operation_id, _code, _kind, _requires_identity, op_label, *_rest = operation
            if _normalize_phrase(op_label) == normalized:
                return operation_id
        operation_ids = [operation[0] for operation in operations]
        if not operation_ids:
            return None
        placeholders = ",".join("?" for _ in operation_ids)
        for operation_id, phrase in self.conn.execute(
            f"SELECT operation_id, phrase FROM bot_operation_synonyms WHERE operation_id IN ({placeholders})",
            operation_ids,
        ).fetchall():
            if _normalize_phrase(phrase) == normalized:
                return operation_id
        return None

    # Одноразовий сідінг — "osb" (латиницею) як синонім для ОБОХ ОСБ-
    # операцій (income_osb/sale_osb): старий хардкоджений словник
    # _category_from_text мав "osb" і "осб" як ДВІ окремі мітки на один і
    # той самий (ОСБ, None), а мітка операції — лише "ОСБ" (кирилицею), яка
    # нормалізується в "осб", але НЕ в "osb". Без цього сідінгу клієнти, що
    # звикли писати латиницею, втратили б розпізнавання. Ідемпотентно:
    # перевіряє наявність цього конкретного синоніма перед вставкою.
    def _seed_operation_category_synonyms(self):
        now = datetime.now().isoformat(timespec="seconds")
        for builtin_key in ("income_osb", "sale_osb"):
            operation = self.get_operation_by_code(builtin_key)
            if not operation:
                continue
            operation_id = operation[0]
            exists = self.conn.execute(
                "SELECT 1 FROM bot_operation_synonyms WHERE operation_id = ? AND phrase = ?",
                (operation_id, "osb"),
            ).fetchone()
            if exists:
                continue
            with self.conn:
                self.conn.execute(
                    "INSERT INTO bot_operation_synonyms (operation_id, phrase, created_at) VALUES (?, ?, ?)",
                    (operation_id, "osb", now),
                )


# =============================================================================
# Бізнес-логіка приходу/продажу: запис операції в SQLite + (за потреби)
# синхронізація в Excel. Викликається з TelegramBotWorker, яка сама вирішує
# режим синхронізації (self._excel_sync_mode()) і передає його параметром
# sync_mode — цей модуль нічого не знає про Telegram чи налаштування бота.
# =============================================================================

# Толерантність порівняння кількості/об'єму з фактичним залишком на складі —
# щоб дрібні похибки округлення float не сприймались як "недостатньо товару".
INCOME_QUANTITY_TOLERANCE = 0.01
INCOME_VOLUME_TOLERANCE = 0.01

SALES_SHEET_NAME = "ПРОДАЖА МАТЕРИАЛА"
ANTISEPTIC_SHEET_NAME = "АНТИСЕПТИРОВАНИЕ"
# Реальний баг (2026-08-14): "приход завів, а його не видно ніде... тому що
# він не записується" — apply_income_operation ЗАВЖДИ оновлював лише СКЛАД
# (баланс) + stock_movements (внутрішній аудит-лог), і ніколи не писав
# журнальний рядок у вже наявний реальний лист Excel "ПРИХОД МАТЕРИАЛА" (той
# самий, вже задокументований у sheet_meta, з колонками майже ідентичними
# ПРОДАЖА МАТЕРИАЛА) — на відміну від продажу/списання/антисептирования, у
# кожного з яких давно є власний журнальний запис. Тепер прихід так само
# отримує рядок у цьому листі (income_columns/income_sheet_values нижче).
INCOME_SHEET_NAME = "ПРИХОД МАТЕРИАЛА"
# Задача користувача: "чому заміняє коментар? можливо придумати вкладку
# списання?" - на відміну від продажу/антисептирования, списання досі не
# мало власного листа-журналу взагалі (лише перезаписуване поле "Причина
# списания" на самому СКЛАД - єдине поточне значення, без історії). Новий
# лист з тим самим принципом, що вже мають ПРОДАЖА МАТЕРИАЛА/АНТИСЕПТИРОВАНИЕ.
WRITEOFF_SHEET_NAME = "СПИСАНИЕ"
# Задача користувача: "нащо в стовбцю дата час з секундами? прибери. лівіше
# додай окрему колонку точний час" - "Дата" лишається чистою датою (як і
# була, _parse_date_text завжди повертає .date()), а реальний момент
# створення запису (раніше приймався параметром now і ніде не
# використовувався) тепер іде в окрему колонку ЛІВОРУЧ від "Дата".
_WRITEOFF_TIME_HEADER = "Точное время"
_WRITEOFF_SHEET_HEADERS = [
    _WRITEOFF_TIME_HEADER,
    "Дата",
    "Документ",
    "Продукт",
    "Порода",
    "Состояние",
    "Толщина, мм",
    "Ширина, мм",
    "Длина, мм",
    "Количество, шт",
    "Причина списания",
    "Менеджер",
]

# Задача користувача (2026-08-14): "якщо цих стовпців немає в таблиці, то
# хай створить їх у першому вільному стовпці(цях)" - "Автор" для всіх
# чотирьох листів, де операція вже записується (Продажи/Списание/
# Антисептирование/Приход). На відміну від _WRITEOFF_TIME_HEADER вище
# (insert_cols(1) - свідомо ЗЛІВА, старіший випадок) - тут САМЕ "перший
# вільний" - тобто в КІНЕЦЬ (max_column + 1), не зсуваючи вже наявні
# колонки та їхні позиційні посилання.
_REQUIRED_OPERATION_AUTHOR_COLUMNS = [
    (WRITEOFF_SHEET_NAME, "Менеджер"),
    (SALES_SHEET_NAME, "Менеджер (итог)"),
    (ANTISEPTIC_SHEET_NAME, "Ответственный"),
    (INCOME_SHEET_NAME, "Менеджер"),
]

# Задача користувача (2026-08-14): "якщо приєднати порожній ексель, то має
# створитись программою красивенька табличка з відповідними вкладками" -
# точні заголовки нижче зняті напряму з реального test_sklad.xlsx (не
# вигадані/не виведені з fuzzy-match словників xxx_columns() вище - той
# порядок призначений лише для розпізнавання вже наявних колонок, не для
# відтворення "правильного" вигляду листа з нуля).
_SALES_SHEET_HEADERS = [
    "Дата",
    "Документ",
    "Клиент",
    "Продукт",
    "Порода",
    "Состояние",
    "Толщина, мм",
    "Ширина, мм",
    "Длина, мм",
    "Количество, шт",
    "Ввод вручную, м3",
    "Ввод вручную, м2",
    "Расчетный объем, м3",
    "Расчетная площадь, м2",
    "Итоговый объем, м3",
    "Итоговая площадь, м2",
    "Ед. цены",
    "Цена за ед.",
    "Сумма",
    "Код позиции (SKU)",
    "Форма оплаты",
    "Менеджер (итог)",
    "Комментарий",
    "Менеджер вручную",
    "Ввод вручную, мп",
    "Расчетный метраж, мп",
    "Итоговый метраж, мп",
    "Адрес выгрузки",
]

# АНТИСЕПТИРОВАНИЕ: реальний файл має інформаційний KPI-блок (формули)
# ПЕРЕД заголовками (рядки 1-6) - навмисно НЕ відтворюється тут (складні
# формули, легко зробити неправильно) - лист "з нуля" отримує заголовки
# ОДРАЗУ в рядку 1. _find_header_row (вище) шукає рядок за вмістом "Дата",
# тож новий лист із заголовком у рядку 1 розпізнається так само коректно.
_ANTISEPTIC_SHEET_HEADERS = [
    "Дата",
    "№ услуги",
    "Клиент",
    "Услуга",
    "Ед. изм.",
    "Объем, м3",
    "Цена за м3, MDL",
    "Стоимость, MDL",
    "Тип расчета",
    "Статус оплаты",
    "№ документа",
    "Приход наличных, MDL",
    "Приход по банку, MDL",
    "Отражение в расчетах",
    "Ответственный",
    "Комментарий",
    "Адрес выгрузки",
]

_INCOME_SHEET_HEADERS = [
    "Дата",
    "Документ",
    "Поставщик",
    "Продукт",
    "Порода",
    "Состояние",
    "Толщина, мм",
    "Ширина, мм",
    "Длина, мм",
    "Количество, шт",
    "Ввод вручную, м3",
    "Ввод вручную, м2",
    "Расчетный объем, м3",
    "Расчетная площадь, м2",
    "Итоговый объем, м3",
    "Итоговая площадь, м2",
    "Ед. цены",
    "Цена за ед.",
    "Сумма",
    "Код позиции (SKU)",
    "Комментарий",
    "Менеджер",
]

_WAREHOUSE_SHEET_HEADERS = [
    "Код позиции (SKU)",
    "Продукт",
    "Порода",
    "Состояние",
    "Толщина, мм",
    "Ширина, мм",
    "Длина, мм",
    "Основная ед. учета",
    "Начальный остаток, шт",
    "Начальный остаток, м3",
    "Начальный остаток, м2",
    "Приход, шт",
    "Приход, м3",
    "Приход, м2",
    "Продано, шт",
    "Продано, м3",
    "Продано, м2",
    "Остаток, шт",
    "Остаток, м3",
    "Остаток, м2",
    "Комментарий",
    "Начальный остаток, мп",
    "Приход, мп",
    "Продано, мп",
    "Остаток, мп",
]

# Порядок свідомо СКЛАД спочатку (продажі/списання/прихід посилаються на
# нього під час запису) - не критично для самого створення листів (кожен
# лист незалежний), але логічно найближче до того, як людина сама читала б
# структуру "спочатку склад, потім документи руху".
_REQUIRED_SHEETS_FULL = [
    ("СКЛАД", _WAREHOUSE_SHEET_HEADERS),
    (INCOME_SHEET_NAME, _INCOME_SHEET_HEADERS),
    (SALES_SHEET_NAME, _SALES_SHEET_HEADERS),
    (WRITEOFF_SHEET_NAME, _WRITEOFF_SHEET_HEADERS),
    (ANTISEPTIC_SHEET_NAME, _ANTISEPTIC_SHEET_HEADERS),
]


# Значення поза перелічeними в payment_method_options НЕ відкидаємо —
# повертаємо як є (вільний ввід лишається можливим, той самий принцип, що
# й раніше). store.resolve_payment_method — розпізнавання (поточні мітки +
# синоніми); ця функція — те, що ФАКТИЧНО пишеться в лист продажу/
# антисептирования, з тим самим фолбеком, що діяв до Кроку 4.4.
def normalize_payment_method(store, value):
    if not value:
        return value
    resolved = store.resolve_payment_method(value)
    return resolved if resolved is not None else value


def product_requires_type(product):
    normalized = _normalize_phrase(product)
    return normalized in {"доска", "doska"}


def display_product_name(payload):
    # Задача користувача: "ніяких КД АД в продукті, лише в состоянии" -
    # раніше сюди дописувалась condition (AD/КД) для "Доска", хоча вона й
    # так завжди записується окремо в колонку "Состояние" (set_value(...,
    # columns.get("condition"), payload.get("condition")) - всюди поруч із
    # цим викликом). product_requires_type лишається НЕ займаним - він і
    # далі потрібен окремо, щоб визначати, чи condition взагалі обов'язкове
    # поле для чек-листа цього товару.
    return payload.get("product") or ""


# Спільна форма рядків звіту "Низкий остаток" - той самий вигляд, що вже
# показує _render_low_stock_report (telegram_dialog_reports.py) і webapp/
# data.html "Низкий остаток". Живе тут (не в telegram_dialog_core.py), бо
# потрібна ОБОМ: боту (через CoreDialogMixin) і webapp_server.py (окремий
# HTTP-хендлер, не міксин) - webapp_server.py імпортувати telegram_dialog
# було б циклічним імпортом (telegram_dialog_core.py вже імпортує
# webapp_server).
def low_stock_report_rows(store, threshold):
    items = store.low_stock_warehouse_items(threshold)
    rows = []
    for item in items:
        size = "x".join(
            _display_bot_number(value)
            for value in (item.get("thickness"), item.get("width"), item.get("length"))
            if value not in (None, "")
        )
        rows.append({
            "product": _display_value(item.get("product")),
            "breed": _display_value(item.get("breed")),
            "condition": _display_value(item.get("condition")),
            "size": size,
            # Задача користувача: "фільтри мають бути всі схожими за їх
            # типами інформації" - та сама причина, що й у writeoff_report_
            # rows вище - "Низкий остаток" переиспользує ТОЙ САМИЙ модал
            # розміру, що й СКЛАД.
            "thickness": item.get("thickness"),
            "width": item.get("width"),
            "length": item.get("length"),
            "quantity": item.get("balance_qty"),
            "volume": item.get("balance_volume"),
            "area": item.get("balance_area"),
        })
    return rows


# Задача користувача: "додай вкладку списання... будемо бачити що і коли
# списали і чому" - той самий принцип, що вже має low_stock_report_rows
# вище (окрема, спільна функція в warehouse_data.py, не прив'язана лише до
# webapp - reusable і для бот-звіту, якщо колись знадобиться).
def writeoff_report_rows(store):
    headers = store.get_headers(WRITEOFF_SHEET_NAME)
    if not headers:
        return []
    columns = writeoff_columns(headers)
    rows = store.fetch_all_rows(WRITEOFF_SHEET_NAME)
    result = []
    for values in rows:
        size = "x".join(
            _display_bot_number(value)
            for value in (
                row_value(values, columns.get("thickness")),
                row_value(values, columns.get("width")),
                row_value(values, columns.get("length")),
            )
            if value not in (None, "")
        )
        raw_date = row_value(values, columns.get("date"))
        if isinstance(raw_date, datetime):
            display_date = raw_date.date().strftime("%d.%m.%Y")
        elif isinstance(raw_date, date):
            display_date = raw_date.strftime("%d.%m.%Y")
        else:
            display_date = raw_date or ""
        result.append({
            "date": display_date,
            "document": row_value(values, columns.get("document")),
            "product": row_value(values, columns.get("product")),
            "breed": row_value(values, columns.get("breed")),
            "condition": row_value(values, columns.get("condition")),
            "size": size,
            # Задача користувача: "фільтри мають бути всі схожими за їх
            # типами інформації" - окремі thickness/width/length (не лише
            # об'єднаний "size") потрібні, щоб вкладка "Списание" могла
            # переиспользувати ТОЙ САМИЙ модал розміру, що вже має СКЛАД.
            "thickness": row_value(values, columns.get("thickness")),
            "width": row_value(values, columns.get("width")),
            "length": row_value(values, columns.get("length")),
            "quantity": row_value(values, columns.get("quantity")),
            "reason": row_value(values, columns.get("reason")),
            "manager": row_value(values, columns.get("manager")),
        })
    result.reverse()
    return result


# Задача користувача (2026-08-14): "Приход" - нова вкладка "Данные склада".
# Той самий принцип, що й writeoff_report_rows вище - читає напряму з
# INCOME_SHEET_NAME (не з stock_movements - той лише службовий лог, ПРИХОД
# МАТЕРИАЛА тепер повноцінний документ-журнал, як і Продажи/Списание/
# Антисептирование).
def income_report_rows(store):
    headers = store.get_headers(INCOME_SHEET_NAME)
    if not headers:
        return []
    columns = income_columns(headers)
    rows = store.fetch_all_rows(INCOME_SHEET_NAME)
    result = []
    for values in rows:
        size = "x".join(
            _display_bot_number(value)
            for value in (
                row_value(values, columns.get("thickness")),
                row_value(values, columns.get("width")),
                row_value(values, columns.get("length")),
            )
            if value not in (None, "")
        )
        raw_date = row_value(values, columns.get("date"))
        if isinstance(raw_date, datetime):
            display_date = raw_date.date().strftime("%d.%m.%Y")
        elif isinstance(raw_date, date):
            display_date = raw_date.strftime("%d.%m.%Y")
        else:
            display_date = raw_date or ""
        result.append({
            "date": display_date,
            "document": row_value(values, columns.get("document")),
            "product": row_value(values, columns.get("product")),
            "breed": row_value(values, columns.get("breed")),
            "condition": row_value(values, columns.get("condition")),
            "size": size,
            "thickness": row_value(values, columns.get("thickness")),
            "width": row_value(values, columns.get("width")),
            "length": row_value(values, columns.get("length")),
            "quantity": row_value(values, columns.get("quantity")),
            "manager": row_value(values, columns.get("manager")),
        })
    result.reverse()
    return result


# Канонічні назви категорій товару (для кнопкового меню й пошуку). Реальні
# дані містять розбіжності в написанні (напр. "доска AD" і "Доска AD" в
# одному й тому ж листі СКЛАД) — тут зводимо їх до одного варіанту.
PRODUCT_CATEGORIES = {
    "доска ad": "Доска AD",
    "доска kd": "Доска KD",
    "осб": "ОСБ",
    "вагонка": "Вагонка",
}


def normalize_product_category(value):
    if not value:
        return value
    return PRODUCT_CATEGORIES.get(_normalize_phrase(value), value)


def sheet_product_name(payload):
    product = display_product_name(payload)
    return product or payload.get("product")


# Крок 4.4: Спосіб оплати тепер живе в payment_method_options (реальний
# CRUD — додати/перейменувати/видалити, GUI "Способи оплати") — цей список
# лишається ЛИШЕ як одноразові дані сідінгу (_seed_payment_method_options,
# ExcelSqliteStore) з тим самим узгодженим переліком, що й раніше. Для
# всіх способів дані вносяться однаково, різниця лише в цьому полі — воно
# потрібне ВИКЛЮЧНО для статистики/аналітики (обсяг продажів по
# безналу/готівці/АЛЬТ, фільтрація звітів) і НЕ впливає на розрахунок
# залишків складу, вартості продажу чи інші бізнес-процеси.
_DEFAULT_PAYMENT_METHOD_OPTIONS = ["ЕФАКТУРА Б/Н", "ЕФАКТУРА Н", "АЛЬТ"]


def warehouse_columns(headers):
    names = {
        "sku": ["Код позиции (SKU)", "SKU"],
        "product": ["Продукт"],
        "breed": ["Порода"],
        "condition": ["Состояние"],
        "thickness": ["Толщина, мм", "Толщина"],
        "width": ["Ширина, мм", "Ширина"],
        "length": ["Длина, мм", "Длинна, мм", "Длина", "Длинна"],
        "unit": ["Основная ед. учета"],
        "income_qty": ["Приход, шт"],
        "income_volume": ["Приход, м3"],
        "income_area": ["Приход, м2"],
        "income_linear": ["Приход, мп"],
        "sold_qty": ["Продано, шт", "Продажа, шт"],
        "sold_volume": ["Продано, м3", "Продажа, м3"],
        "sold_area": ["Продано, м2", "Продажа, м2"],
        "sold_linear": ["Продано, мп", "Продажа, мп"],
        "balance_qty": ["Остаток, шт"],
        "balance_volume": ["Остаток, м3"],
        "balance_area": ["Остаток, м2"],
        "balance_linear": ["Остаток, мп"],
        "writeoff_reason": ["Причина списания"],
    }
    normalized_headers = {
        _normalize_phrase(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    return {
        target: next(
            (
                normalized_headers[_normalize_phrase(candidate)]
                for candidate in candidates
                if _normalize_phrase(candidate) in normalized_headers
            ),
            None,
        )
        for target, candidates in names.items()
    }


def required_warehouse_columns(columns):
    required = {
        "product": "Продукт",
        "breed": "Порода",
        "thickness": "Толщина, мм",
        "width": "Ширина, мм",
        "length": "Длина, мм",
        "income_qty": "Приход, шт",
        "income_volume": "Приход, м3",
        "balance_qty": "Остаток, шт",
        "balance_volume": "Остаток, м3",
    }
    return [label for key, label in required.items() if columns.get(key) is None]


def required_sale_warehouse_columns(columns):
    required = {
        "product": "Продукт",
        "breed": "Порода",
        "thickness": "Толщина, мм",
        "width": "Ширина, мм",
        "length": "Длина, мм",
        "sold_qty": "Продано, шт",
        "sold_volume": "Продано, м3",
        "balance_qty": "Остаток, шт",
        "balance_volume": "Остаток, м3",
    }
    return [label for key, label in required.items() if columns.get(key) is None]


def warehouse_rows(store):
    headers = store.get_headers("СКЛАД")
    columns = warehouse_columns(headers)
    rows = store.fetch_rows("СКЛАД", 100000, 0)
    return headers, columns, rows


def sales_rows(store):
    headers = store.get_headers(SALES_SHEET_NAME)
    columns = sales_columns(headers)
    rows = store.fetch_rows(SALES_SHEET_NAME, 100000, 0)
    return headers, columns, rows


def antiseptic_rows(store):
    headers = store.get_headers(ANTISEPTIC_SHEET_NAME)
    columns = antiseptic_columns(headers)
    rows = store.fetch_rows(ANTISEPTIC_SHEET_NAME, 100000, 0)
    return headers, columns, rows


# Спільний хелпер для мега-форм (панель "Шаблони"/"Недавні") - викликається
# і з боту (telegram_dialog_core.py, кнопка "Сохранить шаблон" через
# web_app_data), і з локального API-сервера мега-форми (webapp_server.py,
# фонове збереження без закриття Mini App) - тому вільна функція тут, а не
# метод ExcelSqliteStore чи TelegramDialogMixin. Мітка категорії резолвиться
# щоразу заново (не зберігається в самих таблицях) - адмін міг перейменувати
# bot_operations.label з того часу.
def operation_template_entries(store, rows, source):
    entries = []
    for row in rows:
        (
            entry_id, category_operation_id, breed, thickness, width, length,
            client, address, payment_method,
        ) = row
        operation = store.get_operation(category_operation_id) if category_operation_id is not None else None
        if operation is None:
            continue
        entries.append({
            "id": entry_id,
            "source": source,
            "category_operation_id": category_operation_id,
            "category_label": operation[4],
            "breed": breed,
            "thickness": thickness,
            "width": width,
            "length": length,
            "client": client,
            "address": address,
            "payment_method": payment_method,
        })
    return entries


def row_value(row, index):
    # row=None означає "немає рядка складу" (наприклад продаж послуги
    # антисептирования без прив'язки до конкретної позиції складу) —
    # трактуємо так само, як відсутню колонку: порожнє значення, не падіння.
    if row is None or index is None or index >= len(row):
        return ""
    return row[index]


def add_to_row_value(row_values, index, amount):
    if index is None:
        return
    while len(row_values) <= index:
        row_values.append("")
    row_values[index] = _number_value(row_values[index]) + _number_value(amount)


def set_value(row_values, index, value):
    if index is None:
        return
    while len(row_values) <= index:
        row_values.append("")
    row_values[index] = value


# Крок 3+ "Дії": за товаром/типом (payload["product"]/payload.get("condition"))
# знаходить, ЯКА саме сіданa дія (bot_operations) відповідає поточній
# операції — БЕЗ жодної зміни діалогу бота (кнопки категорій і так уже
# заповнюють саме ці два поля, telegram_dialog.py._category_from_text).
# Товар, що НЕ збігається із жодною з 4 заведених категорій (вільний текст
# поза кнопками — бот дозволяє будь-який товар, що вже є на складі) —
# повертає None; викликач тоді має впасти назад на сьогоднішню жорстко
# закодовану поведінку (Задача користувача: "наразі всі внутрішні дії
# мають відповідати теперішнім налаштованим" — для товару поза категоріями
# сьогодні немає конкретної "дії", тож і немає що редагувати).
def resolve_operation_for_payload(store, parent_action_code, kind, payload):
    product = payload.get("product")
    if not product:
        return None
    condition = payload.get("condition")
    for row in store.list_operations(parent_action_code):
        operation_id, _code, op_kind, _requires_identity, _label, _parent, prefill_json, _position, _enabled, _builtin = row
        if op_kind != kind or not prefill_json:
            continue
        prefill = json.loads(prefill_json)
        if _normalize_phrase(prefill.get("product") or "") != _normalize_phrase(product):
            continue
        prefill_condition = prefill.get("condition")
        if prefill_condition and _normalize_phrase(prefill_condition) != _normalize_phrase(condition or ""):
            continue
        return operation_id
    return None


# Крок 3+ "Дії": для яких семантичних колонок execute_operation_write
# взагалі має що робити (write_mode='generic', проста дельта на вже
# знайденому рядку СКЛАД) — і яке значення з item брати для кожної. Лише
# ці 12 ключів колись з'являються з marker='add'/'subtract' у сідінгу
# (_seed_quantity_measure_fields) — решта (product/breed/... і все на
# ПРОДАЖА МАТЕРИАЛА/АНТИСЕПТИРОВАНИЕ) завжди write_mode='ledger'.
_OPERATION_FIELD_ITEM_KEYS = {
    "income_qty": "quantity", "sold_qty": "quantity", "balance_qty": "quantity",
    "income_volume": "volume", "sold_volume": "volume", "balance_volume": "volume",
    "income_area": "area", "sold_area": "area", "balance_area": "area",
    "income_linear": "linear", "sold_linear": "linear", "balance_linear": "linear",
}


# Універсальний виконавець запису для дій (bot_operations) — Крок 3+ "Дії".
# Для кожного увімкненого поля-запиту операції, для кожної його прив'язки з
# write_mode='generic' на вказаній вкладці, бере значення з item за
# семантикою колонки (_OPERATION_FIELD_ITEM_KEYS) і викликає add_to_row_value
# зі знаком за marker ('add'/'subtract'). Значення None (напр. "area" для
# товару, що рахується в м3) просто пропускається — так рівно ОДНА з трьох
# пар income_*/balance_* (volume/area/linear) реально спрацьовує на кожен
# рядок, без потреби окремо визначати тут, яка саме одиниця виміру зараз
# застосовна (це вже вирішено раніше, у telegram_dialog.py._row_measure_kind,
# і відображено лише в тому, яке з полів item не None).
# write_mode='ledger'-прив'язки (усе на ПРОДАЖА МАТЕРИАЛА/АНТИСЕПТИРОВАНИЕ)
# ця функція НЕ чіпає — їх і далі записує sale_sheet_values/
# antiseptic_sheet_values (там є бізнес-логіка: сума, розподіл готівка/банк,
# нумерація документа тощо, а не проста мапа поле->колонка).
def execute_operation_write(store, operation_id, item, row_values, columns, sheet="СКЛАД"):
    for field in store.list_operation_fields(operation_id):
        field_id = field[0]
        for binding in store.list_operation_field_columns(field_id):
            _, _, binding_sheet, column_key, marker, write_mode, _, _ = binding
            if write_mode != "generic" or binding_sheet != sheet:
                continue
            item_key = _OPERATION_FIELD_ITEM_KEYS.get(column_key)
            if item_key is None:
                continue
            value = item.get(item_key)
            if value is None:
                continue
            column_index = columns.get(column_key)
            if column_index is None:
                continue
            amount = _number_value(value) if marker == "add" else -_number_value(value)
            add_to_row_value(row_values, column_index, amount)


def income_item_size(item):
    return (
        f"{_display_bot_number(item['thickness'])}x"
        f"{_display_bot_number(item['width'])}x"
        f"{_display_bot_number(item['length'])}"
    )


# "area"/"linear" (мп, розміри 25x50/30x50/50x50) взаємовиключні з "volume"
# і одне з одним — telegram_dialog.py._row_measure_kind гарантує, що для
# рядка встановлюється РІВНО ОДНЕ з трьох полів; тут лише зчитуємо, яке
# саме, для запису/показу.
def item_measure_kind(item):
    if item.get("area") is not None:
        return "area"
    if item.get("linear") is not None:
        return "linear"
    if item.get("volume") is not None:
        return "volume"
    # ОСБ (і будь-який інший товар без фізичного виміру): жодне з трьох
    # полів ніколи не заповнюється — рядок рахується напряму по кількості.
    return None


ITEM_MEASURE_UNIT = {"volume": "м3", "area": "м2", "linear": "мп"}

_BALANCE_COLUMN_BY_MEASURE_KIND = {"volume": "balance_volume", "area": "balance_area", "linear": "balance_linear"}


# Задача користувача (2026-08-09, скріншот "Списание записано: ... -122 шт,
# -2,6352 м3"): "хочу бачити у всіх місцях, після підрахунків якихось -
# результат: після операції залишилось стільки-то штук та одиниці
# вимірювання" — той самий рядок складу вже мутований (add_to_row_value/
# execute_operation_write) ДО викликів цієї функції, тож тут лише формат
# показу вже готового значення. "Якщо од. вимірювання і є штуки" (ОСБ,
# measure_kind=None) — лише шт, без вигаданого "0 м3".
def _remaining_balance_text(row_values, columns, measure_kind):
    balance_qty = _number_value(row_value(row_values, columns.get("balance_qty")))
    if measure_kind is None:
        return f"{_display_bot_number(balance_qty)} шт"
    balance_column = columns.get(_BALANCE_COLUMN_BY_MEASURE_KIND[measure_kind])
    balance_measure = _number_value(row_value(row_values, balance_column))
    return f"{_display_bot_number(balance_qty)} шт, {_display_bot_number(balance_measure)} {ITEM_MEASURE_UNIT[measure_kind]}"


def income_item_known_size(item):
    values = [item.get(field) for field in ("thickness", "width", "length")]
    if all(value in (None, "") or _number_value(value) <= 0 for value in values):
        return "размер"
    # Завжди 3 позиції товщина x ширина x довжина, "?" на місці ще
    # невідомого виміру — інакше часткові розміри різних рядків при показі
    # разом (_recognized_data_lines) виглядають як безглузді голі числа
    # ("26, 50x5000"), а не як зрозумілі товщина/ширина/довжина.
    return "x".join(
        _display_bot_number(value) if value not in (None, "") and _number_value(value) > 0 else "?"
        for value in values
    )


def sale_position_text(payload, item):
    return (
        f"{display_product_name(payload) or '?'} / "
        f"{payload.get('breed') or '?'} / "
        f"{income_item_size(item)}"
    )


def sales_columns(headers):
    names = {
        "date": ["Дата"],
        "document": ["Документ"],
        "client": ["Клиент"],
        "address": ["Адрес выгрузки"],
        "product": ["Продукт"],
        "breed": ["Порода"],
        "condition": ["Состояние"],
        "thickness": ["Толщина, мм", "Толщина"],
        "width": ["Ширина, мм", "Ширина"],
        "length": ["Длина, мм", "Длинна, мм", "Длина", "Длинна"],
        "quantity": ["Количество, шт"],
        "manual_volume": ["Ввод вручную, м3"],
        "calculated_volume": ["Расчетный объем, м3"],
        "total_volume": ["Итоговый объем, м3"],
        "manual_area": ["Ввод вручную, м2"],
        "calculated_area": ["Расчетная площадь, м2"],
        "total_area": ["Итоговая площадь, м2"],
        "manual_linear": ["Ввод вручную, мп"],
        "calculated_linear": ["Расчетный метраж, мп"],
        "total_linear": ["Итоговый метраж, мп"],
        "price_unit": ["Ед. цены"],
        "price_per_unit": ["Цена за ед."],
        "total_amount": ["Сумма"],
        "payment_method": ["Форма оплаты"],
        "manager_final": ["Менеджер (итог)"],
        "sku": ["Код позиции (SKU)", "SKU"],
        "comment": ["Комментарий"],
        "manual_manager": ["Менеджер вручную"],
    }
    normalized_headers = {
        _normalize_phrase(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    return {
        target: next(
            (
                normalized_headers[_normalize_phrase(candidate)]
                for candidate in candidates
                if _normalize_phrase(candidate) in normalized_headers
            ),
            None,
        )
        for target, candidates in names.items()
    }


def sale_sheet_values(store, payload, item, warehouse_row, warehouse_columns_map, now, document_number=None):
    headers = store.get_headers(SALES_SHEET_NAME)
    values = [""] * len(headers)
    columns = sales_columns(headers)
    user = payload.get("user") or {}
    document_time = datetime.fromisoformat(now).strftime("%Y%m%d-%H%M%S")
    sale_date = _parse_date_text(payload.get("date")) or date.today()
    set_value(values, columns.get("date"), sale_date)
    set_value(
        values,
        columns.get("document"),
        payload.get("document_type") or document_number or f"Telegram продажа {document_time}",
    )
    set_value(values, columns.get("client"), payload.get("client"))
    set_value(values, columns.get("address"), payload.get("address"))
    set_value(
        values,
        columns.get("product"),
        row_value(warehouse_row, warehouse_columns_map["product"]) or sheet_product_name(payload),
    )
    set_value(
        values,
        columns.get("breed"),
        row_value(warehouse_row, warehouse_columns_map["breed"]) or payload.get("breed"),
    )
    set_value(values, columns.get("thickness"), item.get("thickness"))
    set_value(values, columns.get("width"), item.get("width"))
    set_value(values, columns.get("length"), item.get("length"))
    set_value(values, columns.get("quantity"), item.get("quantity"))
    is_area = item.get("area") not in (None, "")
    is_linear = item.get("linear") not in (None, "")
    is_volume = item.get("volume") not in (None, "")
    if item.get("volume_provided"):
        set_value(values, columns.get("manual_volume"), item.get("volume"))
    set_value(values, columns.get("calculated_volume"), item.get("volume"))
    set_value(values, columns.get("total_volume"), item.get("volume"))
    if is_area:
        if item.get("area_provided"):
            set_value(values, columns.get("manual_area"), item.get("area"))
        set_value(values, columns.get("calculated_area"), item.get("area"))
        set_value(values, columns.get("total_area"), item.get("area"))
    if is_linear:
        if item.get("linear_provided"):
            set_value(values, columns.get("manual_linear"), item.get("linear"))
        set_value(values, columns.get("calculated_linear"), item.get("linear"))
        set_value(values, columns.get("total_linear"), item.get("linear"))
    # ОСБ (і будь-який інший товар без фізичного виміру): жодне з трьох
    # is_area/is_linear/is_volume не спрацьовує — рахуємо/показуємо "шт".
    if is_linear:
        price_unit = "мп"
    elif is_area:
        price_unit = "м2"
    elif is_volume:
        price_unit = "м3"
    else:
        price_unit = "шт"
    set_value(values, columns.get("price_unit"), price_unit)
    set_value(values, columns.get("sku"), row_value(warehouse_row, warehouse_columns_map["sku"]))

    price_per_unit = payload.get("price_per_unit")
    set_value(values, columns.get("price_per_unit"), price_per_unit)
    total_amount = payload.get("total_amount")
    if total_amount in (None, "") and price_per_unit not in (None, ""):
        if is_linear:
            measure_for_price = item.get("linear")
        elif is_area:
            measure_for_price = item.get("area")
        elif is_volume:
            measure_for_price = item.get("volume")
        else:
            measure_for_price = item.get("quantity")
        # Аудит коду: раніше рахувалось без округлення — у "Сумма" й Excel
        # потрапляли значення на кшталт 100.31233000000001 (float-шум від
        # множення). _priced_amount (utils.py) - спільна формула, щоб та
        # сама помилка не могла знову розійтись між sale/income/antiseptic.
        total_amount = _priced_amount(price_per_unit, measure_for_price)
    set_value(values, columns.get("total_amount"), total_amount)
    set_value(values, columns.get("payment_method"), normalize_payment_method(store, payload.get("payment_method")))

    manager = payload.get("manager") or user.get("full_name") or user.get("username")
    set_value(values, columns.get("manager_final"), manager)
    set_value(values, columns.get("manual_manager"), user.get("full_name") or user.get("username"))

    comment_parts = [f"Telegram: {payload.get('original_text', '')}"]
    if payload.get("comment"):
        comment_parts.append(payload["comment"])
    set_value(values, columns.get("comment"), " | ".join(part for part in comment_parts if part))
    return values


# INCOME_SHEET_NAME — той самий "документ + позиція" журнал, що вже мають
# ПРОДАЖА МАТЕРИАЛА/СПИСАНИЕ/АНТИСЕПТИРОВАНИЕ, лише без клієнта/оплати
# (прихід — не грошова операція для покупця) і з "Поставщик" замість
# "Клиент" (бот сьогодні постачальника не питає — лишається порожнім, як і
# ціна/сумма, доки цей крок не буде явно доданий до потоку приходу).
def income_columns(headers):
    names = {
        "date": ["Дата"],
        "document": ["Документ"],
        "supplier": ["Поставщик"],
        "product": ["Продукт"],
        "breed": ["Порода"],
        "condition": ["Состояние"],
        "thickness": ["Толщина, мм", "Толщина"],
        "width": ["Ширина, мм", "Ширина"],
        "length": ["Длина, мм", "Длинна, мм", "Длина", "Длинна"],
        "quantity": ["Количество, шт"],
        "manual_volume": ["Ввод вручную, м3"],
        "calculated_volume": ["Расчетный объем, м3"],
        "total_volume": ["Итоговый объем, м3"],
        "manual_area": ["Ввод вручную, м2"],
        "calculated_area": ["Расчетная площадь, м2"],
        "total_area": ["Итоговая площадь, м2"],
        "price_unit": ["Ед. цены"],
        "price_per_unit": ["Цена за ед."],
        "total_amount": ["Сумма"],
        "sku": ["Код позиции (SKU)", "SKU"],
        "comment": ["Комментарий"],
        # Задача користувача (2026-08-14): "Автор" - хто провів прихід -
        # той самий принцип, що вже мають Продажи/Списание/Антисептирование
        # ("Менеджер"/"Менеджер (итог)"/"Ответственный").
        "manager": ["Менеджер"],
    }
    normalized_headers = {
        _normalize_phrase(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    return {
        target: next(
            (
                normalized_headers[_normalize_phrase(candidate)]
                for candidate in candidates
                if _normalize_phrase(candidate) in normalized_headers
            ),
            None,
        )
        for target, candidates in names.items()
    }


def income_sheet_values(store, payload, item, warehouse_row, warehouse_columns_map, now, document_number=None):
    headers = store.get_headers(INCOME_SHEET_NAME)
    values = [""] * len(headers)
    columns = income_columns(headers)
    income_date = _parse_date_text(payload.get("date")) or date.today()
    set_value(values, columns.get("date"), income_date)
    set_value(values, columns.get("document"), document_number)
    set_value(values, columns.get("supplier"), payload.get("supplier"))
    set_value(
        values,
        columns.get("product"),
        row_value(warehouse_row, warehouse_columns_map["product"]) or sheet_product_name(payload),
    )
    set_value(
        values,
        columns.get("breed"),
        row_value(warehouse_row, warehouse_columns_map["breed"]) or payload.get("breed"),
    )
    set_value(values, columns.get("condition"), payload.get("condition"))
    set_value(values, columns.get("thickness"), item.get("thickness"))
    set_value(values, columns.get("width"), item.get("width"))
    set_value(values, columns.get("length"), item.get("length"))
    set_value(values, columns.get("quantity"), item.get("quantity"))
    is_area = item.get("area") not in (None, "")
    is_volume = item.get("volume") not in (None, "")
    if item.get("volume_provided"):
        set_value(values, columns.get("manual_volume"), item.get("volume"))
    set_value(values, columns.get("calculated_volume"), item.get("volume"))
    set_value(values, columns.get("total_volume"), item.get("volume"))
    if is_area:
        if item.get("area_provided"):
            set_value(values, columns.get("manual_area"), item.get("area"))
        set_value(values, columns.get("calculated_area"), item.get("area"))
        set_value(values, columns.get("total_area"), item.get("area"))
    # ОСБ і мп-розміри (сьогодні цей лист не має власних мп-колонок, на
    # відміну від ПРОДАЖА МАТЕРИАЛА) — кількість/розміри й так записані
    # вище, лише об'єм/площа лишаються порожніми, як і мало бути.
    if is_area:
        price_unit = "м2"
    elif is_volume:
        price_unit = "м3"
    else:
        price_unit = "шт"
    set_value(values, columns.get("price_unit"), price_unit)
    set_value(values, columns.get("sku"), row_value(warehouse_row, warehouse_columns_map["sku"]))

    price_per_unit = payload.get("price_per_unit")
    set_value(values, columns.get("price_per_unit"), price_per_unit)
    total_amount = payload.get("total_amount")
    if total_amount in (None, "") and price_per_unit not in (None, ""):
        if is_area:
            measure_for_price = item.get("area")
        elif is_volume:
            measure_for_price = item.get("volume")
        else:
            measure_for_price = item.get("quantity")
        # _priced_amount (utils.py) - та сама формула, що й у
        # sale_sheet_values/antiseptic_sheet_values (аудит коду, 2026-08-14
        # - раніше 4 незалежні копії цього ж округлення в 3 функціях цього
        # файлу + 1 у telegram_dialog_income_sale_parsing.py).
        total_amount = _priced_amount(price_per_unit, measure_for_price)
    set_value(values, columns.get("total_amount"), total_amount)

    comment_parts = [f"Telegram: {payload.get('original_text', '')}"]
    if payload.get("comment"):
        comment_parts.append(payload["comment"])
    set_value(values, columns.get("comment"), " | ".join(part for part in comment_parts if part))

    user = payload.get("user") or {}
    set_value(values, columns.get("manager"), user.get("full_name") or user.get("username"))
    return values


def insert_sheet_row(store, sheet_name, values, now):
    position = store.conn.execute(
        "SELECT COALESCE(MAX(position), 0) + 1 FROM sheet_rows WHERE sheet_name = ?",
        (sheet_name,),
    ).fetchone()[0]
    store.conn.execute(
        """
        INSERT INTO sheet_rows (sheet_name, position, values_json, updated_at)
        VALUES (?, ?, ?, ?)
        """,
        (sheet_name, position, _serialize_row(values), now),
    )


def sync_excel_after_operation(sync_mode, store, sheet_names, dirty_notifier=None):
    if sync_mode != "after_each_operation":
        return None
    # Задача користувача (2026-08-13): раніше цей виклик синхронно
    # переписував ВЕСЬ лист ПІД ЧАС відповіді боту на кожну операцію -
    # dirty_notifier (TelegramBotWorker.mark_excel_dirty, main.py) лише
    # позначає лист і повертається миттю, справжній пакетний запис
    # відбувається пізніше, у поллінг-циклі (TelegramBotWorker._excel_
    # sync_tick). Без notifier (напр. прямий виклик з тестів) - лишається
    # старий синхронний шлях нижче, без зміни поведінки.
    if dirty_notifier is not None:
        dirty_notifier.mark_excel_dirty(sheet_names)
        return None
    try:
        sync_sheets_to_excel(store, sheet_names)
    except PermissionError:
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Закройте файл Excel и нажмите «Обновити Excel»."
        )
    except OSError:
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Проверьте, что файл доступен, и нажмите «Обновити Excel»."
        )
    except RuntimeError:
        # Аудит коду: excel_source.py кидає RuntimeError, коли джерело
        # (локальний файл не обрано чи онлайн-сеанс застарів) взагалі не
        # готове до запису — раніше цей виняток НЕ ловився тут і летів
        # непійманим крізь apply_sale_operation/apply_income_operation,
        # хоча сам запис у SQLite до цього моменту вже успішно закомічений.
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Проверьте настройку источника Excel-таблицы в программе."
        )
    return None


# Товарні позиції продажу (Задача користувача: кілька видів пиломатеріалу
# в одну продажу -> один підсумковий звіт бухгалтеру) — completed_positions,
# якщо є, інакше ОДНА позиція, синтезована зі старих плоских полів payload
# (product/condition/breed/rows/price_per_unit), для повної зворотної
# сумісності з усіма викликами apply_sale_operation, що ще не знають про
# completed_positions (у т.ч. прямі виклики з тестів).
def _sale_positions(payload):
    positions = payload.get("completed_positions")
    if positions:
        return positions
    return [{
        "product": payload.get("product"),
        "condition": payload.get("condition"),
        "breed": payload.get("breed"),
        "rows": payload.get("rows") or [],
        "price_per_unit": payload.get("price_per_unit"),
    }]


# Задача користувача (2026-08-14): "щоб міг продовжувати приход і внести
# кілька різних позицій. так же як це реалізовано в реалізації" - той самий
# принцип, що й _sale_positions вище (completed_positions, якщо кошик
# накопичив кілька позицій; інакше одна поточна "позиція" з полів верхнього
# рівня payload - той самий формат, що вже й був до мультипозиційності).
def _income_positions(payload):
    positions = payload.get("completed_positions")
    if positions:
        return positions
    return [{
        "product": payload.get("product"),
        "condition": payload.get("condition"),
        "breed": payload.get("breed"),
        "rows": payload.get("rows") or [],
        "price_per_unit": payload.get("price_per_unit"),
        "supplier": payload.get("supplier"),
    }]


# Важлива знахідка нового аудиту (28.07.2026, #4) — персистентний лічильник
# номера документа замість len(fetch_rows(...)). Викликати ЛИШЕ УСЕРЕДИНІ
# вже відкритої BEGIN IMMEDIATE-транзакції (як і саме читання/запис
# залишку) — інакше два паралельні записи знову можуть отримати той самий
# номер. current_count_fallback — лише для ПЕРШОГО виклику на ще не
# засіяному лічильнику: продовжує з поточного рахунку рядків, БЕЗ
# ретроактивної переномерації вже виданих номерів.
def _next_document_number(store, sheet_name, current_count_fallback):
    row = store.conn.execute(
        "SELECT next_number FROM document_counters WHERE sheet_name = ?", (sheet_name,)
    ).fetchone()
    if row is None:
        next_number = current_count_fallback + 1
        store.conn.execute(
            "INSERT INTO document_counters (sheet_name, next_number) VALUES (?, ?)",
            (sheet_name, next_number + 1),
        )
        return next_number
    next_number = row[0]
    store.conn.execute(
        "UPDATE document_counters SET next_number = ? WHERE sheet_name = ?",
        (next_number + 1, sheet_name),
    )
    return next_number


# Задача користувача (2026-08-17): "Сумма позиции ... зроби грубим шрифтом" -
# усі 4 підсумкові повідомлення (apply_sale/income/writeoff/antiseptic_
# operation) тепер надсилаються з parse_mode="HTML" (main.py._send_message).
# Це означає, що БУДЬ-ЯКИЙ вільний текст усередині ("<", ">", "&" зокрема)
# ламає відправку всього повідомлення, якщо не екранований - _esc() тут
# обгортає html.escape() для КОЖНОГО динамічного значення (ім'я клієнта,
# адреса, коментар, назва товару/породи), яке потрапляє в ці 4 функції.
def _esc(value):
    return html.escape(str(value)) if value is not None else ""


def apply_sale_operation(store, payload, sync_mode, dirty_notifier=None):
    headers, columns, _ = warehouse_rows(store)
    missing_columns = required_sale_warehouse_columns(columns)
    if missing_columns:
        return {
            "ok": False,
            "message": (
                "Не удалось записать продажу: в листе СКЛАД не найдены нужные колонки: "
                f"{', '.join(missing_columns)}."
            ),
        }

    positions = _sale_positions(payload)
    now = datetime.now().isoformat(timespec="seconds")

    # BEGIN IMMEDIATE одразу, ДО читання залишків — не голий BEGIN
    # (deferred), який набуває лише SHARED-блокування, поки не почнеться
    # запис. Реальний баг (TOCTOU-гонка з аудиту): раніше читання залишку +
    # перевірка "чи вистачає" відбувались ДО BEGIN узагалі, тож два
    # паралельні продажі (два продавці, або бот + окреме з'єднання GUI на
    # тому самому файлі) могли обидва прочитати ОДИН і той самий залишок,
    # обидва пройти перевірку — і обидва закомітити, зробивши залишок
    # від'ємним, ба гірше — store.update_row переписує ВЕСЬ рядок повністю,
    # тож другий commit тихо загубив би й зовсім не пов'язані з продажем
    # правки першого. BEGIN IMMEDIATE набуває RESERVED-блокування одразу,
    # тому інше з'єднання не може розпочати СВІЙ запис, поки це не
    # завершиться (commit/rollback) — читання залишку й перевірка "чи
    # вистачає" всередині цього самого блокування більше не можуть
    # розійтись із фактичним записом.
    updated = 0
    with store.conn:
        store.conn.execute("BEGIN IMMEDIATE")

        # Прохід валідації: тут ще нічого не пишеться в БД. row_values_by_row_id
        # накопичує стан кожного рядка після продажу в пам'яті, щоб дві позиції
        # з однаковим row_id перевірялись проти актуального (ще не збереженого)
        # залишку. Проходимо ВСІ товарні позиції продажу (positions), не лише
        # одну — так само, як і одна позиція раніше.
        row_values_by_row_id = {}
        for position in positions:
            position_payload = {**payload, **position}
            # Аудит коду (перевірка охоплення Fix #9): перевірка нижче на
            # від'ємні quantity/measure не покривала price_per_unit —
            # antiseptic_sheet_values цю перевірку вже має (реверсна ціна
            # сума÷об'єм), а apply_sale_operation ні. Від'ємна ціна не
            # псує залишок складу (те відповідає окрема перевірка нижче),
            # але тихо записала б від'ємну "Цена"/"Сумма" у ПРОДАЖА
            # МАТЕРИАЛА - той самий клас похибки обліку, що Fix #9 вже
            # закривав для об'єму антисептирования.
            price_per_unit_value = _number_value(position_payload.get("price_per_unit"))
            total_amount_value = _number_value(position_payload.get("total_amount"))
            if price_per_unit_value < 0 or total_amount_value < 0:
                return {
                    "ok": False,
                    "message": "Не удалось записать продажу: цена и сумма не могут быть отрицательными.",
                }
            # Свіжий пере-аудит (2026-08-02): price_per_unit/total_amount -
            # поле "або/або" (той самий has_price, що вже перевіряє чек-лист
            # у telegram_dialog_income_sale_flow.py - БУДЬ-ЯКЕ з двох
            # достатнє) - відхиляємо лише коли ОБИДВА не додатні, інакше
            # звичайна продажа "за 5000 MDL" (без явної ціни за одиницю)
            # помилково відхилялась б.
            if price_per_unit_value <= 0 and total_amount_value <= 0:
                return {
                    "ok": False,
                    "message": "Не удалось записать продажу: должна быть указана цена за единицу или сумма.",
                }
            for item in position["rows"]:
                row_id = item.get("row_id")
                if row_id is None:
                    # Рядок-послуга (наприклад антисептирование без конкретної
                    # позиції складу) — не прив'язаний до жодного рядка складу,
                    # тож немає що перевіряти/списувати тут. Просто пропускаємо
                    # валідацію залишку.
                    continue
                row_values = row_values_by_row_id.get(row_id)
                if row_values is None:
                    row_values = store.get_row(row_id)
                    if not row_values:
                        return {
                            "ok": False,
                            "message": (
                                "Не удалось записать продажу: позиция больше не найдена на складе.\n"
                                f"Позиция: {sale_position_text(position_payload, item)}"
                            ),
                        }
                    row_values_by_row_id[row_id] = row_values

                is_area = item.get("area") is not None
                is_linear = item.get("linear") is not None
                measure_key = "area" if is_area else ("linear" if is_linear else "volume")
                # Аудит коду: перевірка нижче ("чи вистачає залишку") НІКОЛИ
                # не відхиляє від'ємне значення — від'ємне завжди МЕНШЕ за
                # будь-який додатний залишок, тож проходить перевірку, а
                # запис нижче (add_to_row_value) цим самим ЗБІЛЬШИВ би
                # залишок складу, хоча в обліку записалось б як "продаж".
                # Другий рубіж захисту на рівні шару даних — незалежно від
                # того, чи бот сьогодні взагалі пропускає такі значення сюди.
                if _number_value(item.get("quantity")) <= 0:
                    return {
                        "ok": False,
                        "message": "Не удалось записать продажу: количество не может быть отрицательным или равным нулю.",
                    }
                # Свіжий пере-аудит (2026-08-02): вимір перевіряємо лише коли
                # він РЕАЛЬНО є в цієї позиції (item_measure_kind, а не
                # measure_key вище, що завжди дефолтиться на "volume") - для
                # ОСБ (кількість без фізичного виміру, вже відвантажена
                # фіча) 0 у "volume" - коректний, очікуваний стан, не помилка.
                measure_kind = item_measure_kind(item)
                if measure_kind is not None and _number_value(item.get(measure_kind)) <= 0:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать продажу: объём/площадь/погонные метры "
                            "не могут быть отрицательными или равными нулю."
                        ),
                    }
                balance_qty = _number_value(row_value(row_values, columns["balance_qty"]))
                if _number_value(item.get("quantity")) > balance_qty + INCOME_QUANTITY_TOLERANCE:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать продажу: на складе уже недостаточно штук.\n"
                            f"Доступно: {_display_bot_number(balance_qty)} шт."
                        ),
                    }
                if is_area:
                    balance_area = _number_value(row_value(row_values, columns.get("balance_area")))
                    if _number_value(item.get("area")) > balance_area + INCOME_VOLUME_TOLERANCE:
                        return {
                            "ok": False,
                            "message": (
                                "Не удалось записать продажу: на складе уже недостаточно площади.\n"
                                f"Доступно: {_display_bot_number(balance_area)} м2."
                            ),
                        }
                elif is_linear:
                    balance_linear = _number_value(row_value(row_values, columns.get("balance_linear")))
                    if _number_value(item.get("linear")) > balance_linear + INCOME_VOLUME_TOLERANCE:
                        return {
                            "ok": False,
                            "message": (
                                "Не удалось записать продажу: на складе уже недостаточно погонных метров.\n"
                                f"Доступно: {_display_bot_number(balance_linear)} мп."
                            ),
                        }
                else:
                    balance_volume = _number_value(row_value(row_values, columns["balance_volume"]))
                    if _number_value(item.get("volume")) > balance_volume + INCOME_VOLUME_TOLERANCE:
                        return {
                            "ok": False,
                            "message": (
                                "Не удалось записать продажу: на складе уже недостаточно объема.\n"
                                f"Доступно: {_display_bot_number(balance_volume)} м3."
                            ),
                        }

                # Крок 3+ "Дії": якщо продукт/тип збігається із заведеною
                # дією (ДОСКА AD/KD/ОСБ/ВАГОНКА) — запис веде конфігурований
                # execute_operation_write (адміністратор може редагувати, які
                # колонки +/- через Редактор кнопок); інакше — сьогоднішня
                # жорстко закодована поведінка (товар поза 4 категоріями).
                operation_id = resolve_operation_for_payload(store, "start_sale", "sale", position_payload)
                if operation_id is not None:
                    execute_operation_write(store, operation_id, item, row_values, columns)
                else:
                    add_to_row_value(row_values, columns["sold_qty"], item["quantity"])
                    add_to_row_value(row_values, columns["balance_qty"], -_number_value(item["quantity"]))
                    if is_area:
                        add_to_row_value(row_values, columns.get("sold_area"), item["area"])
                        add_to_row_value(row_values, columns.get("balance_area"), -_number_value(item["area"]))
                    elif is_linear:
                        add_to_row_value(row_values, columns.get("sold_linear"), item["linear"])
                        add_to_row_value(row_values, columns.get("balance_linear"), -_number_value(item["linear"]))
                    # Реальний ризик (аудит коду, 2026-08-14): раніше тут
                    # був голий else: - для товару без фізичного виміру
                    # (ОСБ і подібні, поза 4 налаштованими "Діями" -
                    # трапляється, якщо адміністратор видалить/вимкне
                    # категорію в редакторі кнопок) item може взагалі не
                    # мати ключа "volume" -> KeyError. Списание/"новий
                    # рядок" приходу вже мають цей самий guard
                    # (item.get("volume") is not None) - тут його бракувало.
                    elif item.get("volume") is not None:
                        add_to_row_value(row_values, columns["sold_volume"], item["volume"])
                        add_to_row_value(row_values, columns["balance_volume"], -_number_value(item["volume"]))

        # Один номер документа на ВЕСЬ продаж (усі позиції/розміри однієї
        # операції), а не по одному на кожен рядок листа — рахується
        # усередині ТІЄЇ Ж транзакції (після BEGIN IMMEDIATE). Персистентний
        # _next_document_number (не len(fetch_rows(...))+1) — імунний і до
        # паралельних записів, і до видалення рядка через GUI (аудит 28.07.2026, #4).
        existing_sale_count = len(store.fetch_rows(SALES_SHEET_NAME, 100000, 0))
        document_number = f"Продажа №{_next_document_number(store, SALES_SHEET_NAME, existing_sale_count)}"

        # Прохід запису: усі позиції вже пройшли валідацію, тож кожен запис
        # нижче відбудеться. store.update_row (не голий SQL) — та сама точка
        # входу, що й ручне редагування в GUI, тож warehouse_items
        # синхронізується автоматично. was_in_transaction всередині
        # store.update_row бачить, що BEGIN IMMEDIATE вже відкрито, і не
        # комітить сам — атомарність усього запису лишається на цьому
        # зовнішньому with store.conn:.
        for row_id, row_values in row_values_by_row_id.items():
            store.update_row(row_id, row_values)

        for position in positions:
            position_payload = {**payload, **position}
            rows_in_position = position["rows"]
            for item in rows_in_position:
                row_values = row_values_by_row_id.get(item.get("row_id"))
                item_payload = position_payload
                if len(rows_in_position) > 1:
                    # Реальний баг (аудит коду, 2026-08-14): position["total_amount"]
                    # (_archive_current_sale_position_and_reset,
                    # telegram_dialog_income_sale_flow.py:511-522) - це СУМА за
                    # ВСІ рядки позиції разом (_sale_total_amount), потрібна лише
                    # для тексту підтвердження ("Сумма позиции" один раз на
                    # позицію). sale_sheet_values нижче писала б це саме число
                    # буквально в "Сумма" КОЖНОГО фізичного рядка листа ПРОДАЖА
                    # МАТЕРИАЛА - при кількох розмірах в одній позиції сума
                    # виходила завищена в N разів (N = рядків), і в самому
                    # листі, і в усіх звітах/PDF/Excel, що з нього рахують
                    # підсумок. Прибираємо total_amount лише для позицій з
                    # ДЕКІЛЬКОМА рядками - тоді вже наявний fallback у
                    # sale_sheet_values порахує частку САМЕ ЦЬОГО рядка
                    # (price_per_unit × його власний вимір). Для одного рядка
                    # лишаємо як є - там total_amount і так дорівнює своїй
                    # єдиній частці.
                    item_payload = {**position_payload, "total_amount": None}
                sale_values = sale_sheet_values(
                    store, item_payload, item, row_values, columns, now, document_number
                )
                insert_sheet_row(store, SALES_SHEET_NAME, sale_values, now)

                user = payload.get("user") or {}
                store.add_stock_movement(
                    {
                        "movement_type": "sale",
                        "source": "telegram",
                        "telegram_user_id": user.get("id"),
                        "username": user.get("username"),
                        "full_name": user.get("full_name"),
                        "product": sheet_product_name(position_payload),
                        "breed": position_payload.get("breed"),
                        "condition": position_payload.get("condition"),
                        "thickness": item.get("thickness"),
                        "width": item.get("width"),
                        "length": item.get("length"),
                        "quantity": item.get("quantity"),
                        "volume": item.get("volume"),
                        "area": item.get("area"),
                        "linear": item.get("linear"),
                        "sheet_row_id": item.get("row_id"),
                        "original_text": payload.get("original_text"),
                        "created_at": now,
                    }
                )
                updated += 1

    excel_warning = sync_excel_after_operation(sync_mode, store, ["СКЛАД", SALES_SHEET_NAME], dirty_notifier)

    # Задача користувача (2026-08-17): "сума має бути у всіх одиничних
    # операціях реалізація/антисептирование... Состояние має показуватись
    # скрізь. якщо состояній декілька - групувати" - прибрано стару розвилку
    # show_position_headers (короткий/детальний вивід залежно від кількості
    # позицій, Задача 67) - тепер ЗАВЖДИ один шаблон: заголовок позиції
    # (Продукт/Порода/Состояние), сума позиції, підсумок продажу. Позиції з
    # ОДНАКОВИМ product/breed/condition зливаються в один блок з одним
    # спільним "Сумма позиции" замість повтору того самого заголовка
    # (вибраний користувачем варіант із трьох показаних мокапів).
    groups = []
    group_index_by_key = {}
    for position in positions:
        position_payload = {**payload, **position}
        key = (position_payload.get("product"), position_payload.get("breed"), position_payload.get("condition"))
        if key not in group_index_by_key:
            group_index_by_key[key] = len(groups)
            groups.append({
                "payload": position_payload,
                "rows": [],
                "total_amount": 0.0,
                "antiseptic_volume": 0.0,
                "antiseptic_sum": 0.0,
            })
        group = groups[group_index_by_key[key]]
        group["rows"].extend(position["rows"])
        group["total_amount"] += _number_value(position.get("total_amount"))
        antiseptic_addon = position.get("antiseptic")
        has_antiseptic = isinstance(antiseptic_addon, dict) and antiseptic_addon.get("volume") and antiseptic_addon.get("price_per_unit")
        if has_antiseptic:
            antiseptic_volume = _number_value(antiseptic_addon.get("volume"))
            antiseptic_price = _number_value(antiseptic_addon.get("price_per_unit"))
            group["antiseptic_volume"] += antiseptic_volume
            group["antiseptic_sum"] += round(antiseptic_volume * antiseptic_price, 2)

    lines = ["Продажа записана:"]
    index = 0
    grand_total = 0.0
    grand_total_goods = 0.0
    grand_total_antiseptic = 0.0
    for group_number, group in enumerate(groups, start=1):
        position_payload = group["payload"]
        if group_number > 1:
            lines.append("")
        header_parts = [_esc(display_product_name(position_payload)), _esc(position_payload.get("breed"))]
        if position_payload.get("condition"):
            header_parts.append(_esc(position_payload["condition"]))
        lines.append(f"Позиция: {' / '.join(part for part in header_parts if part)}")
        for item in group["rows"]:
            index += 1
            measure_kind = item_measure_kind(item)
            if item.get("row_id") is None:
                # Рядок-послуга — розмірів немає, показуємо назву товару/послуги
                # замість "толщинаxширинаxдлина", без мінуса (склад не списаний).
                position_label = _esc(sheet_product_name(position_payload))
                if measure_kind is None:
                    lines.append(f"{index}. {position_label}: {_display_bot_number(item.get('quantity'))} шт")
                else:
                    measure_value = item.get(measure_kind)
                    measure_unit = ITEM_MEASURE_UNIT[measure_kind]
                    lines.append(
                        f"{index}. {position_label}: "
                        f"{_display_bot_number(item.get('quantity'))} шт, "
                        f"{_display_bot_number(measure_value)} {measure_unit}"
                    )
            else:
                row_values = row_values_by_row_id.get(item["row_id"])
                remaining_suffix = (
                    f" (Осталось: {_esc(_remaining_balance_text(row_values, columns, measure_kind))})"
                    if row_values is not None
                    else ""
                )
                if measure_kind is None:
                    lines.append(
                        f"{index}. {_esc(income_item_size(item))}: -"
                        f"{_display_bot_number(item['quantity'])} шт{remaining_suffix}"
                    )
                else:
                    measure_value = item.get(measure_kind)
                    measure_unit = ITEM_MEASURE_UNIT[measure_kind]
                    lines.append(
                        f"{index}. {_esc(income_item_size(item))}: -"
                        f"{_display_bot_number(item['quantity'])} шт, -"
                        f"{_display_bot_number(measure_value)} {measure_unit}{remaining_suffix}"
                    )
        goods_total = round(group["total_amount"], 2)
        antiseptic_sum = round(group["antiseptic_sum"], 2)
        if antiseptic_sum:
            if goods_total:
                lines.append(f"  <b>Сумма за товар: {_display_bot_number(goods_total)} MDL</b>")
            lines.append("  Дополнительная услуга:")
            lines.append(
                f"  Антисептирование: {_display_bot_number(round(group['antiseptic_volume'], 2))} м3 — "
                f"{_display_bot_number(antiseptic_sum)} MDL"
            )
            lines.append(f"  <b>Сумма позиции: {_display_bot_number(round(goods_total + antiseptic_sum, 2))} MDL</b>")
        elif goods_total:
            lines.append(f"  <b>Сумма позиции: {_display_bot_number(goods_total)} MDL</b>")
        grand_total += goods_total + antiseptic_sum
        grand_total_goods += goods_total
        grand_total_antiseptic += antiseptic_sum
    lines.append("")
    client = payload.get("client")
    if client:
        lines.append(f"Клиент: {_esc(client)}")
    address = payload.get("address")
    if address:
        lines.append(f"Адрес выгрузки: {_esc(address)}")
    payment_method = normalize_payment_method(store, payload.get("payment_method"))
    if payment_method:
        lines.append(f"Оплата: {_esc(payment_method)}")
    if grand_total:
        lines.append("")
        # Задача користувача: "хочу бачити загалом за антисепт і загалом за
        # товар, а вже в кінці итог" - той самий принцип, що й у
        # _sale_preview (telegram_dialog_income_sale_flow.py): "Сумма за
        # Антисептирование" лише коли вона реально є, "Сумма за товар"
        # завжди поруч з підсумковим "Итого по всей продаже".
        if grand_total_antiseptic:
            lines.append(f"<b>Сумма за Антисептирование: {_display_bot_number(round(grand_total_antiseptic, 2))} MDL</b>")
        lines.append(f"<b>Сумма за товар: {_display_bot_number(round(grand_total_goods, 2))} MDL</b>")
        lines.append(f"<b>Итого по всей продаже: {_display_bot_number(round(grand_total, 2))} MDL</b>")
    lines.append("")
    lines.append("✅ Выполнено.")
    if excel_warning:
        lines.append("")
        lines.append(_esc(excel_warning))
    return {"ok": True, "message": "\n".join(lines)}


def apply_income_operation(store, payload, sync_mode, dirty_notifier=None):
    headers, columns, _ = warehouse_rows(store)
    # Задача користувача (2026-08-14): "щоб міг продовжувати приход і
    # внести кілька різних позицій. так же як це реалізовано в реалізації" -
    # positions[] тепер може нести КІЛЬКА різних товарів/порід (не лише
    # кілька розмірів ОДНОГО товару, як payload["rows"] раніше). Той самий
    # принцип, що вже має apply_sale_operation (_sale_positions) - при
    # одній позиції (звичний випадок) поведінка НЕ змінюється взагалі.
    positions = _income_positions(payload)

    # Аудит коду (виправлення знайденого при перевірці посилок): ця
    # перевірка раніше стояла ВСЕРЕДИНІ циклу запису нижче — при кількох
    # рядках приходу невалідний НЕ ПЕРШИЙ рядок означав, що попередні
    # (валідні) рядки вже встигали пройти store.update_row/add_row +
    # add_stock_movement і ЗАКОМІТИТИСЬ (return з-під `with store.conn:`
    # комітить транзакцію, а не відкочує її) — перш ніж цикл дійшов би до
    # відхилення. Користувач бачив "не вдалося записати", а частина
    # приходу вже тихо потрапляла на склад. Тепер це окремий прохід ПО
    # ВСІХ рядках УСІХ позицій ДО відкриття транзакції — так само, як уже
    # робить apply_sale_operation (валідація окремо від запису).
    for position in positions:
        for item in position["rows"]:
            is_area = item.get("area") is not None
            is_linear = item.get("linear") is not None
            measure_key = "area" if is_area else ("linear" if is_linear else "volume")
            if _number_value(item.get("quantity")) <= 0:
                return {
                    "ok": False,
                    "message": "Не удалось записать приход: количество не может быть отрицательным или равным нулю.",
                }
            # Свіжий пере-аудит (2026-08-02): той самий OSБ-виняток, що й у
            # apply_sale_operation - вимір перевіряємо лише коли item_measure_
            # kind реально не None (measure_key вище завжди дефолтиться на
            # "volume", навіть для ОСБ, де це поле ніколи не заповнюється).
            measure_kind = item_measure_kind(item)
            if measure_kind is not None and _number_value(item.get(measure_kind)) <= 0:
                return {
                    "ok": False,
                    "message": (
                        "Не удалось записать приход: объём/площадь/погонные метры "
                        "не могут быть отрицательными или равными нулю."
                    ),
                }

    updated = 0
    created = 0
    now = datetime.now().isoformat(timespec="seconds")
    with store.conn:
        # Той самий TOCTOU-фікс, що й у apply_sale_operation/apply_
        # antiseptic_operation: раніше тут був голий "BEGIN" (deferred —
        # блокування набувається лише на першому записі), тож читання
        # поточного залишку (store.get_row) + запис нового (store.
        # update_row) НЕ були одним нероздільним кроком — два паралельних
        # приходи в ОДИН і той самий рядок могли обидва прочитати той самий
        # старий залишок і обидва закомітити, тихо загубивши накопичення
        # одного з них (нижчий ризик, ніж у продажу — тут завжди ДОДАЄМО,
        # тож негативний залишок не виникне, але частина приходу губиться).
        # BEGIN IMMEDIATE набуває RESERVED-блокування одразу, тому інше
        # з'єднання не може почати свій запис, поки це не завершиться.
        store.conn.execute("BEGIN IMMEDIATE")
        # Важлива знахідка нового аудиту (28.07.2026, #5): раніше тут не
        # перевірялось, чи row_id досі існує (напр. видалений через GUI між
        # тим, як бот знайшов рядок, і підтвердженням користувача) —
        # store.get_row повертав би [], store.update_row мовчки оновлював
        # би 0 рядків (SQLite не кидає винятку на 0-рядковий UPDATE), а
        # функція й далі рапортувала "Приход записан" з фантомним updated
        # += 1. Той самий шаблон, що вже має apply_sale_operation: окрема
        # перевірка ВСЕРЕДИНІ цієї ж транзакції, ПЕРЕД будь-яким записом
        # (не окремим проходом до BEGIN IMMEDIATE, як негативні quantity/
        # measure вище — існування рядка залежить від поточного стану БД,
        # тож перевірка має бути атомарною з самим записом).
        row_values_by_row_id = {}
        for position in positions:
            for item in position["rows"]:
                row_id = item.get("row_id")
                if row_id is None:
                    continue
                row_values = store.get_row(row_id)
                if not row_values:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать приход: позиция больше не найдена на складе.\n"
                            f"Позиция: {income_item_size(item)}"
                        ),
                    }
                row_values_by_row_id[row_id] = row_values

        for position in positions:
            # Крок 3+ "Дії": усі рядки ОДНІЄЇ позиції мають той самий
            # товар/тип (position, не по-рядково) — резолвимо ОДИН раз НА
            # ПОЗИЦІЮ (не на весь виклик - позиції мультиприходу можуть
            # бути РІЗНИМИ товарами, той самий принцип, що й apply_sale_
            # operation).
            position_payload = {**payload, **position}
            operation_id = resolve_operation_for_payload(store, "start_income", "income", position_payload)
            for item in position["rows"]:
                is_area = item.get("area") is not None
                is_linear = item.get("linear") is not None
                sheet_row_id = item.get("row_id")
                if item.get("row_id"):
                    row_values = row_values_by_row_id[item["row_id"]]
                    if operation_id is not None:
                        execute_operation_write(store, operation_id, item, row_values, columns)
                    else:
                        add_to_row_value(row_values, columns["income_qty"], item["quantity"])
                        add_to_row_value(row_values, columns["balance_qty"], item["quantity"])
                        if is_area:
                            add_to_row_value(row_values, columns.get("income_area"), item["area"])
                            add_to_row_value(row_values, columns.get("balance_area"), item["area"])
                        elif is_linear:
                            add_to_row_value(row_values, columns.get("income_linear"), item["linear"])
                            add_to_row_value(row_values, columns.get("balance_linear"), item["linear"])
                        # Той самий guard, що вже має гілка "новий рядок"
                        # нижче (item.get("volume") is not None) - раніше
                        # тут був голий else:, несумісний з товаром без
                        # ключа "volume" (аудит коду, 2026-08-14).
                        elif item.get("volume") is not None:
                            add_to_row_value(row_values, columns["income_volume"], item["volume"])
                            add_to_row_value(row_values, columns["balance_volume"], item["volume"])
                    store.update_row(item["row_id"], row_values)
                    updated += 1
                else:
                    values = [""] * len(headers)
                    sheet_product = sheet_product_name(position_payload)
                    set_value(values, columns.get("product"), sheet_product)
                    set_value(values, columns.get("breed"), position_payload.get("breed"))
                    # Реальний баг (2026-08-14): "чому в них типу немає? я ж
                    # в приході виставляв" - при СТВОРЕННІ нового рядка
                    # складу "Состояние" (Тип) ніколи не записувався, хоча
                    # condition уже був заповнений з приходу - колонка
                    # лишалась порожньою назавжди (наступні прихід/продаж
                    # теж її не чіпають, лише кількість). Разом із фіксом
                    # _warehouse_row_matches (telegram_dialog_income_sale_
                    # flow.py) це й було причиною "не просумовані" - без
                    # цього запису наступний прихід того самого розміру/
                    # породи/типу не міг знайти вже створений рядок і завжди
                    # створював ще один.
                    set_value(values, columns.get("condition"), position_payload.get("condition"))
                    set_value(values, columns.get("thickness"), item["thickness"])
                    set_value(values, columns.get("width"), item["width"])
                    set_value(values, columns.get("length"), item["length"])
                    if is_area:
                        unit_label = "м2"
                    elif is_linear:
                        unit_label = "мп"
                    elif item.get("volume") is not None:
                        unit_label = "м3"
                    else:
                        unit_label = "шт"
                    set_value(values, columns.get("unit"), unit_label)
                    # Свіжий пере-аудит (New-Notable #6): гілка "рядок
                    # ЗНАЙДЕНО" (вище) уже викликає execute_operation_write,
                    # коли є bot_operations-конфігурація для цієї категорії -
                    # ця гілка ("новий рядок") донедавна завжди писала
                    # жорстко, ігноруючи будь-яку кастомізацію адміна через
                    # "Дії" саме для ПЕРШОГО приходу нового розміру/породи.
                    # execute_operation_write на щойно побудованому
                    # порожньому рядку дає ідентичний результат прямому
                    # set_value (add_to_row_value: _number_value("") == 0,
                    # тож 0 + amount == amount).
                    if operation_id is not None:
                        execute_operation_write(store, operation_id, item, values, columns)
                    else:
                        set_value(values, columns.get("income_qty"), item["quantity"])
                        set_value(values, columns.get("balance_qty"), item["quantity"])
                        if is_area:
                            set_value(values, columns.get("income_area"), item["area"])
                            set_value(values, columns.get("balance_area"), item["area"])
                        elif is_linear:
                            set_value(values, columns.get("income_linear"), item["linear"])
                            set_value(values, columns.get("balance_linear"), item["linear"])
                        elif item.get("volume") is not None:
                            set_value(values, columns.get("income_volume"), item["volume"])
                            set_value(values, columns.get("balance_volume"), item["volume"])
                    size = income_item_size(item)
                    sku = f"{sheet_product}|{position_payload.get('breed')}|{size}"
                    set_value(values, columns.get("sku"), sku)
                    sheet_row_id = store.add_row("СКЛАД", values)
                    item["row_id"] = sheet_row_id
                    # Задача користувача: "після операції залишилось
                    # стільки-то штук" — рядок ще НЕМАЄ в row_values_by_
                    # row_id (він заповнювався лише для ВЖЕ ІСНУЮЧИХ рядків
                    # вище), а message-loop нижче читає саме з нього — той
                    # самий словник обслуговує обидві гілки (знайдено/
                    # створено).
                    row_values_by_row_id[sheet_row_id] = values
                    created += 1

                user = payload.get("user") or {}
                store.add_stock_movement(
                    {
                        "movement_type": "income",
                        "source": "telegram",
                        "telegram_user_id": user.get("id"),
                        "username": user.get("username"),
                        "full_name": user.get("full_name"),
                        "product": sheet_product_name(position_payload),
                        "breed": position_payload.get("breed"),
                        "condition": position_payload.get("condition"),
                        "thickness": item.get("thickness"),
                        "width": item.get("width"),
                        "length": item.get("length"),
                        "quantity": item.get("quantity"),
                        "volume": item.get("volume"),
                        "area": item.get("area"),
                        "linear": item.get("linear"),
                        "sheet_row_id": sheet_row_id,
                        "original_text": payload.get("original_text"),
                        "created_at": now,
                    }
                )

        # Той самий персистентний _next_document_number, що вже мають
        # продаж/списання/антисептирование — один номер на ВЕСЬ виклик (усі
        # позиції цього приходу), не по одному на кожен рядок.
        existing_income_count = len(store.fetch_rows(INCOME_SHEET_NAME, 100000, 0))
        income_document_number = f"Приход №{_next_document_number(store, INCOME_SHEET_NAME, existing_income_count)}"
        for position in positions:
            position_payload = {**payload, **position}
            for item in position["rows"]:
                income_sheet_row = income_sheet_values(
                    store, position_payload, item, row_values_by_row_id[item["row_id"]], columns, now,
                    income_document_number,
                )
                insert_sheet_row(store, INCOME_SHEET_NAME, income_sheet_row, now)

    excel_warning = sync_excel_after_operation(sync_mode, store, ["СКЛАД", INCOME_SHEET_NAME], dirty_notifier)

    # Задача користувача (2026-08-17): "Состояние має показуватись скрізь.
    # якщо состояній декілька - групувати" - той самий принцип, що вже
    # застосований в apply_sale_operation: єдиний заголовок "Позиция:
    # Продукт / Порода / Состояние", позиції з ОДНАКОВИМ product/breed/
    # condition зливаються в один блок замість повтору заголовка (раніше
    # одна позиція мала окремий 3-рядковий блок "Продукт:/Порода:/
    # Состояние:", кілька позицій - "Позиция: Товар / Порода" без
    # Состояние; тепер один спільний формат для будь-якої кількості).
    groups = []
    group_index_by_key = {}
    for position in positions:
        position_payload = {**payload, **position}
        key = (position_payload.get("product"), position_payload.get("breed"), position_payload.get("condition"))
        if key not in group_index_by_key:
            group_index_by_key[key] = len(groups)
            groups.append({"payload": position_payload, "rows": []})
        groups[group_index_by_key[key]]["rows"].extend(position["rows"])

    lines = ["Приход записан:", ""]
    index = 0
    for group_number, group in enumerate(groups, start=1):
        position_payload = group["payload"]
        if group_number > 1:
            lines.append("")
        header_parts = [_esc(display_product_name(position_payload)), _esc(position_payload.get("breed"))]
        if position_payload.get("condition"):
            header_parts.append(_esc(position_payload["condition"]))
        lines.append(f"Позиция: {' / '.join(part for part in header_parts if part)}")
        # "Всего" замість "Осталось" для приходу свідомо - тут завжди
        # ДОДАЄМО на склад, тож підсумковий залишок логічніше назвати
        # "всього", а не "залишилось" (те слово лишається лише для
        # продажу/списання, де залишок справді ЗМЕНШУЄТЬСЯ).
        for item in group["rows"]:
            index += 1
            measure_kind = item_measure_kind(item)
            row_values = row_values_by_row_id.get(item.get("row_id"))
            remaining_suffix = (
                f" (Всего: {_esc(_remaining_balance_text(row_values, columns, measure_kind))})"
                if row_values is not None
                else ""
            )
            if measure_kind is None:
                lines.append(
                    f"{index}. {_esc(income_item_size(item))}: +{_display_bot_number(item['quantity'])} шт{remaining_suffix}"
                )
                continue
            measure_value = item.get(measure_kind)
            measure_unit = ITEM_MEASURE_UNIT[measure_kind]
            lines.append(
                f"{index}. {_esc(income_item_size(item))}: +"
                f"{_display_bot_number(item['quantity'])} шт, +"
                f"{_display_bot_number(measure_value)} {measure_unit}{remaining_suffix}"
            )
    lines.append("")
    lines.append(f"Обновлено позиций: {updated}")
    if created:
        lines.append(f"Создано новых позиций: {created}")
    lines.append("")
    lines.append("✅ Выполнено.")
    if excel_warning:
        lines.append("")
        lines.append(_esc(excel_warning))
    return {"ok": True, "message": "\n".join(lines)}


# Списання (writeoff) — "прихід навпаки": та сама ідентифікація рядка складу
# (порода/товщина/ширина/довжина), що й прихід/продаж, але БЕЗ клієнта/ціни/
# оплати. На відміну від приходу — немає гілки "рядок не знайдено ->
# створити новий" (не можна списати те, чого ніколи не було на складі); на
# відміну від продажу — немає кількох товарних позицій/накопичення, ціни чи
# способу оплати.
# Задача користувача: "чому заміняє коментар? можливо придумати вкладку
# списання?" — тепер, окрім оновлення СКЛАД (як і раніше) і stock_movements
# (внутрішній аудит-лог), кожне списання ще й пишеться окремим рядком у
# WRITEOFF_SHEET_NAME — та сама, вже перевірена схема "документ + позиція",
# що вже мають ПРОДАЖА МАТЕРИАЛА/АНТИСЕПТИРОВАНИЕ.
def writeoff_columns(headers):
    names = {
        "time": [_WRITEOFF_TIME_HEADER],
        "date": ["Дата"],
        "document": ["Документ"],
        "product": ["Продукт"],
        "breed": ["Порода"],
        "condition": ["Состояние"],
        "thickness": ["Толщина, мм", "Толщина"],
        "width": ["Ширина, мм", "Ширина"],
        "length": ["Длина, мм", "Длинна, мм", "Длина", "Длинна"],
        "quantity": ["Количество, шт"],
        "reason": ["Причина списания"],
        "manager": ["Менеджер"],
    }
    normalized_headers = {
        _normalize_phrase(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    return {
        target: next(
            (
                normalized_headers[_normalize_phrase(candidate)]
                for candidate in candidates
                if _normalize_phrase(candidate) in normalized_headers
            ),
            None,
        )
        for target, candidates in names.items()
    }


def writeoff_sheet_values(store, payload, item, now, document_number=None):
    headers = store.get_headers(WRITEOFF_SHEET_NAME)
    values = [""] * len(headers)
    columns = writeoff_columns(headers)
    set_value(values, columns.get("time"), datetime.fromisoformat(now).time())
    writeoff_date = _parse_date_text(payload.get("date")) or date.today()
    set_value(values, columns.get("date"), writeoff_date)
    set_value(values, columns.get("document"), document_number)
    set_value(values, columns.get("product"), sheet_product_name(payload))
    set_value(values, columns.get("breed"), payload.get("breed"))
    set_value(values, columns.get("condition"), payload.get("condition"))
    set_value(values, columns.get("thickness"), item.get("thickness"))
    set_value(values, columns.get("width"), item.get("width"))
    set_value(values, columns.get("length"), item.get("length"))
    set_value(values, columns.get("quantity"), item.get("quantity"))
    set_value(values, columns.get("reason"), payload.get("comment"))
    user = payload.get("user") or {}
    set_value(values, columns.get("manager"), user.get("full_name") or user.get("username"))
    return values


def apply_writeoff_operation(store, payload, sync_mode, dirty_notifier=None):
    headers, columns, _ = warehouse_rows(store)
    for item in payload["rows"]:
        if _number_value(item.get("quantity")) <= 0:
            return {
                "ok": False,
                "message": "Не удалось записать списание: количество не может быть отрицательным или равным нулю.",
            }
        # Той самий ОСБ-виняток, що й у apply_sale_operation/apply_income_
        # operation: вимір перевіряємо лише коли він РЕАЛЬНО є в цієї
        # позиції (item_measure_kind) - для ОСБ (кількість без фізичного
        # виміру) 0 у "volume" - коректний, очікуваний стан, не помилка.
        measure_kind = item_measure_kind(item)
        if measure_kind is not None and _number_value(item.get(measure_kind)) <= 0:
            return {
                "ok": False,
                "message": (
                    "Не удалось записать списание: объём/площадь/погонные метры "
                    "не могут быть отрицательными или равными нулю."
                ),
            }

    now = datetime.now().isoformat(timespec="seconds")
    updated = 0
    with store.conn:
        # Той самий TOCTOU-фікс (BEGIN IMMEDIATE), що й у apply_sale_
        # operation/apply_income_operation — читання залишку й перевірка
        # "чи вистачає" мають бути в одній нероздільній транзакції з
        # фактичним записом.
        store.conn.execute("BEGIN IMMEDIATE")
        operation_id = resolve_operation_for_payload(store, "start_writeoff", "writeoff", payload)

        row_values_by_row_id = {}
        for item in payload["rows"]:
            row_id = item.get("row_id")
            if row_id is None:
                continue
            row_values = row_values_by_row_id.get(row_id)
            if row_values is None:
                row_values = store.get_row(row_id)
                if not row_values:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать списание: позиция больше не найдена на складе.\n"
                            f"Позиция: {income_item_size(item)}"
                        ),
                    }
                row_values_by_row_id[row_id] = row_values

            is_area = item.get("area") is not None
            is_linear = item.get("linear") is not None
            balance_qty = _number_value(row_value(row_values, columns["balance_qty"]))
            if _number_value(item.get("quantity")) > balance_qty + INCOME_QUANTITY_TOLERANCE:
                return {
                    "ok": False,
                    "message": (
                        "Не удалось записать списание: на складе уже недостаточно штук.\n"
                        f"Доступно: {_display_bot_number(balance_qty)} шт."
                    ),
                }
            if is_area:
                balance_area = _number_value(row_value(row_values, columns.get("balance_area")))
                if _number_value(item.get("area")) > balance_area + INCOME_VOLUME_TOLERANCE:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать списание: на складе уже недостаточно площади.\n"
                            f"Доступно: {_display_bot_number(balance_area)} м2."
                        ),
                    }
            elif is_linear:
                balance_linear = _number_value(row_value(row_values, columns.get("balance_linear")))
                if _number_value(item.get("linear")) > balance_linear + INCOME_VOLUME_TOLERANCE:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать списание: на складе уже недостаточно погонных метров.\n"
                            f"Доступно: {_display_bot_number(balance_linear)} мп."
                        ),
                    }
            elif item.get("volume") is not None:
                balance_volume = _number_value(row_value(row_values, columns["balance_volume"]))
                if _number_value(item.get("volume")) > balance_volume + INCOME_VOLUME_TOLERANCE:
                    return {
                        "ok": False,
                        "message": (
                            "Не удалось записать списание: на складе уже недостаточно объема.\n"
                            f"Доступно: {_display_bot_number(balance_volume)} м3."
                        ),
                    }

            if operation_id is not None:
                execute_operation_write(store, operation_id, item, row_values, columns)
            else:
                add_to_row_value(row_values, columns["balance_qty"], -_number_value(item["quantity"]))
                if is_area:
                    add_to_row_value(row_values, columns.get("balance_area"), -_number_value(item["area"]))
                elif is_linear:
                    add_to_row_value(row_values, columns.get("balance_linear"), -_number_value(item["linear"]))
                elif item.get("volume") is not None:
                    add_to_row_value(row_values, columns["balance_volume"], -_number_value(item["volume"]))

            # Задача користувача: "додамо колонку причина списания в
            # складі... там буде відображатись весь текст причини" — пише
            # ПОВЕРХ (не accumulate), бо СКЛАД — агрегований рядок за
            # SKU, а не журнал операцій; це і є "поточна причина останнього
            # списання цього рядка", той самий принцип, що вже діє для
            # balance_qty (завжди поточне значення, не історія). set_value
            # безпечно ігнорує відсутню колонку (columns.get(...) is None)
            # — старі встановлення без цієї колонки не ламаються.
            if payload.get("comment"):
                set_value(row_values, columns.get("writeoff_reason"), payload["comment"])

        for row_id, row_values in row_values_by_row_id.items():
            store.update_row(row_id, row_values)

        # Той самий персистентний _next_document_number, що вже мають
        # продаж/антисептирование - один номер на ВЕСЬ виклик (усі позиції
        # цього списання), не по одному на кожен рядок.
        existing_writeoff_count = len(store.fetch_rows(WRITEOFF_SHEET_NAME, 100000, 0))
        writeoff_document_number = (
            f"Списание №{_next_document_number(store, WRITEOFF_SHEET_NAME, existing_writeoff_count)}"
        )
        for item in payload["rows"]:
            writeoff_sheet_row = writeoff_sheet_values(store, payload, item, now, writeoff_document_number)
            insert_sheet_row(store, WRITEOFF_SHEET_NAME, writeoff_sheet_row, now)

        user = payload.get("user") or {}
        for item in payload["rows"]:
            store.add_stock_movement(
                {
                    "movement_type": "writeoff",
                    "source": "telegram",
                    "telegram_user_id": user.get("id"),
                    "username": user.get("username"),
                    "full_name": user.get("full_name"),
                    "product": sheet_product_name(payload),
                    "breed": payload.get("breed"),
                    "condition": payload.get("condition"),
                    "thickness": item.get("thickness"),
                    "width": item.get("width"),
                    "length": item.get("length"),
                    "quantity": item.get("quantity"),
                    "volume": item.get("volume"),
                    "area": item.get("area"),
                    "linear": item.get("linear"),
                    "reason": payload.get("comment"),
                    "sheet_row_id": item.get("row_id"),
                    "original_text": payload.get("original_text"),
                    "created_at": now,
                }
            )
            updated += 1

    excel_warning = sync_excel_after_operation(sync_mode, store, ["СКЛАД", WRITEOFF_SHEET_NAME], dirty_notifier)

    # Задача користувача (2026-08-17): "Состояние має показуватись скрізь" -
    # той самий заголовок "Позиция: Продукт / Порода / Состояние", що вже
    # має apply_sale_operation/apply_income_operation. На відміну від них,
    # списання завжди має РІВНО один product/breed/condition на весь виклик
    # (payload-рівень, не по позиціях/рядках) - групувати нема чого,
    # заголовок друкується один раз.
    lines = ["Списание записано:"]
    header_parts = [
        _esc(part) for part in (display_product_name(payload), payload.get("breed"), payload.get("condition")) if part
    ]
    if header_parts:
        lines.append(f"Позиция: {' / '.join(header_parts)}")
    for index, item in enumerate(payload["rows"], start=1):
        measure_kind = item_measure_kind(item)
        row_values = row_values_by_row_id.get(item.get("row_id"))
        remaining_suffix = (
            f" (Осталось: {_esc(_remaining_balance_text(row_values, columns, measure_kind))})"
            if row_values is not None
            else ""
        )
        if measure_kind is None:
            lines.append(
                f"{index}. {_esc(income_item_size(item))}: -{_display_bot_number(item['quantity'])} шт{remaining_suffix}"
            )
            continue
        measure_value = item.get(measure_kind)
        measure_unit = ITEM_MEASURE_UNIT[measure_kind]
        lines.append(
            f"{index}. {_esc(income_item_size(item))}: -"
            f"{_display_bot_number(item['quantity'])} шт, -"
            f"{_display_bot_number(measure_value)} {measure_unit}{remaining_suffix}"
        )
    lines.append(f"Обновлено позиций: {updated}")
    if payload.get("comment"):
        lines.append(f"Причина: {_esc(payload['comment'])}")
    lines.append("")
    lines.append("✅ Выполнено.")
    if excel_warning:
        lines.append("")
        lines.append(_esc(excel_warning))
    return {"ok": True, "message": "\n".join(lines)}


# =============================================================================
# Услуга антисептирования — окремий лист (АНТИСЕПТИРОВАНИЕ), НЕ рядок у
# ПРОДАЖА МАТЕРИАЛА: інша структура колонок (Тип расчета/Статус оплаты/
# Приход наличных/Приход по банку/Отражение в расчетах — ведеться окремий
# готівка/банк розподіл, якого нема у звичайній продажі), і склад НЕ
# списується (доска клієнта, не наша). Зверху листа (перед реальними
# заголовками колонок) є інформаційний блок зведення (об'єм/вартість/
# кількість послуг, скільки готівкою/по банку) — НЕ формули, статичні
# значення в шаблоні, тому sync_antiseptic_to_excel перераховує їх сама при
# кожній синхронізації (на відміну від sync_sheets_to_excel, яка для
# СКЛАД/ПРОДАЖА просто дописує рядки без жодного зведення зверху).
# =============================================================================


def antiseptic_columns(headers):
    names = {
        "date": ["Дата"],
        "service_number": ["№ услуги"],
        "client": ["Клиент"],
        "address": ["Адрес выгрузки"],
        "service": ["Услуга"],
        "unit": ["Ед. изм."],
        "volume": ["Объем, м3"],
        "price_per_unit": ["Цена за м3, MDL"],
        "total_amount": ["Стоимость, MDL"],
        "payment_method": ["Тип расчета"],
        "payment_status": ["Статус оплаты"],
        "document": ["№ документа"],
        "cash_amount": ["Приход наличных, MDL"],
        "bank_amount": ["Приход по банку, MDL"],
        "reflection": ["Отражение в расчетах"],
        "manager": ["Ответственный"],
        "comment": ["Комментарий"],
    }
    normalized_headers = {
        _normalize_phrase(header): index
        for index, header in enumerate(headers)
        if header is not None
    }
    return {
        target: next(
            (
                normalized_headers[_normalize_phrase(candidate)]
                for candidate in candidates
                if _normalize_phrase(candidate) in normalized_headers
            ),
            None,
        )
        for target, candidates in names.items()
    }


def antiseptic_sheet_values(store, payload, now):
    headers = store.get_headers(ANTISEPTIC_SHEET_NAME)
    values = [""] * len(headers)
    columns = antiseptic_columns(headers)

    service_date = _parse_date_text(payload.get("date")) or date.today()
    set_value(values, columns.get("date"), service_date)

    # Персистентний _next_document_number (не len(fetch_rows(...))+1) -
    # імунний і до паралельних записів, і до видалення рядка через GUI
    # (аудит 28.07.2026, #4).
    existing_count = len(store.fetch_rows(ANTISEPTIC_SHEET_NAME, 100000, 0))
    set_value(values, columns.get("service_number"), f"Услуга №{_next_document_number(store, ANTISEPTIC_SHEET_NAME, existing_count)}")

    set_value(values, columns.get("client"), payload.get("client"))
    set_value(values, columns.get("address"), payload.get("address"))
    set_value(values, columns.get("service"), "Антисептирование")
    set_value(values, columns.get("unit"), "м3")
    volume = _number_value(payload.get("volume"))
    set_value(values, columns.get("volume"), volume)

    price_per_unit = payload.get("price_per_unit")
    total_amount = payload.get("total_amount")
    # Аудит коду: обидва напрямки рахувались без округлення — float-шум
    # (напр. 100.31233000000001) потрапляв просто в "Сумма"/"Цена за м3"
    # у фінансовому обліку. _priced_amount (utils.py) - спільна формула
    # прямого напрямку з sale_sheet_values/income_sheet_values вище;
    # зворотний напрямок (ціна ЗІ суми) — окрема формула, лишається тут.
    if total_amount in (None, "") and price_per_unit not in (None, ""):
        total_amount = _priced_amount(price_per_unit, volume)
    elif price_per_unit in (None, "") and total_amount not in (None, "") and volume:
        price_per_unit = round(_number_value(total_amount) / volume, 2)
    set_value(values, columns.get("price_per_unit"), price_per_unit)
    set_value(values, columns.get("total_amount"), total_amount)

    payment_method = normalize_payment_method(store, payload.get("payment_method"))
    set_value(values, columns.get("payment_method"), payment_method)
    # Часткова/відкладена оплата зараз ніде в системі не ведеться (те саме
    # припущення, що й у звичайній продажі) — записана послуга завжди вважається
    # одразу повністю оплаченою.
    set_value(values, columns.get("payment_status"), "Оплачено")
    set_value(values, columns.get("document"), payload.get("document_type") or "")

    # Аудит коду: раніше тут звіряли payment_method з жорстким рядком
    # "ЕФАКТУРА Б/Н" напряму — перейменування варіанту оплати (штатна дія в
    # "Способи оплати") тихо ламало розподіл готівка/банк без жодної помилки.
    # get_payment_method_kind — стабільний прапорець (payment_method_options.
    # kind), що переживає перейменування; NULL (жоден варіант не позначений
    # адміном як банк) означає "готівка", як і завжди.
    is_bank = store.get_payment_method_kind(payment_method) == "bank"
    total_number = _number_value(total_amount)
    set_value(values, columns.get("cash_amount"), 0 if is_bank else total_number)
    set_value(values, columns.get("bank_amount"), total_number if is_bank else 0)
    set_value(values, columns.get("reflection"), "БАНК" if is_bank else "НАЛИЧНЫЕ")

    user = payload.get("user") or {}
    manager = payload.get("manager") or user.get("full_name") or user.get("username")
    set_value(values, columns.get("manager"), manager)

    comment_parts = []
    if payload.get("original_text"):
        comment_parts.append(f"Telegram: {payload['original_text']}")
    if payload.get("comment"):
        comment_parts.append(payload["comment"])
    set_value(values, columns.get("comment"), " | ".join(comment_parts))
    return values


def apply_antiseptic_operation(store, payload, sync_mode, dirty_notifier=None):
    headers = store.get_headers(ANTISEPTIC_SHEET_NAME)
    if not headers:
        return {
            "ok": False,
            "message": f"Не удалось записать услугу: лист {ANTISEPTIC_SHEET_NAME} не найден.",
        }

    now = datetime.now().isoformat(timespec="seconds")
    columns = antiseptic_columns(headers)

    # Аудит коду: antiseptic_sheet_values не має жодного захисту від
    # від'ємного об'єму/ціни — від'ємний об'єм і так пройшов би повз
    # перевірку "volume == 0" (лише ділення на нуль), даючи від'ємну
    # ціну при зворотному розрахунку (сума ÷ об'єм) без жодного відхилення.
    # Послуга не має складського залишку (немає що "роздути"), але
    # від'ємний запис у фінансовому обліку так само безглуздий.
    # Свіжий пере-аудит (2026-08-02): об'єм послуги, на відміну від ціни/
    # суми нижче, не має "або/або"-альтернативи - завжди обов'язковий сам
    # по собі, тож нуль тут відхиляємо безумовно.
    if _number_value(payload.get("volume")) <= 0:
        return {
            "ok": False,
            "message": "Не удалось записать услугу: объём не может быть отрицательным или равным нулю.",
        }
    price_per_unit_value = _number_value(payload.get("price_per_unit"))
    total_amount_value = _number_value(payload.get("total_amount"))
    if price_per_unit_value < 0 or total_amount_value < 0:
        return {"ok": False, "message": "Не удалось записать услугу: цена и сумма не могут быть отрицательными."}
    # price_per_unit/total_amount - те саме "або/або" поле, що й у продажу
    # (has_price у telegram_dialog_antiseptic.py приймає будь-яке з двох) -
    # відхиляємо лише коли ОБИДВА не додатні.
    if price_per_unit_value <= 0 and total_amount_value <= 0:
        return {
            "ok": False,
            "message": "Не удалось записать услугу: должна быть указана цена за м3 или сумма.",
        }

    # Реальний баг з аудиту: той самий клас TOCTOU-гонки, що й виправлений
    # в apply_sale_operation — antiseptic_sheet_values рахує "Услуга №N" за
    # поточною кількістю рядків (existing_count), а раніше це відбувалось
    # ДО BEGIN узагалі (і навіть після — на голому "BEGIN", deferred). Два
    # паралельні антисептирования могли порахувати ОДНАКОВИЙ номер послуги.
    # Тепер підрахунок теж усередині BEGIN IMMEDIATE (RESERVED-блокування
    # одразу), як і в apply_sale_operation.
    with store.conn:
        store.conn.execute("BEGIN IMMEDIATE")
        values = antiseptic_sheet_values(store, payload, now)
        insert_sheet_row(store, ANTISEPTIC_SHEET_NAME, values, now)
        user = payload.get("user") or {}
        store.add_stock_movement(
            {
                "movement_type": "antiseptic",
                "source": "telegram",
                "telegram_user_id": user.get("id"),
                "username": user.get("username"),
                "full_name": user.get("full_name"),
                "product": "Антисептирование",
                "volume": payload.get("volume"),
                "original_text": payload.get("original_text"),
                "created_at": now,
            }
        )

    excel_warning = sync_antiseptic_after_operation(sync_mode, store, dirty_notifier)

    volume = _number_value(payload.get("volume"))
    total_amount = row_value(values, columns.get("total_amount"))
    payment_method = row_value(values, columns.get("payment_method"))
    lines = [
        "Услуга записана:",
        f"Антисептирование: {_display_bot_number(volume)} м3",
        f"Клиент: {_esc(payload.get('client'))}",
    ]
    address = payload.get("address")
    if address:
        lines.append(f"Адрес выгрузки: {_esc(address)}")
    if payment_method:
        lines.append(f"Оплата: {_esc(payment_method)}")
    # Задача користувача: "сумма має бути завжди знизу, скрізь" - Сумма
    # переїхала в самий кінець (після клієнта/адреси/оплати), той самий
    # порядок, що вже узгоджено для продажу й екрана підтвердження.
    lines.append(f"<b>Сумма: {_display_bot_number(total_amount)} MDL</b>")
    lines.append("")
    lines.append("✅ Выполнено.")
    if excel_warning:
        lines.append("")
        lines.append(_esc(excel_warning))
    return {"ok": True, "message": "\n".join(lines)}


# Зведення зверху листа (об'єм/вартість/кількість/готівка/банк) - НЕ формули
# в шаблоні, тож перераховуються тут щоразу заново з реальних рядків, а не
# лишаються "як було в Excel до цього".
def sync_antiseptic_to_excel(store):
    create_excel_backup()
    workbook = excel_source.open_workbook()
    try:
        worksheet = workbook[ANTISEPTIC_SHEET_NAME]
        header_row = _find_header_row(worksheet, ANTISEPTIC_SHEET_NAME, fallback=7)
        summary_row_1 = header_row - 4
        summary_row_2 = header_row - 3
        data_start_row = header_row + 1

        rows = store.fetch_all_rows(ANTISEPTIC_SHEET_NAME)
        columns = antiseptic_columns(store.get_headers(ANTISEPTIC_SHEET_NAME))

        # Свіжий пере-аудит (2026-08-02): volume_total випав із фіксу нижче -
        # той самий клас багу (сума вже округлених float все одно накопичує
        # похибку), просто для об'єму, а не для грошових підсумків.
        volume_total = round(sum(_number_value(row_value(row, columns.get("volume"))) for row in rows), 2)
        # Аудит коду (перевірка охоплення Fix #4): кожен рядок тут уже
        # округлений (antiseptic_sheet_values), але plain sum() над кількома
        # округленими float все одно накопичує похибку (напр. sum([19.99]*7)
        # == 139.92999999999998) - той самий клас багу, лише на рівні суми
        # в шапці Excel, а не окремого рядка.
        cost_total = round(sum(_number_value(row_value(row, columns.get("total_amount"))) for row in rows), 2)
        cash_total = round(sum(_number_value(row_value(row, columns.get("cash_amount"))) for row in rows), 2)
        bank_total = round(sum(_number_value(row_value(row, columns.get("bank_amount"))) for row in rows), 2)

        if summary_row_1 >= 1:
            worksheet.cell(row=summary_row_1, column=2, value=volume_total)
            worksheet.cell(row=summary_row_1, column=5, value=cost_total)
            worksheet.cell(row=summary_row_1, column=8, value=len(rows))
        if summary_row_2 >= 1:
            worksheet.cell(row=summary_row_2, column=2, value=cash_total)
            worksheet.cell(row=summary_row_2, column=5, value=bank_total)
            # Статус оплаты завжди "Оплачено" (див. antiseptic_sheet_values) -
            # неоплачених послуг зараз бути не може.
            worksheet.cell(row=summary_row_2, column=8, value=0)

        if worksheet.max_row >= data_start_row:
            worksheet.delete_rows(data_start_row, worksheet.max_row - data_start_row + 1)
        for row_offset, values in enumerate(rows):
            for col_offset, value in enumerate(values):
                worksheet.cell(
                    row=data_start_row + row_offset, column=col_offset + 1,
                    value=_sanitize_excel_value(value),
                )

        excel_source.save_workbook(workbook)
    finally:
        workbook.close()


# sync_antiseptic_to_excel записує ІНШИЙ лист/логіку (шапка+підсумкові
# рядки), ніж загальний sync_sheets_to_excel - позначаємо це окремим
# рядком-маркером у тому самому наборі "брудних" листів (TelegramBotWorker.
# mark_excel_dirty/_excel_sync_tick, main.py) замість окремого набору стану.
ANTISEPTIC_DIRTY_MARKER = "__antiseptic__"


def sync_antiseptic_after_operation(sync_mode, store, dirty_notifier=None):
    if sync_mode != "after_each_operation":
        return None
    if dirty_notifier is not None:
        dirty_notifier.mark_excel_dirty([ANTISEPTIC_DIRTY_MARKER])
        return None
    try:
        sync_antiseptic_to_excel(store)
    except PermissionError:
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Закройте файл Excel и нажмите «Обновити Excel»."
        )
    except OSError:
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Проверьте, что файл доступен, и нажмите «Обновити Excel»."
        )
    except RuntimeError:
        return (
            "Операция сохранена в программе, но Excel сейчас не обновлен.\n"
            "Проверьте настройку источника Excel-таблицы в программе."
        )
    return None
