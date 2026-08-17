"""Єдина точка входу для читання/запису Excel — незалежно від того, чи
таблиця локальна, чи на OneDrive/SharePoint (Задача користувача: "можна як
локальний так і онлайн. потрібно вибрати або або").

gui.py (читання: _load_excel_into_store/_discard_current_sheet_changes) і
warehouse_data.py (запис: create_excel_backup/sync_sheet(s)_to_excel)
раніше читали/писали paths.FILE_PATH напряму (і, окремо, gui.py's
self.file_path — уже сьогодні розбіжність з тим, що реально пише
warehouse_data.py). Обидва місця тепер ідуть через open_workbook/
save_workbook/backup_workbook_bytes нижче — єдине місце, де вирішується,
що таке "поточний Excel-файл".

Задача користувача (наступне уточнення): "зроби так щоб працював лише
вибраний ексель один... і більше ніяких інших файлів не читає, ні
внутрепроєктних, ніяких інших" — excel_local_path БІЛЬШЕ НЕ має мовчазного
запасного варіанту на paths.FILE_PATH: якщо файл не обрано явно (порожній
excel_local_path), _local_path кидає чітку помилку, а не тихо підставляє
внутрішній файл проєкту. Одноразова міграція (gui.py.__init__, поруч із
self.settings) записує ПОТОЧНИЙ файл (paths.FILE_PATH) як явний вибір у
settings.json РІВНО один раз, при першому запуску після цієї зміни — щоб
нічого не зламалось для вже працюючих встановлень, а обраний файл відтоді
й надалі був явним записом у JSON, а не кодовим запасним варіантом.

settings.json читається напряму (SettingsStore(SETTINGS_PATH)) в кожній
функції, а не передається параметром — так само, як сам settings.py читає
свій файл, і щоб не протягувати новий аргумент через усі виклики
sync_sheets_to_excel/_load_excel_into_store по всьому проєкту.
"""

import os
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

import onedrive_sync
from paths import SETTINGS_PATH
from settings import SettingsStore


def _settings():
    return SettingsStore(SETTINGS_PATH)


def _local_path(settings):
    custom = settings.get("excel_local_path")
    if not custom:
        raise RuntimeError(
            "Локальний файл Excel не обрано. Відкрийте «Таблиця Excel» у Налаштуваннях і оберіть файл."
        )
    return Path(custom)


def _require_online_access_token():
    token, _username = onedrive_sync.get_access_token_silent()
    if not token:
        raise RuntimeError(
            "Онлайн-джерело Excel не підключено або сеанс входу застарів. "
            "Відкрийте «Таблиця Excel» у Налаштуваннях і увійдіть через Microsoft ще раз."
        )
    return token


# Важлива знахідка нового аудиту (28.07.2026, #10): режим "Онлайн" можна
# зберегти через "Таблиця Excel" (save_source, gui.py), навіть якщо
# "Підключити файл" (посилання) так і не було успішно виконано — токен при
# цьому може бути (людина увійшла через Microsoft), але excel_online_
# drive_id/item_id лишаються порожніми. Без цієї перевірки
# onedrive_sync.download_workbook_bytes(token, "", "") пішов би з порожніми
# ID напряму до Graph API, отримавши незрозумілу низькорівневу HTTP-помилку
# замість чіткої дії-підказки.
def _require_online_selection(settings):
    drive_id = settings.get("excel_online_drive_id")
    item_id = settings.get("excel_online_item_id")
    if not drive_id or not item_id:
        raise RuntimeError(
            "Онлайн-файл ще не обрано. Відкрийте «Таблиця Excel» у Налаштуваннях і підключіть файл через посилання."
        )
    return drive_id, item_id


def open_workbook(data_only=False):
    settings = _settings()
    if settings.get("excel_source_mode") == "online":
        token = _require_online_access_token()
        drive_id, item_id = _require_online_selection(settings)
        data = onedrive_sync.download_workbook_bytes(token, drive_id, item_id)
        return load_workbook(BytesIO(data), data_only=data_only)
    local_path = _local_path(settings)
    if not local_path.exists():
        _create_blank_workbook(local_path)
    return load_workbook(local_path, data_only=data_only)


def _create_blank_workbook(path):
    # Пакування в .exe (перший запуск на "чистій" машині - там немає жодного
    # test_sklad.xlsx поруч, на відміну від dev-теки з проєктом, де такий
    # файл завжди вже лежав): обраний локальний шлях (у т.ч. з одноразової
    # міграції gui.py.__init__) ще не існує на диску взагалі - без цього
    # load_workbook() нижче падав би FileNotFoundError ще ДО того, як
    # ensure_workbook_has_required_sheets (warehouse_data.py) встигне додати
    # потрібні листи в порожній файл ("якщо приєднати порожній ексель, то
    # має створитись программою красивенька табличка").
    # openpyxl не дозволяє зберегти книгу зовсім без жодного видимого листа
    # ("At least one sheet must be visible") - лишаємо дефолтний "Sheet",
    # так само як і будь-який порожній .xlsx, створений вручну в Excel;
    # ensure_workbook_has_required_sheets (виклик одразу після цього) додасть
    # усі 5 обов'язкових листів поруч із ним.
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.save(path)


def save_workbook(workbook):
    settings = _settings()
    if settings.get("excel_source_mode") == "online":
        token = _require_online_access_token()
        drive_id, item_id = _require_online_selection(settings)
        buffer = BytesIO()
        workbook.save(buffer)
        onedrive_sync.upload_workbook_bytes(token, drive_id, item_id, buffer.getvalue())
        return
    # Реальний баг (аудит коду, 2026-08-14): раніше workbook.save() писав
    # НАПРЯМУ в кінцевий файл - цей запис відбувається після КОЖНОЇ операції
    # приходу/продажу/списання/антисептирування. Крах/зникнення живлення
    # саме посеред запису лишав би пошкоджений .xlsx (openpyxl такий файл
    # більше не відкриє), і жодне з наявних except-оточень навколо
    # синхронізації такого пошкодження не ловить - усі наступні операції
    # почали б падати. Той самий прийом, що вже рятує settings.json
    # (settings.py) - спершу пишемо в тимчасовий файл у ТІЙ САМІЙ теці (щоб
    # заміна нижче лишалась атомарною навіть на Windows), і лише коли запис
    # повністю вдався - замінюємо ним кінцевий файл. Крах під час запису
    # тепер лишає в найгіршому разі пошкоджений .tmp поруч, а справжній
    # Excel-файл - недоторканим.
    target_path = _local_path(settings)
    tmp_path = target_path.with_name(target_path.name + ".tmp")
    workbook.save(tmp_path)
    os.replace(tmp_path, target_path)


def backup_workbook_bytes():
    """Сирі байти поточного файлу — для create_excel_backup (локальна
    резервна копія робиться однаково в обох режимах, це додатковий
    запобіжник, а не заміна самого джерела)."""
    settings = _settings()
    if settings.get("excel_source_mode") == "online":
        token = _require_online_access_token()
        drive_id, item_id = _require_online_selection(settings)
        return onedrive_sync.download_workbook_bytes(token, drive_id, item_id)
    return _local_path(settings).read_bytes()


def backup_file_name_parts():
    settings = _settings()
    if settings.get("excel_source_mode") == "online":
        name = settings.get("excel_online_file_name") or "excel_online.xlsx"
        path = Path(name)
        return path.stem, path.suffix or ".xlsx"
    path = _local_path(settings)
    return path.stem, path.suffix


def current_source_identity(settings=None):
    """Стабільний рядок-ідентифікатор ПІДКЛЮЧЕНОГО файлу (не вмісту) —
    щоб відрізнити "той самий файл перепідключили" від "підключили інший
    файл". Задача користувача: "ніяких перенесень. всі таблиці і дані з
    таблиць ЛИШЕ ПЕРСОНАЛЬНІ і не мають ЖОДНІ дані бути пов'язаними між
    таблицями" — Store порівнює це значення з тим, що зберіг минулого
    разу (app_meta), і якщо воно змінилось, скидає похідний від вмісту
    попереднього файлу стан (лічильники номерів документів, підказки
    "останні використані", вивчені написання клієнтів), щоб він не
    протікав у щойно підключений, зовсім інший файл."""
    settings = settings or _settings()
    if settings.get("excel_source_mode") == "online":
        drive_id = settings.get("excel_online_drive_id") or ""
        item_id = settings.get("excel_online_item_id") or ""
        return f"online:{drive_id}:{item_id}"
    local_path = settings.get("excel_local_path") or ""
    return f"local:{local_path}"


def is_real_source_switch(new_identity, settings=None):
    """True, якщо new_identity відрізняється від того, що зараз активне
    в settings.json, І поточне джерело вже було колись реально обрано
    (не порожній placeholder "local:"/"online::") — щоб діалог вибору
    файлу не лякав попередженням про "зміну файлу" при найпершому
    підключенні, коли міняти ще нічого. Задача користувача (2026-08-14,
    одразу після [[feedback_file_scoped_data_isolation]]): "давай тепер
    зробимо коли новий підключаємо файл щоб питало підтвердження" —
    викликається З ДІАЛОГУ вибору файлу, ДО збереження нового джерела в
    settings.json (сам скид похідного стану відбувається пізніше, при
    реімпорті — Store._reset_file_scoped_state_if_source_changed)."""
    old_identity = current_source_identity(settings)
    if old_identity in ("local:", "online::"):
        return False
    return old_identity != new_identity


def current_source_label():
    """Рядок для показу в GUI (діалог "Таблиця Excel") — НЕ через
    _local_path(), щоб сам показ статусу не падав з помилкою, якщо файл
    ще не обрано (тоді показуємо про це прямим текстом, а не винятком)."""
    settings = _settings()
    if settings.get("excel_source_mode") == "online":
        file_name = settings.get("excel_online_file_name")
        if not file_name:
            return "Онлайн (OneDrive/SharePoint) — файл ще не підключено"
        account = settings.get("excel_online_account")
        return f"Підключено: {file_name}" + (f" ({account})" if account else "")
    local_path = settings.get("excel_local_path")
    if not local_path:
        return "Локальний файл ще не обрано"
    return f"Локально: {local_path}"
