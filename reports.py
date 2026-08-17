"""Узагальнений report-builder. Один опис звіту (title, columns, rows) ->
три рендери (повідомлення, Excel, PDF). Кожен майбутній звіт (продажі за
період, історія, повний звіт по складу — Фаза B) описує лише СВОЇ колонки
й дані та викликає ці рендери, а не пише свій PDF/Excel-генератор з нуля.

Опис звіту (report spec) — простий dict:
    {
        "title": "Остаток по складу",
        "generated_at": "10.07.2026 12:00",
        "columns": [
            {"key": "product", "label": "Продукт", "width_mm": 52,
             "align": "left", "total": False},
            {"key": "quantity", "label": "Кол-во, шт", "width_mm": 28,
             "align": "right", "total": True},
        ],
        "rows": [{"product": "Доска AD", "quantity": 12.0}, ...],
    }

"total": True на колонці означає "порахувати суму по всіх рядках і показати
підсумковий рядок/значення" — однаково для всіх трьох форматів.

Без залежностей від telegram_dialog.py (щоб не було циклічного імпорту) —
лише paths.py, stdlib і openpyxl.
"""

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from paths import BASE_DIR, BUNDLED_PYTHON_EXE, PDF_HELPER_EXE, PDF_REPORT_SCRIPT
from utils import _sanitize_excel_value

FORMAT_MESSAGE = "message"
FORMAT_EXCEL = "excel"
FORMAT_PDF = "pdf"


def _slug(value):
    text = str(value or "").strip().upper().replace(" ", "_")
    return text or "ВСЕ"


def build_report_filename(kind, scope=None, date_from=None, date_to=None, ext="xlsx"):
    """Авто-ім'я файлу звіту: ОСТАТОК_AD_09-07-2026.xlsx,
    ПРОДАЖИ_ВСЕ_01-07-2026_09-07-2026.xlsx. Дати приймає вже відформатованими
    рядками (наприклад datetime.strftime("%d-%m-%Y")) — ця функція не працює
    з датами сама, лише складає ім'я. kind/scope мають бути російськими
    рядками (без англійських слів) — застосунок мусить лишатись повністю
    російськомовним, назви файлів включно."""
    parts = [_slug(kind), _slug(scope)]
    if date_from and date_to and date_from != date_to:
        parts.append(f"{date_from}_{date_to}")
    elif date_from:
        parts.append(date_from)
    return "_".join(parts) + f".{ext}"


def _number_value(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return 0


def _display_number(value):
    if value in (None, ""):
        return "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    try:
        number = float(text.replace(",", "."))
        if number.is_integer():
            return str(int(number))
        return str(round(number, 4)).replace(".", ",")
    except ValueError:
        return text


def _display_cell(value, column):
    if column.get("total") or column.get("numeric"):
        return _display_number(value)
    return "" if value is None else str(value)


def column_totals(spec):
    rows = spec["rows"]
    return {
        column["key"]: sum(_number_value(row.get(column["key"])) for row in rows)
        for column in spec["columns"]
        if column.get("total")
    }


def _message_header(spec):
    lines = [spec["title"], f"Позиций: {len(spec['rows'])}"]
    if spec.get("generated_at"):
        lines.append(f"Сформировано: {spec['generated_at']}")
    return lines


def _message_totals_line(spec):
    totals = column_totals(spec)
    if not totals:
        return None
    columns_by_key = {column["key"]: column for column in spec["columns"]}
    parts = []
    for key, value in totals.items():
        unit = columns_by_key[key].get("unit", "")
        display = _display_number(round(value, 4))
        parts.append(f"{display}{' ' + unit if unit else ''}")
    return "Итого: " + ", ".join(parts)


def render_report_message(spec):
    """Формат повідомлення. Якщо spec["message_group_by"] задано (список
    ключів колонок, напр. ["product", "breed", "condition"]) — рядки
    групуються під одним заголовком замість повторення тих самих полів на
    кожному рядку (саме це робило довгий звіт нечитабельним: "Продукт: X |
    Порода: Y | ..." на кожній з 30 позицій). spec["message_compact"] — для
    звітів, де рядок = окрема операція (продажі за період тощо), тож
    групувати нема за чим, але повторювати "Поле: значення" на кожному
    рядку так само нечитабельно — значення просто йдуть підряд без міток
    (одиниці шт/м3/м2/MDL лишаються самопоясненими). Без жодного з двох —
    старий плаский формат "N. поле: значення | поле: значення"."""
    if spec.get("message_group_by"):
        return _render_grouped_message(spec, spec["message_group_by"])
    if spec.get("message_compact"):
        return _render_compact_flat_message(spec)
    return _render_flat_message(spec)


def _row_note(row, columns):
    # "note"-колонка — довільний текст для конкретного рядка (напр.
    # попередження про розбіжність даних). Показуємо ОКРЕМО, після значень,
    # а не змішуємо з рештою деталей рядка — і лише якщо не порожній.
    parts = [str(row.get(column["key"])) for column in columns if column.get("note") and row.get(column["key"])]
    return " ".join(parts)


def _append_footer(lines, spec):
    totals_line = _message_totals_line(spec)
    if totals_line:
        lines.append("")
        lines.append(totals_line)
    warnings = spec.get("warnings")
    if warnings:
        lines.append("")
        lines.extend(warnings)


def _render_flat_message(spec):
    columns = spec["columns"]
    rows = spec["rows"]
    message_columns = [
        column for column in columns if column.get("in_message", True) and not column.get("note")
    ]
    lines = _message_header(spec)
    lines.append("")
    # message_row_limit — для звітів, де рядків може бути багато (продажі за
    # період тощо) і групування (як у "Остатке") не рятує, бо кожен рядок —
    # окрема операція, а не дублікат тієї самої позиції. Обмежує лише
    # ПОВІДОМЛЕННЯ (щоб влізти у ~2 екрани) — PDF/Excel завжди показують усі
    # рядки без обмеження.
    limit = spec.get("message_row_limit")
    display_rows = rows[:limit] if limit else rows
    for index, row in enumerate(display_rows, start=1):
        # Пропускаємо колонку, якщо саме в ЦЬОМУ рядку значення відсутнє
        # (None/"") — інакше товар, що ведеться в м2 (площа), показував би
        # оманливе "Площадь, м2: 0" поряд із реальним об'ємом (і навпаки).
        parts = [
            f"{column['label']}: {_display_cell(row.get(column['key']), column)}"
            for column in message_columns
            if row.get(column["key"]) not in (None, "")
        ]
        line = f"{index}. " + " | ".join(parts)
        note = _row_note(row, columns)
        if note:
            line += f" {note}"
        lines.append(line)
    if limit and len(rows) > limit:
        lines.append(f"...и еще {len(rows) - limit} позиций.")

    _append_footer(lines, spec)
    return "\n".join(lines)


# Для звітів-журналів (кожен рядок — окрема операція: продажа, прихід тощо),
# де групувати нема за чим (на відміну від "Остатка", де багато рядків
# ділять один product/breed/condition), але повторювати "Поле: значення" на
# кожному рядку так само нечитабельно. Нетотal-колонки йдуть підряд через
# " — " без міток (дата/клієнт/товар самопояснені як текст), а total-колонки
# (кількість/обсяг/сума) — через кому з одиницею виміру (шт/м3/MDL), той
# самий принцип, що й у _render_grouped_message.
def _render_compact_flat_message(spec):
    columns = spec["columns"]
    rows = spec["rows"]
    detail_columns = [
        column
        for column in columns
        if column.get("in_message", True) and not column.get("total") and not column.get("note")
    ]
    value_columns = [column for column in columns if column.get("total")]

    lines = _message_header(spec)
    lines.append("")
    limit = spec.get("message_row_limit")
    display_rows = rows[:limit] if limit else rows
    for index, row in enumerate(display_rows, start=1):
        detail = " — ".join(
            _display_cell(row.get(column["key"]), column)
            for column in detail_columns
            if row.get(column["key"]) not in (None, "")
        )
        values = ", ".join(
            f"{_display_cell(row.get(column['key']), column)}{' ' + column['unit'] if column.get('unit') else ''}"
            for column in value_columns
            if row.get(column["key"]) not in (None, "")
        )
        line = f"{index}. {detail} — {values}" if detail else f"{index}. {values}"
        note = _row_note(row, columns)
        if note:
            line += f" {note}"
        lines.append(line)
    if limit and len(rows) > limit:
        lines.append(f"...и еще {len(rows) - limit} позиций.")

    _append_footer(lines, spec)
    return "\n".join(lines)


def _render_grouped_message(spec, group_by):
    columns = spec["columns"]
    # Копія, відсортована за group_by — інакше однакові товар/порода/стан,
    # розкидані не поспіль у вихідних рядках, дають ту саму групу двічі
    # (рендер лише зливає ПОСПІЛЬ однакові рядки, не шукає їх по всьому списку).
    rows = sorted(spec["rows"], key=lambda row: tuple(str(row.get(key) or "") for key in group_by))
    detail_columns = [
        column
        for column in columns
        if column.get("in_message", True) and column["key"] not in group_by
        and not column.get("total") and not column.get("note")
    ]
    value_columns = [column for column in columns if column.get("total")]

    lines = _message_header(spec)
    lines.append("")

    # Реальний баг з аудиту: на відміну від _render_flat_message/
    # _render_compact_flat_message, тут узагалі не було message_row_limit —
    # групування рятує лише коли позицій МАЛО ГРУП з багатьма рядками
    # всередині; широкий/непідфільтрований "Остаток" з десятками РІЗНИХ
    # товар/порода/розмір комбінацій групування не стискає взагалі (кожна
    # своя група), і повідомлення могло вийти дуже довгим. PDF/Excel і надалі
    # показують усі рядки без обмеження.
    limit = spec.get("message_row_limit")
    display_rows = rows[:limit] if limit else rows
    skipped_count = len(rows) - len(display_rows)

    current_group = object()  # сентінел — гарантовано не дорівнює жодному кортежу
    for row in display_rows:
        group_key = tuple(row.get(key) for key in group_by)
        if group_key != current_group:
            current_group = group_key
            group_label = ", ".join(
                str(row.get(key)) for key in group_by if row.get(key) not in (None, "")
            )
            lines.append(f"{group_label}:")

        detail = " ".join(
            _display_cell(row.get(column["key"]), column)
            for column in detail_columns
            if row.get(column["key"]) not in (None, "")
        )
        # Пропускаємо колонки-підсумки, де саме в ЦЬОМУ рядку значення
        # відсутнє (None) — інакше товар, що ведеться в м2 (площа), показував
        # би оманливе "0 м3" поряд із реальною площею.
        values = ", ".join(
            f"{_display_cell(row.get(column['key']), column)}{' ' + column['unit'] if column.get('unit') else ''}"
            for column in value_columns
            if row.get(column["key"]) not in (None, "")
        )
        line = f"  {detail} — {values}" if detail else f"  {values}"
        note = _row_note(row, columns)
        if note:
            line += f" {note}"
        lines.append(line)

    if skipped_count > 0:
        lines.append(f"...и еще {skipped_count} позиций.")

    _append_footer(lines, spec)
    return "\n".join(lines)


# _sanitize_excel_value/_FORMULA_TRIGGER_CHARS живуть у utils.py (критична
# знахідка аудиту 28.07.2026, #2) — щоб warehouse_data.py (головне
# Excel-дзеркало) теж могло їх перевикористати без циклічного імпорту
# reports.py<->warehouse_data.py.
def render_report_excel(spec, path):
    columns = spec["columns"]
    rows = spec["rows"]
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"

    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    sheet.cell(row=1, column=1, value=spec["title"]).font = Font(bold=True, size=14)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    subtitle = f"Сформировано: {spec.get('generated_at', '')} | Позиций: {len(rows)}"
    sheet.cell(row=2, column=1, value=subtitle)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))
    sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    header_row = 3
    for col_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=_sanitize_excel_value(column["label"]))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_offset, row in enumerate(rows, start=1):
        for col_index, column in enumerate(columns, start=1):
            value = row.get(column["key"])
            if value is None:
                # None означає "не застосовується до цього рядка" (напр.
                # обʼєм для товару, що ведеться в площі) — лишаємо порожню
                # клітинку, а не "0", щоб не виглядало як помилка даних.
                value = ""
            elif column.get("total") or column.get("numeric"):
                value = _number_value(value)
            else:
                value = _sanitize_excel_value(value)
            cell = sheet.cell(row=header_row + row_offset, column=col_index, value=value)
            if column.get("align") == "right":
                cell.alignment = Alignment(horizontal="right")
            elif column.get("align") == "center":
                cell.alignment = Alignment(horizontal="center")

    totals = column_totals(spec)
    if totals:
        total_row = header_row + len(rows) + 1
        label_written = False
        for col_index, column in enumerate(columns, start=1):
            if column["key"] in totals:
                cell = sheet.cell(row=total_row, column=col_index, value=round(totals[column["key"]], 4))
            elif not label_written:
                cell = sheet.cell(row=total_row, column=col_index, value="Итого")
                label_written = True
            else:
                cell = sheet.cell(row=total_row, column=col_index, value="")
            cell.font = bold_font

    for col_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = max(10, column.get("width_mm", 25) / 2.2)

    workbook.save(str(path))
    return path


def render_report_pdf(spec, path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    data_path = path.with_suffix(".json")
    data_path.write_text(
        json.dumps(
            {
                "title": spec["title"],
                "generated_at": spec.get("generated_at", ""),
                "columns": spec["columns"],
                "rows": spec["rows"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    try:
        # Пакування в .exe: у зібраній версії немає ні реального Python, ні
        # вихідного system/pdf_stock_report.py - PDF генерує окремий
        # допоміжний .exe (paths.PDF_HELPER_EXE). У dev-режимі (python
        # client_app.py/main.py) все лишається як було - той самий
        # BUNDLED_PYTHON_EXE + PDF_REPORT_SCRIPT.
        if getattr(sys, "frozen", False):
            if not PDF_HELPER_EXE.exists():
                raise RuntimeError("Не найден модуль генерации PDF.")
            argv = [str(PDF_HELPER_EXE), str(data_path), str(path)]
        else:
            if not BUNDLED_PYTHON_EXE.exists():
                raise RuntimeError("Не найден Python для генерации PDF.")
            if not PDF_REPORT_SCRIPT.exists():
                raise RuntimeError("Не найден скрипт генерации PDF.")
            argv = [str(BUNDLED_PYTHON_EXE), str(PDF_REPORT_SCRIPT), str(data_path), str(path)]

        try:
            result = subprocess.run(
                argv,
                cwd=str(BASE_DIR),
                capture_output=True,
                text=True,
                timeout=30,
                # Без цього кожен експорт PDF блимав би консольним вікном
                # поверх --windowed головного застосунку (pdf_stock_report.exe
                # свідомо зібраний як --console, щоб його можна було
                # запустити окремо для діагностики).
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
            )
        except subprocess.TimeoutExpired:
            # Реальний ризик (аудит коду, 2026-08-14): timeout=30 сам по
            # собі кидає subprocess.TimeoutExpired - інший клас винятку, ніж
            # усі решта помилок тут (навмисно RuntimeError із зрозумілим
            # текстом). Без цього перехоплення виклик все одно НЕ падає
            # непіймано (у _render_sales_report вже є широкий except
            # Exception), але користувач бачив би загальне "попробуйте еще
            # раз позже" - без жодного натяку, що саме тайм-аут, і що
            # повторна спроба з ТИМИ САМИМИ великими даними, ймовірно, знову
            # впреться в ту саму межу.
            raise RuntimeError(
                "Формирование PDF заняло слишком много времени (свыше 30 секунд). "
                "Попробуйте уменьшить отчет (сузить период или фильтр) или выбрать другой формат."
            )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "Не удалось сформировать PDF.")
        if not path.exists() or path.stat().st_size == 0:
            raise RuntimeError("PDF-файл не был создан.")
        return path
    finally:
        # Аудит коду: цей .json — лише вхідні дані для одноразового
        # підпроцесу, ніде далі не потрібен — раніше лишався на диску
        # назавжди після кожного згенерованого PDF, і при успіху, і при помилці.
        data_path.unlink(missing_ok=True)


def render_report(spec, fmt, path=None):
    """Єдина точка входу для викликача: обирає рендер за форматом."""
    if fmt == FORMAT_MESSAGE:
        return {"type": "message", "text": render_report_message(spec)}
    if fmt == FORMAT_EXCEL:
        result_path = render_report_excel(spec, path)
        return {"type": "document", "path": str(result_path)}
    if fmt == FORMAT_PDF:
        result_path = render_report_pdf(spec, path)
        return {"type": "document", "path": str(result_path)}
    raise ValueError(f"Неизвестный формат отчета: {fmt}")
